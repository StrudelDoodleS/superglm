"""Reporting and export for a ``Piecewise`` term.

The design's fourth property lives here: a rating table exported from a
piecewise term must reproduce ``model.predict`` from the workbook alone.  That
is the claim the feature exists to make, so the test reconstructs predictions
out of cell values and a parsed sheet note -- never out of the fitted spec --
and the matching sabotage is aimed at the workbook WRITE path.  Sabotaging the
basis would move both sides of the comparison and prove nothing.

Tolerances are derived rather than chosen.  ``_RECONSTRUCTION_RTOL`` counts the
roundings in the two paths being compared and is asserted to sit inside the
1e-12 the design asks for, so the design's own bar is pinned alongside the
tighter measured one.
"""

from __future__ import annotations

import re
from io import BytesIO

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from superglm import Categorical, Numeric, Spline, SuperGLM
from superglm.export.excel import write_rating_table_workbook
from superglm.export.rating_tables import build_rating_table_payload
from superglm.export.summary import _adapt_compact_summary, build_summary_export_payload
from superglm.inference._term_covariance import feature_se_from_cov
from tests._piecewise_cases import CASE_NAMES, make_case

_EPS = float(np.finfo(np.float64).eps)

# Roundings along the two paths the export test compares.  Reconstruction: a
# two-point interpolation (5), an exp (1), a power (3) and three products (3).
# Model: the hat weight (3), a two-term dot product (3), three additive term
# contributions (3) and an exp (1).  Twenty-two, plus the comparison's own two,
# rounded up to 32 for headroom on a differently-ordered BLAS.  Terms in the
# log-scale sum are of similar magnitude, so there is no cancellation to
# inflate this beyond a flop count.
_RECONSTRUCTION_OPS = 32
_RECONSTRUCTION_RTOL = _RECONSTRUCTION_OPS * _EPS

# The edf equality is a statement about a trace, which accumulates over the
# whole active design rather than over one dot product; 1e-9 is the design's
# own bar and is orders above the round-off, so it is used as written.
_EDF_ATOL = 1e-9

_NOTE_SLOPES = re.compile(r"below (\S+) use slope (\S+); above (\S+) use slope (\S+)\.")
_NOTE_FLAT = re.compile(r"below (\S+) use the (\S+) row; above (\S+) use the (\S+) row\.")

_TITLE_ROW = 5
_NOTE_ROW = 6
_HEADER_ROW = 7
_STRIDE = 3

_FITTED: dict[tuple[str, str | None], SuperGLM] = {}


def _fit(case_name: str, extrapolation: str | None = None) -> tuple[SuperGLM, object]:
    """Fit the named fixture once and reuse it; the tests only read from it.

    Every term is exactly tabulable on purpose -- intercept, Piecewise,
    Categorical, Numeric and no spline, whose block is binned by construction.
    A spline in the model would put discretisation error into the comparison
    that has nothing to do with the piecewise block.
    """
    key = (case_name, extrapolation)
    if key not in _FITTED:
        case = make_case(case_name, extrapolation=extrapolation)
        model = SuperGLM(
            features={
                "x": case.spec,
                "region": Categorical(base="first"),
                "density": Numeric(),
            },
        )
        model.fit(case.X, case.y, sample_weight=case.sample_weight)
        _FITTED[key] = model
    return _FITTED[key], make_case(case_name, extrapolation=extrapolation)


def _piecewise_spec(model: SuperGLM):
    """The FITTED spec.  ``SuperGLM`` deep-copies specs, so the caller's is unbuilt."""
    return model._specs["x"]


def _group(model: SuperGLM):
    return next(g for g in model._groups if g.feature_name == "x")


def _read_main_effect_blocks(ws) -> list[dict[str, object]]:
    """Read the main-effect blocks back out of a written sheet, cells only."""
    blocks: list[dict[str, object]] = []
    idx = 0
    while True:
        start_col = 1 + idx * _STRIDE
        title = ws.cell(row=_TITLE_ROW, column=start_col).value
        if title is None:
            return blocks
        headers = [
            ws.cell(row=_HEADER_ROW, column=start_col + offset).value for offset in range(_STRIDE)
        ]
        rows: list[list[object]] = []
        row = _HEADER_ROW + 1
        while ws.cell(row=row, column=start_col).value is not None:
            rows.append(
                [ws.cell(row=row, column=start_col + offset).value for offset in range(_STRIDE)]
            )
            row += 1
        blocks.append(
            {
                "title": title,
                "note": ws.cell(row=_NOTE_ROW, column=start_col).value,
                "headers": headers,
                "rows": rows,
                "start_col": start_col,
            }
        )
        idx += 1


def _workbook_sheet(payload) -> object:
    buffer = BytesIO()
    write_rating_table_workbook(
        payload,
        buffer,
        sheet_name="Rating Tables",
        summary_sheet_name="Model Summary",
        impact_sheet_name="Discretization Impact",
    )
    buffer.seek(0)
    return load_workbook(buffer, data_only=True)["Rating Tables"]


def _predict_from_workbook(ws, X: pd.DataFrame) -> np.ndarray:
    """Rebuild predictions from cell values and the sheet note alone.

    Nothing fitted is consulted: the knots, the log relativities, the two
    boundary slopes, the categorical lookup and the per-unit relativity all
    come out of the workbook, which is exactly what a consumer has.
    """
    base = float(ws["C2"].value)
    blocks = {str(block["title"]): block for block in _read_main_effect_blocks(ws)}

    piecewise = blocks["x"]
    assert piecewise["headers"] == ["x", "Relativity", "Log relativity"]
    knots = np.array([float(row[0]) for row in piecewise["rows"]], dtype=np.float64)
    log_relativity = np.array([float(row[2]) for row in piecewise["rows"]], dtype=np.float64)

    x = X["x"].to_numpy(dtype=np.float64)
    # ``np.interp`` clamps to the end values outside the grid, which is
    # already the flat rule; the slopes rule then overwrites the tails.  The
    # note is parsed, not the spec: whichever rule the sheet states is the
    # rule this consumer applies.
    log_effect = np.interp(x, knots, log_relativity)
    note = str(piecewise["note"])
    match = _NOTE_SLOPES.search(note)
    if match is not None:
        lower, slope_low, upper, slope_high = (float(value) for value in match.groups())
        assert lower == knots[0]
        assert upper == knots[-1]
        below = x < lower
        above = x > upper
        log_effect[below] = log_relativity[0] + slope_low * (x[below] - lower)
        log_effect[above] = log_relativity[-1] + slope_high * (x[above] - upper)
    else:
        flat = _NOTE_FLAT.search(note)
        assert flat is not None, note
        lower, low_row, upper, high_row = (float(value) for value in flat.groups())
        assert lower == knots[0] == low_row
        assert upper == knots[-1] == high_row

    lookup = {str(row[0]): float(row[1]) for row in blocks["region"]["rows"]}
    region = np.array([lookup[str(value)] for value in X["region"]], dtype=np.float64)

    per_unit = float(blocks["density"]["rows"][0][1])
    density = per_unit ** X["density"].to_numpy(dtype=np.float64)

    return base * np.exp(log_effect) * region * density


class TestTermInference:
    @pytest.mark.parametrize("case_name", CASE_NAMES)
    def test_term_inference_reports_one_value_per_knot_with_a_zero_at_the_base(self, case_name):
        """kind, the knot vector, and the base knot's zero -- the whole contract."""
        model, _ = _fit(case_name)
        spec = _piecewise_spec(model)
        ti = model.term_inference("x")

        assert ti.kind == "piecewise"
        assert ti.absorbs_intercept is False
        assert ti.centering_mode == "base_knot"
        np.testing.assert_array_equal(ti.x, spec._knots)
        assert ti.log_relativity.shape == spec._knots.shape
        assert ti.log_relativity[spec._base_index] == 0.0
        assert ti.relativity[spec._base_index] == 1.0
        assert ti.se_log_relativity[spec._base_index] == 0.0
        assert np.all(np.isfinite(ti.se_log_relativity))
        assert np.count_nonzero(ti.se_log_relativity) == spec._non_base_indices.size

    @pytest.mark.parametrize("case_name", CASE_NAMES)
    def test_to_dataframe_returns_the_x_bearing_shape(self, case_name):
        """A knot-indexed term whose frame lost x would say nothing about position."""
        model, _ = _fit(case_name)
        spec = _piecewise_spec(model)
        frame = model.term_inference("x").to_dataframe()

        assert list(frame.columns)[:3] == ["x", "log_relativity", "relativity"]
        assert "level" not in frame.columns
        assert "label" not in frame.columns
        assert len(frame) == spec._knots.size
        np.testing.assert_array_equal(frame["x"].to_numpy(), spec._knots)

    def test_an_inactive_term_returns_one_standard_error_per_knot(self):
        """The dropped-term early return must stay the length of log_relativity.

        Reached directly: the term-inference path gates on ``active`` before it
        asks for standard errors, so the only way to exercise the fallback is
        to call it with no active subgroup, which is exactly what a caller
        holding a covariance for a reduced design does.
        """
        model, _ = _fit("interior_base")
        spec = _piecewise_spec(model)
        se = feature_se_from_cov(
            "x",
            np.zeros((0, 0)),
            [],
            model.result,
            model._groups,
            model._specs,
            model._interaction_specs,
        )
        assert se.shape == (spec._knots.size,)
        np.testing.assert_array_equal(se, np.zeros(spec._knots.size))


class TestSummaryRows:
    @pytest.mark.parametrize("case_name", CASE_NAMES)
    def test_one_coefficient_row_per_non_base_knot_plus_one_chi_square_group_row(self, case_name):
        """The fallback branch would have emitted ONE row and dropped J of them."""
        model, _ = _fit(case_name)
        spec = _piecewise_spec(model)
        n_coefficients = spec._non_base_indices.size

        payload = build_summary_export_payload(model)
        term_rows = [row for row in payload.terms if row.group == "x"]
        coefficient_rows = [row for row in term_rows if row.kind == "coefficient"]
        group_rows = [row for row in term_rows if row.kind == "group"]

        assert len(coefficient_rows) == n_coefficients
        assert [row.term for row in coefficient_rows] == [
            f"x[{float(spec._knots[j]):.10g}]" for j in spec._non_base_indices
        ]
        for row in coefficient_rows:
            assert row.statistic_type == "z"
            assert row.p_value is not None and np.isfinite(row.p_value)
            assert row.ci_lower is not None and np.isfinite(row.ci_lower)
            assert row.ci_upper is not None and np.isfinite(row.ci_upper)
            assert row.ci_lower < row.estimate < row.ci_upper

        assert len(group_rows) == 1
        assert group_rows[0].term == "x"
        assert group_rows[0].statistic_type == "chi2"
        assert group_rows[0].p_value is not None and np.isfinite(group_rows[0].p_value)

        # The term's edf is reported once, on the term-level row.  Repeating it
        # on the first knot row -- the Categorical convention, which exists
        # because a categorical term has no term-level row -- would put the same
        # degrees of freedom in the summary's parametric and smooth buckets.
        carrying_edf = [row.term for row in term_rows if row.edf is not None]
        assert carrying_edf == ["x"]

        coef_rows = _adapt_compact_summary(model.summary(detail="compact")).rows
        whole_term = next(row for row in coef_rows if row.name == "x" and row.is_spline)
        assert whole_term.ref_df == float(n_coefficients)
        assert whole_term.n_params == n_coefficients

    def test_the_console_summary_names_the_term_piecewise_not_spline(self):
        """The group-test row is rendered by the spline path; it must not say so."""
        model, _ = _fit("interior_base")
        text = str(model.summary())
        assert "[piecewise, 4 params" in text
        assert "[spline, 4 params" not in text

    @pytest.mark.parametrize("case_name", CASE_NAMES)
    def test_the_summary_names_the_base_knot_in_a_reference_row(self, case_name):
        """Every coefficient is a contrast against a knot the summary must name.

        The workbook prints all ``J + 2`` knots and the editor puts a handle on
        each, so without this row the summary is the one surface of the three
        that never states the reference -- and with ``breaks=int`` the knot
        vector is not printed either, so it cannot be inferred by elimination.
        """
        model, _ = _fit(case_name)
        spec = _piecewise_spec(model)
        base_row = f"x[{float(spec._knots[spec._base_index]):.10g}]"

        text = str(model.summary())
        assert base_row in text
        # Rendered as a reference, not as an estimate of zero.
        line = next(line for line in text.splitlines() if base_row in line)
        assert "ref" in line

        # Knot order, not "the non-base knots then the base".
        rendered = [
            row.name
            for row in model.summary()._display_rows
            if row.group == "x" and row.name.startswith("x[")
        ]
        assert rendered == [f"x[{float(knot):.10g}]" for knot in spec._knots]

    def test_the_reference_row_stays_out_of_the_canonical_coefficient_rows(self):
        """It carries no coefficient, so nothing that counts parameters may see it.

        ``Categorical`` synthesises its reference row on the same display-only
        path for the same reason; putting it in ``_coef_rows`` would add a
        phantom row to the exported Summary sheet and to every edf bucket.
        """
        model, _ = _fit("interior_base")
        spec = _piecewise_spec(model)
        base_row = f"x[{float(spec._knots[spec._base_index]):.10g}]"

        assert all(row.name != base_row for row in model.summary()._coef_rows)
        payload = build_summary_export_payload(model)
        assert all(row.term != base_row for row in payload.terms)


class TestDegreesOfFreedomAttribution:
    def test_the_fixed_df_is_not_reported_as_smooth_df(self):
        """§4: edf is J+1 exactly, not an estimated trace, and the test is not Wood's.

        ``is_spline`` is only what routes the row through the group-test
        renderer.  Read as "this term was smoothed" it put four parametric df in
        the header's smooth bucket and printed the Wood footnote over a model
        with no smooth in it.
        """
        model, _ = _fit("interior_base")
        text = str(model.summary())

        assert "(0 smooth)" in text
        assert "Wood (2013)" not in text
        assert "Parametric p-values are Wald approximations." in text
        # Total df is unaffected: intercept + 4 knots + 2 region + 1 density.
        assert "8.000 (0 smooth)" in text

    def test_a_real_smooth_in_the_same_model_still_reports_as_smooth(self):
        """The exclusion is scoped to the piecewise row, not to the bucket itself."""
        case = make_case("interior_base")
        model = SuperGLM(
            features={
                "x": case.spec,
                "density": Spline(n_knots=6),
                "region": Categorical(base="first"),
            },
        )
        model.fit(case.X, case.y, sample_weight=case.sample_weight)
        text = str(model.summary())

        assert "(0 smooth)" not in text
        assert "Wood (2013)" in text


class TestRelativities:
    @pytest.mark.parametrize("case_name", CASE_NAMES)
    def test_relativities_carries_the_piecewise_term(self, case_name):
        """`relativities()` dispatches on an if/elif chain with no else branch.

        A spec whose ``reconstruct()`` names none of ``x`` / ``levels`` /
        ``relativity_per_unit`` is dropped from the documented "relativity
        DataFrames for all features" silently -- a tariff built off it ships one
        factor short, with no error and no warning.
        """
        model, _ = _fit(case_name)
        spec = _piecewise_spec(model)
        tables = model.relativities()

        assert list(tables) == list(model._feature_order)
        frame = tables["x"]
        np.testing.assert_allclose(frame["x"].to_numpy(), spec._knots, rtol=0.0, atol=0.0)
        np.testing.assert_allclose(
            frame["log_relativity"].to_numpy(),
            model.term_inference("x").log_relativity,
            rtol=0.0,
            atol=_RECONSTRUCTION_RTOL,
        )
        assert float(frame["relativity"].to_numpy()[spec._base_index]) == 1.0

    def test_relativities_with_se_reports_one_standard_error_per_knot(self):
        model, _ = _fit("interior_base")
        spec = _piecewise_spec(model)
        frame = model.relativities(with_se=True)["x"]

        assert len(frame) == spec._knots.size
        se = frame["se_log_relativity"].to_numpy()
        assert se[spec._base_index] == 0.0
        assert np.all(se[[j for j in range(se.size) if j != spec._base_index]] > 0.0)

    @pytest.mark.parametrize("case_name", CASE_NAMES)
    def test_feature_se_is_labelled_by_knot_like_every_other_piecewise_surface(self, case_name):
        """The unlabelled ``{"se": diag}`` fallback is off by one from the base on.

        It returns the J+1 per-coefficient values with no knot vector beside
        them, so a caller zipping them against the term's knots pairs the knot
        after the base with the base's SE, and so on to the end.
        """
        model, case = _fit(case_name)
        spec = _piecewise_spec(model)
        metrics = model.metrics(case.X, case.y, sample_weight=case.sample_weight)
        out = metrics.feature_se("x")

        assert set(out) == {"x", "se_log_relativity"}
        np.testing.assert_array_equal(out["x"], spec._knots)
        se = out["se_log_relativity"]
        assert se.shape == (spec._knots.size,)
        assert se[spec._base_index] == 0.0

    def test_the_four_condition_comment_is_present_at_the_reporting_site(self):
        """The reporting contract's conditions live in code, not only in a design doc."""
        from pathlib import Path

        import superglm.inference.coef_tables as coef_tables

        source = Path(coef_tables.__file__).read_text()
        branch = source.split("elif isinstance(spec, Piecewise):", 1)[1]
        comment = branch.split("knots = spec._knots", 1)[0]
        assert "unpenalized" in comment
        assert "selection shrinkage" in comment or "selection_penalty" in comment
        assert "FIXED INPUTS" in comment
        assert "unconstrained" in comment


class TestEffectiveDegreesOfFreedom:
    @pytest.mark.parametrize("case_name", CASE_NAMES)
    def test_the_measured_edf_equals_j_plus_one_on_an_unshrunk_fit(self, case_name):
        """The design ASSERTS edf = J+1; this makes it an observation instead.

        The value reported is the measured trace, not the nominal count, so
        this test is what turns the design's claim into evidence -- and the
        shrunk fit below is why reporting the nominal count would be a lie.
        """
        model, _ = _fit(case_name)
        spec = _piecewise_spec(model)
        nominal = float(spec._non_base_indices.size)

        assert model.term_inference("x").edf == pytest.approx(nominal, abs=_EDF_ATOL)

    def test_a_shrunk_fit_reports_less_than_j_plus_one(self):
        """Hardcoding J+1 would have silently over-reported this fit's df."""
        case = make_case("interior_base")
        model = SuperGLM(
            selection_penalty="auto",
            features={
                "x": case.spec,
                "region": Categorical(base="first"),
                "density": Numeric(),
            },
        )
        model.fit(case.X, case.y, sample_weight=case.sample_weight)
        spec = _piecewise_spec(model)
        nominal = float(spec._non_base_indices.size)
        edf = model.term_inference("x").edf

        assert 0.0 < edf < nominal - _EDF_ATOL

    def test_a_selection_penalty_auto_fit_completes_and_reports_rows(self):
        """A smoke test only: the four-condition comment withdraws the Wald numbers here."""
        case = make_case("interior_base")
        model = SuperGLM(
            selection_penalty="auto",
            features={
                "x": case.spec,
                "region": Categorical(base="first"),
                "density": Numeric(),
            },
        )
        model.fit(case.X, case.y, sample_weight=case.sample_weight)
        spec = _piecewise_spec(model)

        payload = build_summary_export_payload(model)
        term_rows = [row for row in payload.terms if row.group == "x"]
        assert sum(row.kind == "coefficient" for row in term_rows) == spec._non_base_indices.size
        assert sum(row.kind == "group" for row in term_rows) == 1
        assert str(model.summary())


class TestDerivedSlopes:
    @pytest.mark.parametrize("case_name", CASE_NAMES)
    def test_reconstruct_slopes_are_the_knot_differences_over_the_widths(self, case_name):
        """The Emblem reading, derived from the coefficients rather than fitted."""
        model, _ = _fit(case_name)
        spec = _piecewise_spec(model)
        group = _group(model)
        beta = np.asarray(model.result.beta[group.sl], dtype=np.float64)
        raw = spec.reconstruct(beta)

        values = np.zeros(spec._knots.size, dtype=np.float64)
        values[spec._non_base_indices] = beta
        assert values[spec._base_index] == 0.0
        expected = np.array(
            [
                (values[j + 1] - values[j]) / (spec._knots[j + 1] - spec._knots[j])
                for j in range(spec._knots.size - 1)
            ]
        )

        np.testing.assert_allclose(raw["slopes"], expected, rtol=0.0, atol=0.0)
        assert raw["boundary_slopes"] == (float(expected[0]), float(expected[-1]))


class TestDiscreteAndImpact:
    def test_a_discrete_fit_predicts_identically_to_a_dense_one(self):
        """``should_discretize`` is False for a non-spline; this pins that it stays so."""
        dense_case = make_case("interior_base")
        dense = SuperGLM(
            discrete=False,
            features={
                "x": dense_case.spec,
                "region": Categorical(base="first"),
                "density": Numeric(),
            },
        )
        dense.fit(dense_case.X, dense_case.y, sample_weight=dense_case.sample_weight)

        discrete_case = make_case("interior_base")
        discrete = SuperGLM(
            discrete=True,
            features={
                "x": discrete_case.spec,
                "region": Categorical(base="first"),
                "density": Numeric(),
            },
        )
        discrete.fit(discrete_case.X, discrete_case.y, sample_weight=discrete_case.sample_weight)

        np.testing.assert_array_equal(
            _piecewise_spec(discrete)._knots,
            _piecewise_spec(dense)._knots,
        )
        np.testing.assert_array_equal(
            discrete.result.beta[_group(discrete).sl],
            dense.result.beta[_group(dense).sl],
        )
        np.testing.assert_array_equal(
            discrete.term_inference("x").log_relativity,
            dense.term_inference("x").log_relativity,
        )
        np.testing.assert_array_equal(
            discrete.predict(discrete_case.X),
            dense.predict(dense_case.X),
        )

    def test_the_discretization_impact_sheet_has_no_row_for_the_piecewise_term(self):
        """A piecewise export has no discretisation error, so it has nothing to report.

        The model carries a spline as well, so the sheet is genuinely populated
        -- an empty sheet would satisfy the assertion for the wrong reason.
        """
        case = make_case("interior_base")
        X = case.X.copy()
        X["age"] = np.random.default_rng(77).uniform(18.0, 80.0, len(X))
        model = SuperGLM(
            features={
                "x": case.spec,
                "age": Spline(n_knots=6),
                "region": Categorical(base="first"),
            },
        )
        model.fit(X, case.y, sample_weight=case.sample_weight)
        payload = build_rating_table_payload(
            model, X, case.y, sample_weight=case.sample_weight, n_bins=20
        )

        assert [(block.name, block.kind) for block in payload.main_effects] == [
            ("x", "piecewise"),
            ("age", "continuous"),
            ("region", "categorical"),
        ]
        features = set(payload.discretization_impact["feature"])
        assert features == {"age"}


class TestWorkbookExactness:
    # ``pinned_narrower`` is not a duplicate of ``interior_base``: its rated
    # range is inside the data, so a third of its rows sit beyond the boundary
    # knots and the reconstruction has to apply the note's stated tail rule --
    # the parsed boundary slopes under "extend", the end-row clamp under
    # "clip" -- to reach them at all.  Both modes are swept so neither rule
    # can silently stand in for the other.
    @pytest.mark.parametrize("extrapolation", ["clip", "extend"])
    @pytest.mark.parametrize("case_name", ["interior_base", "pinned_narrower"])
    def test_the_workbook_alone_reproduces_the_predictions(self, case_name, extrapolation):
        assert _RECONSTRUCTION_RTOL <= 1e-12

        model, case = _fit(case_name, extrapolation=extrapolation)
        payload = build_rating_table_payload(
            model, case.X, case.y, sample_weight=case.sample_weight, n_bins=20
        )
        ws = _workbook_sheet(payload)

        reconstructed = _predict_from_workbook(ws, case.X)
        np.testing.assert_allclose(
            reconstructed,
            model.predict(case.X),
            rtol=_RECONSTRUCTION_RTOL,
            atol=0.0,
        )

    def test_the_pinned_narrower_case_really_does_extrapolate(self):
        """Guards the test above from passing because no row left the range."""
        model, case = _fit("pinned_narrower")
        spec = _piecewise_spec(model)
        x = case.X["x"].to_numpy(dtype=np.float64)

        assert np.count_nonzero(x < spec._knots[0]) > 50
        assert np.count_nonzero(x > spec._knots[-1]) > 50

    def test_the_block_is_three_columns_of_knots_relativities_and_logs(self):
        """Four columns would overwrite the neighbouring block on the fixed stride."""
        model, case = _fit("interior_base")
        payload = build_rating_table_payload(
            model, case.X, case.y, sample_weight=case.sample_weight, n_bins=20
        )
        block = next(b for b in payload.main_effects if b.name == "x")
        spec = _piecewise_spec(model)

        assert block.kind == "piecewise"
        assert list(block.table.columns) == ["x", "Relativity", "Log relativity"]
        assert len(block.table) == spec._knots.size
        np.testing.assert_array_equal(block.table["x"].to_numpy(), spec._knots)
        np.testing.assert_allclose(
            block.table["Relativity"].to_numpy(),
            np.exp(block.table["Log relativity"].to_numpy()),
            rtol=0.0,
            atol=0.0,
        )

    def test_the_extend_sheet_note_states_the_rule_and_both_boundary_slopes(self):
        """A consumer outside the tabulated range needs the rule, not just the rows."""
        model, case = _fit("pinned_narrower", extrapolation="extend")
        payload = build_rating_table_payload(
            model, case.X, case.y, sample_weight=case.sample_weight, n_bins=20
        )
        block = next(b for b in payload.main_effects if b.name == "x")
        assert block.extrapolation == "extend"
        ws = _workbook_sheet(payload)
        note = str(ws.cell(row=_NOTE_ROW, column=1).value)

        assert "Log relativity" in note
        assert "geometrically on Relativity" in note

        knots = block.table["x"].to_numpy(dtype=np.float64)
        log_relativity = block.table["Log relativity"].to_numpy(dtype=np.float64)
        slopes = np.diff(log_relativity) / np.diff(knots)
        match = _NOTE_SLOPES.search(note)
        assert match is not None, note
        lower, slope_low, upper, slope_high = (float(value) for value in match.groups())

        # Round-trip exact, not merely close: a boundary slope printed to a
        # readable number of digits reintroduces the discrepancy between model
        # and tariff that this feature exists to remove.
        assert lower == knots[0]
        assert upper == knots[-1]
        assert slope_low == slopes[0]
        assert slope_high == slopes[-1]

    def test_the_default_sheet_note_states_the_flat_rule_and_no_slopes(self):
        """Clip is the default, so the default workbook must state the flat rule."""
        model, case = _fit("pinned_narrower")
        payload = build_rating_table_payload(
            model, case.X, case.y, sample_weight=case.sample_weight, n_bins=20
        )
        block = next(b for b in payload.main_effects if b.name == "x")
        assert block.extrapolation == "clip"
        ws = _workbook_sheet(payload)
        note = str(ws.cell(row=_NOTE_ROW, column=1).value)

        assert "Log relativity" in note
        assert "geometrically on Relativity" in note
        assert _NOTE_SLOPES.search(note) is None
        flat = _NOTE_FLAT.search(note)
        assert flat is not None, note
        knots = block.table["x"].to_numpy(dtype=np.float64)
        lower, low_row, upper, high_row = (float(value) for value in flat.groups())
        assert lower == knots[0] == low_row
        assert upper == knots[-1] == high_row

    def test_the_error_sheet_note_states_that_out_of_range_is_not_rated(self):
        """interior_base, because an error-mode pinned_narrower refuses to fit at all."""
        model, case = _fit("interior_base", extrapolation="error")
        payload = build_rating_table_payload(
            model, case.X, case.y, sample_weight=case.sample_weight, n_bins=20
        )
        block = next(b for b in payload.main_effects if b.name == "x")
        assert block.extrapolation == "error"
        ws = _workbook_sheet(payload)
        note = str(ws.cell(row=_NOTE_ROW, column=1).value)

        assert "not rated" in note
        assert "extrapolation='error'" in note
        assert _NOTE_SLOPES.search(note) is None
        assert _NOTE_FLAT.search(note) is None

    def test_the_log_relativity_column_is_not_left_at_two_decimal_places(self):
        """Stored exactly but rendered as 0.00 defeats the column's whole purpose."""
        model, case = _fit("interior_base")
        payload = build_rating_table_payload(
            model, case.X, case.y, sample_weight=case.sample_weight, n_bins=20
        )
        spec = _piecewise_spec(model)
        ws = _workbook_sheet(payload)

        for row in range(_HEADER_ROW + 1, _HEADER_ROW + 1 + spec._knots.size):
            assert ws.cell(row=row, column=2).number_format == "0.000000000000"
            assert ws.cell(row=row, column=3).number_format == "0.000000000000"
