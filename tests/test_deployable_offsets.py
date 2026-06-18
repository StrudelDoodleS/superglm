from __future__ import annotations

import numpy as np
import pandas as pd
import pytest
from openpyxl import load_workbook

from superglm import LogRatioOffset, Numeric, SuperGLM


def _fit_model_with_term_offset():
    rng = np.random.default_rng(20260618)
    n = 240
    X = pd.DataFrame(
        {
            "score": rng.normal(size=n),
            "term": rng.choice([12.0, 36.0], size=n),
        }
    )
    offset = LogRatioOffset(
        name="policy_term",
        source_feature="term",
        reference=12.0,
    )
    eta = -1.2 + 0.15 * X["score"].to_numpy() + offset.evaluate(X)
    y = rng.poisson(np.exp(eta)).astype(float)
    exposure = rng.uniform(0.5, 2.0, size=n)
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={"score": Numeric()},
    )
    model.fit(X, y, sample_weight=exposure, offset=offset)
    return model, X, y, exposure, offset


def test_log_ratio_offset_is_reused_for_prediction():
    model, X, _y, _exposure, offset = _fit_model_with_term_offset()

    automatic = model.predict(X)
    explicit = model.predict(X, offset=offset.evaluate(X))

    np.testing.assert_allclose(automatic, explicit, rtol=1e-12, atol=1e-12)


def test_log_ratio_offset_is_reported_separately_from_fitted_coefficients():
    model, _X, _y, _exposure, _offset = _fit_model_with_term_offset()

    summary = model.summary()
    text = str(summary)

    assert "Fixed offsets (not estimated):" in text
    assert "policy_term" in text
    assert "log(term / 12)" in text
    assert "response multiplier = (term / 12) ^ 1" in text
    assert "fixed_offsets" in summary
    assert summary["fixed_offsets"][0]["Coefficient"] == 1.0


def test_structured_offset_replaces_generic_offset_multiplier_block():
    model, X, y, exposure, _offset = _fit_model_with_term_offset()

    payload = model.rating_table_payload(X, y, sample_weight=exposure)

    assert "Offset Multiplier" not in [block.name for block in payload.main_effects]


def test_log_ratio_offset_is_exported_to_dedicated_sheet(tmp_path):
    model, X, y, exposure, _offset = _fit_model_with_term_offset()
    output = tmp_path / "rating_tables.xlsx"

    model.export_rating_tables(output, X, y, sample_weight=exposure)

    workbook = load_workbook(output, data_only=True)
    assert workbook.sheetnames == [
        "Rating Tables",
        "Fixed Offsets",
        "Discretization Impact",
        "Model Summary",
    ]
    worksheet = workbook["Fixed Offsets"]
    headers = [worksheet.cell(row=1, column=i).value for i in range(1, 10)]
    assert headers == [
        "Term",
        "Term Type",
        "Source Feature",
        "Transform",
        "Reference Value",
        "Coefficient",
        "Sequence",
        "Link Expression",
        "Response Multiplier",
    ]
    assert worksheet["A2"].value == "policy_term"
    assert worksheet["B2"].value == "FIXED_OFFSET"
    assert worksheet["C2"].value == "term"
    assert worksheet["D2"].value == "LOG_RATIO"
    assert worksheet["E2"].value == 12.0
    assert worksheet["F2"].value == 1.0


def test_raw_array_offset_keeps_generic_multiplier_without_claiming_expression(tmp_path):
    rng = np.random.default_rng(9)
    X = pd.DataFrame(
        {
            "score": rng.normal(size=80),
            "term": rng.choice([12.0, 36.0], 80),
        }
    )
    offset = np.log(X["term"].to_numpy() / 12.0)
    y = rng.poisson(np.exp(-1.0 + 0.1 * X["score"].to_numpy() + offset)).astype(float)
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={"score": Numeric()},
    ).fit(X, y, offset=offset)
    output = tmp_path / "raw_offset.xlsx"

    payload = model.rating_table_payload(X, y)
    model.export_rating_tables(output, X, y)

    assert "Offset Multiplier" in [block.name for block in payload.main_effects]
    assert "Fixed Offsets" not in load_workbook(output, read_only=True).sheetnames


def test_log_ratio_offset_rejects_non_positive_source_values():
    offset = LogRatioOffset("policy_term", "term", reference=12.0)
    with pytest.raises(ValueError, match="strictly positive"):
        offset.evaluate(pd.DataFrame({"term": [12.0, 0.0]}))


def test_log_ratio_offset_requires_log_link():
    X = pd.DataFrame({"score": [0.0, 1.0, 2.0], "term": [12.0, 12.0, 36.0]})
    offset = LogRatioOffset("policy_term", "term", reference=12.0)
    model = SuperGLM(
        family="gaussian",
        link="identity",
        selection_penalty=0.0,
        features={"score": Numeric()},
    )

    with pytest.raises(ValueError, match="requires a log-link model"):
        model.fit(X, np.array([1.0, 2.0, 3.0]), offset=offset)
