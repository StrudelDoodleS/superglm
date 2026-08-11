from __future__ import annotations

import json
import re
import threading
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pandas as pd
import polars as pl
import pytest

from superglm import (
    Categorical,
    Numeric,
    OrderedCategorical,
    Polynomial,
    Spline,
    SuperGLM,
    families,
    generate_tweedie_cpg,
)
from superglm.editor import EditableTerm, EditorSession
from superglm.editor.apply import apply_edits_to_model_copy_with_data
from superglm.editor.evaluation import coerce_dataset
from superglm.editor.persistence import (
    ModelArtifactValidation,
    _spread_row_indices,
    serialize_validated_model,
)
from superglm.inference.summary import ModelSummary, _BasisDetailRow, _CoefRow


def test_editor_demo_notebook_includes_k_adequacy_sweep():
    notebook_path = Path(__file__).resolve().parents[1] / "docs/notebooks/editor_demo.ipynb"
    notebook = json.loads(notebook_path.read_text())
    source = "\n".join("".join(cell["source"]) for cell in notebook["cells"])

    assert "## Choose Spline Capacity With A k Sweep" in source
    assert "n = 150_000" in source
    assert "## In-Sample k Adequacy" in source
    assert "## Cross-Validated k Check" in source
    assert "from sklearn.model_selection import KFold" in source
    assert "from sklearn.model_selection import KFold, train_test_split" in source
    assert "families" in source
    assert "generate_tweedie_cpg" in source
    assert "X_train_val, X_test" in source
    assert "X_train, X_val" in source
    assert "TRUE_TWEEDIE_P" in source
    assert "TRUE_TWEEDIE_P_INIT" in source
    assert "TRUE_TWEEDIE_PHI" in source
    assert "age = np.clip(rng.normal(" in source
    assert "age_effect -= np.average(age_effect, weights=exposure)" in source
    assert "mu = np.exp(eta)" in source
    assert "y = generate_tweedie_cpg(" in source
    assert "TERRITORY_LEVELS" in source
    assert "territory_effect_map" in source
    assert "territory_probs" in source
    assert "territory = rng.choice(TERRITORY_LEVELS" in source
    assert '"T01"' in source
    assert '"T10"' in source
    assert "territory_effect" in source
    assert "family=families.tweedie(p=TRUE_TWEEDIE_P_INIT)" in source
    assert "discrete=True" in source
    assert "n_bins=512" in source
    assert "model.estimate_p(" in source
    assert "tweedie_profile.search_trace" in source
    assert 'method="brent"' in source
    assert 'phi_method="mle"' in source
    assert '"territory": Categorical(base="most_exposed")' in source
    assert 'terms=["age", "mileage", "region", "age_band", "territory"]' in source
    assert "## Collapse Sparse Categorical Levels" in source
    assert 'collapse_session.select_levels("territory", ["T01", "T03"])' in source
    assert 'collapse_session.select_levels("territory", ["T08", "T10"])' in source
    assert "grouped_preview = session_payload(collapse_session)" in source
    assert 'collapse_session.replace_with_collapsed_levels("territory", method="fit")' in source
    assert "cross_validate(" in source
    assert "cv = KFold(" in source
    assert "cv=cv" in source
    assert "k_grid" in source
    assert "cv_sweep" in source
    assert "cv_report = {" in source
    assert '"method": "KFold"' in source
    assert "mean_val_deviance" in source
    assert "recommend_k_from_cv" in source
    assert "cv_one_se_k" in source
    assert "cv_recommended_k" in source
    assert 'basis=Spline(kind="ps", k=5)' in source
    assert "median_age_band_p" in source
    assert "median_age_band_min_p" not in source
    assert "np.nanmin(p_values)" not in source
    assert "row.name == term" in source
    assert "chosen_k = cv_recommended_k" in source
    assert "age_edf" in source
    assert "total_edf" in source
    assert "edf_share" in source
    assert "chosen_k" in source
    assert 'Spline(kind="bs", k=age_k' in source
    assert "model = fit_age_model(chosen_k, X_fit=X_train, y_fit=y_train, w_fit=w_train)" in source
    assert "cv_loss_records = []" in source
    assert "fold_train_loss" in source
    assert "fold_validation_loss" in source
    assert "cv_loss_summary" in source
    assert "split_loss" in source
    assert '"metric": "weighted mean unit deviance"' in source
    assert '"rows": cv_loss.round(6).to_dict("records")' in source
    assert '"summary": cv_loss_summary.round(6).to_dict("records")' in source
    assert '"split_loss": split_loss.round(6).to_dict("records")' in source
    assert "train_data=(X_train, y_train, w_train)" in source
    assert "validation_data=(X_val, y_val, w_val)" in source
    assert "test_data=(X_test, y_test, w_test)" in source
    assert "cv_report=cv_report" in source


@pytest.fixture
def editor_frame():
    rng = np.random.default_rng(20260703)
    n = 450
    x_spline = rng.uniform(0.0, 10.0, n)
    x_poly = rng.uniform(-2.0, 3.0, n)
    x_num = rng.normal(0.0, 1.0, n)
    region = rng.choice(["A", "B", "C"], n, p=[0.4, 0.35, 0.25])
    band = rng.choice(["low", "medium", "high"], n, p=[0.45, 0.35, 0.20])

    eta = (
        0.6
        + 0.18 * np.sin(x_spline / 1.8)
        + 0.08 * x_poly**2
        + 0.12 * x_num
        + 0.25 * (region == "B")
        - 0.18 * (region == "C")
        + 0.15 * (band == "medium")
        + 0.32 * (band == "high")
    )
    y = eta + rng.normal(0.0, 0.08, n)
    X = pd.DataFrame(
        {
            "x_spline": x_spline,
            "x_poly": x_poly,
            "x_num": x_num,
            "region": region,
            "band": band,
        }
    )
    return X, y


def _to_polars(frame: pd.DataFrame) -> pl.DataFrame:
    return pl.DataFrame({name: frame[name].to_numpy() for name in frame.columns})


@pytest.fixture
def editor_model(editor_frame):
    X, y = editor_frame
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        spline_penalty=0.1,
        features={
            "x_spline": Spline(n_knots=8),
            "x_poly": Polynomial(degree=2),
            "x_num": Numeric(),
            "region": Categorical(base="first"),
            "band": OrderedCategorical(
                order=["low", "medium", "high"], basis=Spline(kind="ps", n_knots=2), base="first"
            ),
        },
    )
    model.fit(X, y)
    return model


def test_polars_editor_loading_edits_and_artifact_validation_match_pandas(
    editor_model,
    editor_frame,
):
    from superglm.editor.metrics import compute_dataset_metrics
    from superglm.editor.payloads import session_payload

    X, y = editor_frame
    X_polars = _to_polars(X)
    polars_model = editor_model.clone_unfitted().fit(X_polars, y)
    pandas_session = EditorSession.from_model(
        editor_model,
        terms=["x_spline", "region"],
        train_data=(X, y),
    )
    polars_session = EditorSession.from_model(
        polars_model,
        terms=["x_spline", "region"],
        train_data=(X_polars, y),
    )

    assert polars_session._evaluation_data["train"].X is X_polars
    assert session_payload(polars_session) == session_payload(pandas_session)
    for name in ("x_spline", "region"):
        np.testing.assert_allclose(
            polars_session.terms[name].weights,
            pandas_session.terms[name].weights,
            rtol=0.0,
            atol=0.0,
        )

    for session in (pandas_session, polars_session):
        session.select_indices("x_spline", [2, 3])
        session.shift("x_spline", 0.04)
        session.undo()
        session.redo()
        session.select_levels("region", ["B"])
        session.shift("region", 0.03)
    assert session_payload(polars_session) == session_payload(pandas_session)
    np.testing.assert_allclose(
        polars_session.edited_offset(["x_spline", "region"]),
        pandas_session.edited_offset(["x_spline", "region"]),
        rtol=0.0,
        atol=0.0,
    )

    pandas_edited = pandas_session.to_model()
    polars_edited = polars_session.to_model()
    assert polars_edited._fit_X_ref is X_polars
    np.testing.assert_allclose(
        polars_edited.predict(X_polars),
        pandas_edited.predict(X),
        rtol=1e-12,
        atol=1e-12,
    )

    pandas_dataset = coerce_dataset("validation", (X, y))
    polars_dataset = coerce_dataset("validation", (X_polars, y))
    assert polars_dataset is not None
    assert polars_dataset.X is X_polars
    assert polars_dataset.n_obs == len(X_polars)
    assert compute_dataset_metrics(polars_edited, polars_dataset) == pytest.approx(
        compute_dataset_metrics(pandas_edited, pandas_dataset)
    )

    data, validation = serialize_validated_model(
        polars_edited,
        dataset=polars_dataset,
        max_rows=19,
    )
    assert data
    assert validation.prediction_rows == 19


def test_polars_editor_category_collapse_matches_pandas_and_retains_native_frame(
    editor_model,
    editor_frame,
):
    from superglm.editor.payloads import session_payload

    X, y = editor_frame
    X_polars = _to_polars(X)
    polars_model = editor_model.clone_unfitted().fit(X_polars, y)
    pandas_session = EditorSession.from_model(
        editor_model,
        terms=["region"],
        train_data=(X, y),
    )
    polars_session = EditorSession.from_model(
        polars_model,
        terms=["region"],
        train_data=(X_polars, y),
    )

    for session in (pandas_session, polars_session):
        session.select_levels("region", ["B", "C"])
        session.replace_with_collapsed_levels("region", method="fit")

    assert polars_session.model._fit_X_ref is X_polars
    assert session_payload(polars_session) == session_payload(pandas_session)
    np.testing.assert_allclose(
        polars_session.model.predict(X_polars),
        pandas_session.model.predict(X),
        rtol=1e-12,
        atol=1e-12,
    )


def test_serialize_validated_model_round_trip_predictions(editor_model, editor_frame):
    import io

    import joblib

    X, y = editor_frame
    dataset = coerce_dataset("validation", (X, y))

    data, validation = serialize_validated_model(editor_model, dataset=dataset, max_rows=17)
    loaded = joblib.load(io.BytesIO(data))

    assert isinstance(data, bytes)
    assert validation == ModelArtifactValidation(
        artifact_round_trip=True,
        prediction_rows=17,
        scope="artifact+predictions",
    )
    indices = _spread_row_indices(len(X), 17)
    np.testing.assert_allclose(
        loaded.predict(X.iloc[indices]),
        editor_model.predict(X.iloc[indices]),
        rtol=1e-12,
        atol=1e-12,
    )


def test_spread_row_indices_are_deterministic_bounded_and_include_endpoints():
    first = _spread_row_indices(101, 17)
    second = _spread_row_indices(101, 17)

    np.testing.assert_array_equal(first, second)
    assert len(first) == 17
    assert first[0] == 0
    assert first[-1] == 100
    assert len(np.unique(first)) == len(first)
    np.testing.assert_array_equal(_spread_row_indices(4, 17), np.arange(4))
    np.testing.assert_array_equal(_spread_row_indices(0, 17), np.array([], dtype=np.intp))


def test_serialize_validated_model_slices_spread_rows_and_offset_consistently(
    editor_model,
    editor_frame,
    monkeypatch,
):
    import superglm.editor.persistence as persistence

    X, y = editor_frame
    offset = np.linspace(-0.2, 0.2, len(X))
    dataset = coerce_dataset("validation", (X, y, None, offset))
    expected_indices = _spread_row_indices(len(X), 17)
    seen: dict[str, np.ndarray] = {}
    real_load = persistence.joblib_load_bytes

    def load_with_recording_predict(data):
        loaded = real_load(data)
        real_predict = loaded.predict

        def recording_predict(X_rows, offset=None):
            seen["indices"] = X_rows.index.to_numpy()
            seen["offset"] = np.asarray(offset)
            return real_predict(X_rows, offset=offset)

        loaded.predict = recording_predict
        return loaded

    monkeypatch.setattr(persistence, "joblib_load_bytes", load_with_recording_predict)

    _, validation = serialize_validated_model(editor_model, dataset=dataset, max_rows=17)

    np.testing.assert_array_equal(seen["indices"], expected_indices)
    np.testing.assert_array_equal(seen["offset"], offset[expected_indices])
    assert validation.prediction_rows == 17


@pytest.mark.parametrize("failure", ["mismatch", "nonfinite"])
def test_serialize_validated_model_rejects_bad_loaded_predictions(
    editor_model,
    editor_frame,
    monkeypatch,
    failure,
):
    import superglm.editor.persistence as persistence

    X, y = editor_frame
    dataset = coerce_dataset("validation", (X, y))
    real_load = persistence.joblib_load_bytes

    def load_with_bad_predict(data):
        loaded = real_load(data)
        real_predict = loaded.predict

        def bad_predict(X_rows, offset=None):
            predictions = np.asarray(real_predict(X_rows, offset=offset), dtype=np.float64)
            if failure == "nonfinite":
                return np.full_like(predictions, np.nan)
            return predictions + 1.0

        loaded.predict = bad_predict
        return loaded

    monkeypatch.setattr(persistence, "joblib_load_bytes", load_with_bad_predict)

    with pytest.raises(ValueError, match="prediction validation failed"):
        serialize_validated_model(editor_model, dataset=dataset, max_rows=17)


def test_serialize_validated_model_contract_only(editor_model):
    data, validation = serialize_validated_model(editor_model)

    assert isinstance(data, bytes)
    assert validation == ModelArtifactValidation(
        artifact_round_trip=True,
        prediction_rows=0,
        scope="artifact",
    )


def test_serialize_validated_model_empty_dataset_is_contract_only(editor_model, editor_frame):
    X, y = editor_frame
    dataset = coerce_dataset("validation", (X.iloc[:0], y[:0]))

    _, validation = serialize_validated_model(editor_model, dataset=dataset)

    assert validation.scope == "artifact"
    assert validation.prediction_rows == 0


def test_serialize_validated_model_rejects_invalid_max_rows(editor_model):
    with pytest.raises(ValueError, match="max_rows"):
        serialize_validated_model(editor_model, max_rows=0)


@pytest.mark.parametrize(
    "contract_failure",
    [
        "wrong_type",
        "missing_result",
        "noncallable_predict",
        "beta_shape",
        "nonfinite_beta",
        "nonfinite_intercept",
        "feature_order",
        "spec_key_order",
        "spec_type",
    ],
)
def test_serialize_validated_model_rejects_bad_artifact_contract(
    editor_model,
    monkeypatch,
    contract_failure,
):
    from dataclasses import replace

    import superglm.editor.persistence as persistence

    real_load = persistence.joblib_load_bytes

    def load_invalid_artifact(data):
        if contract_failure == "wrong_type":
            return SimpleNamespace()
        loaded = real_load(data)
        if contract_failure == "missing_result":
            loaded._result = None
        elif contract_failure == "noncallable_predict":
            loaded.predict = None
        elif contract_failure == "beta_shape":
            loaded._result = replace(loaded.result, beta=loaded.result.beta[:-1])
        elif contract_failure == "nonfinite_beta":
            beta = loaded.result.beta.copy()
            beta[0] = np.nan
            loaded._result = replace(loaded.result, beta=beta)
        elif contract_failure == "nonfinite_intercept":
            loaded._result = replace(loaded.result, intercept=np.inf)
        elif contract_failure == "feature_order":
            loaded._feature_order = list(reversed(loaded._feature_order))
        elif contract_failure == "spec_key_order":
            loaded._specs = dict(reversed(list(loaded._specs.items())))
        elif contract_failure == "spec_type":
            first_name = next(iter(loaded._specs))
            loaded._specs[first_name] = Numeric()
        return loaded

    monkeypatch.setattr(persistence, "joblib_load_bytes", load_invalid_artifact)

    with pytest.raises(ValueError, match="artifact validation failed"):
        serialize_validated_model(editor_model)


def test_session_extracts_1d_main_effects(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline", "region"])

    assert list(session.terms) == ["x_spline", "region"]
    assert session.terms["x_spline"].x is not None
    assert session.terms["region"].levels == ["A", "B", "C"]
    assert session.terms["x_spline"].metadata["term_type"] == "spline"
    assert session.terms["region"].metadata["term_type"] == "categorical"


def test_select_x_and_levels(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline", "region"])

    session.select_x("x_spline", 2.0, 5.0)
    assert session.selection("x_spline").size > 0

    session.select_levels("region", ["B", "C"])
    assert session.selection("region").tolist() == [1, 2]


def test_editor_model_revision_changes_only_for_prediction_mutations(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline", "region"])
    assert session.model_revision == 0
    assert session.edit_epoch == 0

    session.select_indices("x_spline", [3, 4])
    assert session.model_revision == 0

    session.shift("x_spline", 0.1)
    assert session.model_revision == 1
    assert session.edit_epoch == 1

    session.undo()
    assert session.model_revision == 2
    assert session.edit_epoch == 2

    session.redo()
    assert session.model_revision == 3
    assert session.edit_epoch == 3

    session.select_indices("region", [1])
    session.reorder_levels("region", target_index=0)
    assert session.model_revision == 3


def test_editor_model_revision_noops_reset_and_cache_invalidation(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    assert session._materialized_edit_model is None
    assert session._materialized_edit_epoch is None

    with pytest.raises(KeyError):
        session.shift("missing", 0.1)
    assert session.model_revision == session.edit_epoch == 0

    session.select_indices("x_spline", [3])
    session.shift("x_spline", 0.1)
    session.undo()
    assert session.model_revision == session.edit_epoch == 2
    assert len(session.redo_stack) == 1

    cached_model = object()
    session._materialized_edit_model = cached_model
    session._materialized_edit_epoch = session.edit_epoch
    session.set_value("x_spline", float(term.edited_log_effect[3]))

    assert session.model_revision == session.edit_epoch == 2
    assert session._materialized_edit_model is cached_model
    assert session._materialized_edit_epoch == 2
    assert session.history[-1].operation == "set_value"
    assert session.redo_stack == []

    session.shift("x_spline", 0.1)
    assert session.model_revision == session.edit_epoch == 3
    assert session._materialized_edit_model is None
    assert session._materialized_edit_epoch is None

    session._materialized_edit_model = object()
    session._materialized_edit_epoch = session.edit_epoch
    session.reset("x_spline")
    assert session.model_revision == session.edit_epoch == 4
    assert session._materialized_edit_model is None
    assert session._materialized_edit_epoch is None
    assert term.edited_log_effect[3] == pytest.approx(term.original_log_effect[3])

    cached_model = object()
    session._materialized_edit_model = cached_model
    session._materialized_edit_epoch = session.edit_epoch
    session.reset("x_spline")
    assert session.model_revision == session.edit_epoch == 4
    assert session._materialized_edit_model is cached_model
    assert session._materialized_edit_epoch == 4


def test_set_values_collapses_same_value_duplicates_without_revision_change(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    value = float(term.edited_log_effect[3])
    cached_model = object()
    session._materialized_edit_model = cached_model
    session._materialized_edit_epoch = session.edit_epoch

    session.set_values("x_spline", [3, 3], [value, value])

    assert session.model_revision == session.edit_epoch == 0
    assert session._materialized_edit_model is cached_model
    assert session._materialized_edit_epoch == 0
    np.testing.assert_array_equal(session.history[-1].indices, [3])
    np.testing.assert_array_equal(session.history[-1].before, [value])
    np.testing.assert_array_equal(session.history[-1].after, [value])


def test_set_values_rejects_conflicting_duplicates_without_mutation(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_indices("x_spline", [4])
    session.shift("x_spline", 0.1)
    before = session.terms["x_spline"].edited_log_effect.copy()
    old_history = session.history
    old_records = list(session.history)
    old_revision = session.model_revision
    old_epoch = session.edit_epoch
    cached_model = object()
    session._materialized_edit_model = cached_model
    session._materialized_edit_epoch = session.edit_epoch

    with pytest.raises(ValueError, match="Conflicting values"):
        session.set_values("x_spline", [3, 3], [0.1, 0.2])

    np.testing.assert_array_equal(session.terms["x_spline"].edited_log_effect, before)
    assert session.history is old_history
    assert len(session.history) == len(old_records)
    assert all(actual is expected for actual, expected in zip(session.history, old_records))
    assert session.model_revision == old_revision
    assert session.edit_epoch == old_epoch
    assert session._materialized_edit_model is cached_model
    assert session._materialized_edit_epoch == old_epoch


def test_replace_in_force_model_advances_revision_once(editor_model):
    replacement = editor_model._clone_without_features(set())
    replacement.fit(editor_model._fit_X_ref, editor_model._fit_y_ref)
    session = EditorSession.from_model(editor_model, terms=["x_spline"])

    session.replace_in_force_model(replacement)

    assert session.model is replacement
    assert session.model_revision == 1
    assert session.edit_epoch == 1


def test_replace_in_force_model_failure_restores_complete_session_state(
    editor_model,
    monkeypatch,
):
    replacement = editor_model._clone_without_features(set())
    replacement.fit(editor_model._fit_X_ref, editor_model._fit_y_ref)
    session = EditorSession.from_model(editor_model, terms=["x_spline", "region"])
    session.select_levels("region", ["B"])
    session.shift("region", 0.1)
    session.select_indices("x_spline", [3])
    session.shift("x_spline", 0.2)
    session.undo("x_spline")
    session.reorder_levels("region", target_index=0)

    cached_model = object()
    session._materialized_edit_model = cached_model
    session._materialized_edit_epoch = session.edit_epoch

    old_model = session.model
    old_terms = session.terms
    old_term_values = {name: term.edited_log_effect.copy() for name, term in session.terms.items()}
    old_term_levels = {
        name: None if term.levels is None else list(term.levels)
        for name, term in session.terms.items()
    }
    old_selection = session._selection
    old_selection_state = {
        name: (indices, indices.copy()) for name, indices in session._selection.items()
    }
    old_history = session.history
    old_redo = session.redo_stack
    old_history_state = [
        (record, record.indices, record.indices.copy()) for record in session.history
    ]
    old_redo_state = [
        (record, record.indices, record.indices.copy()) for record in session.redo_stack
    ]
    old_level_orders = session._level_orders
    old_level_order_state = {
        name: (labels, list(labels)) for name, labels in session._level_orders.items()
    }
    old_revision = session.model_revision
    old_epoch = session.edit_epoch

    original_reapply = session._reapply_level_orders
    staged: dict[str, bool] = {}

    def fail_after_staged_reorder():
        staged["model"] = session.model is replacement
        staged["terms"] = session.terms is not old_terms
        staged["selection"] = session._selection is not old_selection
        staged["history"] = session.history is not old_history
        staged["redo"] = session.redo_stack is not old_redo
        staged["level_orders"] = session._level_orders is not old_level_orders
        original_reapply()
        raise RuntimeError("staged level reorder failed")

    monkeypatch.setattr(session, "_reapply_level_orders", fail_after_staged_reorder)

    with pytest.raises(RuntimeError, match="staged level reorder failed"):
        session.replace_in_force_model(replacement)

    assert all(staged.values())
    assert session.model is old_model
    assert session.terms is old_terms
    for name, term in session.terms.items():
        np.testing.assert_array_equal(term.edited_log_effect, old_term_values[name])
        assert term.levels == old_term_levels[name]
    assert session._selection is old_selection
    for name, (indices, values) in old_selection_state.items():
        assert session._selection[name] is indices
        np.testing.assert_array_equal(session._selection[name], values)
    assert session.history is old_history
    assert session.redo_stack is old_redo
    for stack, expected in (
        (session.history, old_history_state),
        (session.redo_stack, old_redo_state),
    ):
        assert len(stack) == len(expected)
        for record, (old_record, old_indices, old_values) in zip(stack, expected, strict=True):
            assert record is old_record
            assert record.indices is old_indices
            np.testing.assert_array_equal(record.indices, old_values)
    assert session._level_orders is old_level_orders
    for name, (labels, values) in old_level_order_state.items():
        assert session._level_orders[name] is labels
        assert session._level_orders[name] == values
    assert session.model_revision == old_revision
    assert session.edit_epoch == old_epoch
    assert session._materialized_edit_model is cached_model
    assert session._materialized_edit_epoch == old_epoch


def test_shift_set_interpolate_isotonic_smooth_and_history(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    session.select_indices("x_spline", [10, 11, 12, 13])

    before = term.edited_log_effect.copy()
    session.shift("x_spline", 0.25)
    np.testing.assert_allclose(term.edited_log_effect[10:14], before[10:14] + 0.25)

    session.set_value("x_spline", -0.1)
    np.testing.assert_allclose(term.edited_log_effect[10:14], -0.1)

    term.edited_log_effect[10:14] = [0.4, -0.2, 0.3, -0.1]
    session.linear_interpolate("x_spline")
    np.testing.assert_allclose(term.edited_log_effect[[10, 13]], [0.4, -0.1])
    assert np.all(np.diff(term.edited_log_effect[10:14]) < 0)

    term.edited_log_effect[10:14] = [0.3, -0.2, 0.2, -0.1]
    session.isotonic("x_spline", direction="increasing")
    assert np.all(np.diff(term.edited_log_effect[10:14]) >= -1e-12)

    session.undo()
    np.testing.assert_allclose(term.edited_log_effect[10:14], [0.3, -0.2, 0.2, -0.1])

    session.redo()
    assert np.all(np.diff(term.edited_log_effect[10:14]) >= -1e-12)

    term.edited_log_effect[10:14] = [0.0, 1.0, -0.5, 2.0]
    session.linear_interpolate("x_spline", strength=0.5)
    full_target = np.interp(
        term.x[10:14],
        [term.x[10], term.x[13]],
        [0.0, 2.0],
    )
    expected_blend = 0.5 * np.array([0.0, 1.0, -0.5, 2.0]) + 0.5 * full_target
    np.testing.assert_allclose(term.edited_log_effect[10:14], expected_blend)

    term.edited_log_effect[10:14] = [0.0, 1.0, 0.0, 1.0]
    session.smooth("x_spline", strength=1.0)
    assert np.max(term.edited_log_effect[10:14]) < 1.0


def test_level_and_snap_selected_values(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    session.select_indices("x_spline", [10, 11, 12, 13])

    term.edited_log_effect[10:14] = [0.2, 1.0, -0.4, 0.8]
    session.level_left("x_spline")
    np.testing.assert_allclose(term.edited_log_effect[10:14], 0.2)

    term.edited_log_effect[10:14] = [0.2, 1.0, -0.4, 0.8]
    session.level_right("x_spline")
    np.testing.assert_allclose(term.edited_log_effect[10:14], 0.8)

    term.edited_log_effect[10:14] = [0.2, 1.0, -0.4, 0.8]
    session.snap_highest("x_spline")
    np.testing.assert_allclose(term.edited_log_effect[10:14], 1.0)

    term.edited_log_effect[10:14] = [0.2, 1.0, -0.4, 0.8]
    session.snap_lowest("x_spline")
    np.testing.assert_allclose(term.edited_log_effect[10:14], -0.4)


def test_isotonic_anchors_selected_region_to_neighbors(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    term.edited_log_effect[:] = np.linspace(0.0, 1.0, term.size)
    term.edited_log_effect[10:15] = [0.8, -0.2, 0.6, 0.1, 0.5]

    left_neighbor = term.edited_log_effect[9]
    right_neighbor = term.edited_log_effect[15]
    session.select_indices("x_spline", range(10, 15))
    session.isotonic("x_spline", direction="increasing")

    after = term.edited_log_effect
    assert after[10] >= left_neighbor
    assert after[14] <= right_neighbor
    assert np.all(np.diff(after[10:15]) >= -1e-12)


def test_isotonic_neighbor_anchors_constrain_without_pinning_to_both_sides():
    values = np.array([-0.007, 0.043, 0.027, 0.052, 0.113], dtype=np.float64)
    session = EditorSession(
        model=None,
        terms={
            "age_band": EditableTerm(
                name="age_band",
                kind="categorical",
                levels=["18-24", "25-34", "35-49", "50-64", "65-80"],
                original_log_effect=values.copy(),
                edited_log_effect=values.copy(),
                weights=np.ones_like(values),
            )
        },
    )

    session.select_indices("age_band", [2, 3])
    session.isotonic("age_band", direction="increasing")

    after = session.terms["age_band"].edited_log_effect
    assert after[2] == pytest.approx(values[1])
    assert after[3] == pytest.approx(values[3])
    assert after[3] < values[4]
    assert np.all(np.diff(after) >= -1e-12)


def test_categorical_monotone_uses_directional_clamp_instead_of_pooling():
    values = np.array([-0.007, 0.039, 0.023, 0.047, 0.113], dtype=np.float64)
    session = EditorSession(
        model=None,
        terms={
            "age_band": EditableTerm(
                name="age_band",
                kind="categorical",
                levels=["18-24", "25-34", "35-49", "50-64", "65-80"],
                original_log_effect=values.copy(),
                edited_log_effect=values.copy(),
                weights=np.array([10.0, 10.0, 30.0, 20.0, 30.0]),
            )
        },
    )

    session.select_indices("age_band", [0, 1, 2, 3, 4])
    session.isotonic("age_band", direction="increasing")
    np.testing.assert_allclose(
        session.terms["age_band"].edited_log_effect,
        [-0.007, 0.039, 0.039, 0.047, 0.113],
    )

    session.reset("age_band")
    session.select_indices("age_band", [1, 2, 3, 4])
    session.isotonic("age_band", direction="decreasing")
    np.testing.assert_allclose(
        session.terms["age_band"].edited_log_effect,
        [-0.007, -0.007, -0.007, -0.007, -0.007],
    )


def test_smooth_anchors_selected_region_to_neighbors(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    term.edited_log_effect[:] = np.linspace(0.0, 1.0, term.size)
    term.edited_log_effect[10:15] = [0.8, -0.2, 0.6, 0.1, 0.5]

    before = term.edited_log_effect.copy()
    session.select_indices("x_spline", range(10, 15))
    session.smooth("x_spline", strength=1.0)

    after = term.edited_log_effect
    assert abs(after[10] - before[9]) <= abs(before[10] - before[9]) + 1e-12
    assert abs(after[14] - before[15]) <= abs(before[14] - before[15]) + 1e-12
    assert after[11] != pytest.approx(before[11])
    assert after[12] != pytest.approx(before[12])


def test_smooth_does_not_create_boundary_discontinuities(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    term.edited_log_effect[:] = np.sin(np.linspace(0.0, 16.0, term.size))
    before = term.edited_log_effect.copy()
    start, stop = 40, 95

    session.select_indices("x_spline", range(start, stop))
    session.smooth("x_spline", strength=1.0)

    after = term.edited_log_effect
    assert abs(after[start] - after[start - 1]) <= abs(before[start] - before[start - 1]) + 1e-12
    assert abs(after[stop - 1] - after[stop]) <= abs(before[stop - 1] - before[stop]) + 1e-12
    assert not np.allclose(after[start:stop], before[start:stop])


def test_smooth_without_selection_applies_to_whole_term(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    term.edited_log_effect[:] = np.sin(np.linspace(0.0, 20.0, term.size))
    before = term.edited_log_effect.copy()

    session.smooth("x_spline", strength=1.0)

    assert not np.allclose(term.edited_log_effect, before)
    assert session.history[-1].indices.size == term.size


def test_smooth_is_aggressive_for_narrow_spikes(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    term.edited_log_effect[:] = 0.0
    term.edited_log_effect[term.size // 2] = 1.0

    session.smooth("x_spline", strength=1.0)

    assert np.max(term.edited_log_effect) < 0.2


def test_spline_control_points_are_fixed_x_vertical_handles(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline", "region"])
    term = session.terms["x_spline"]
    controls = session.control_points("x_spline")

    assert 6 <= controls["x"].size <= 12
    assert np.all(np.diff(controls["x"]) > 0)
    assert controls["x"][0] >= float(term.x.min()) - 1e-12
    assert controls["x"][-1] <= float(term.x.max()) + 1e-12

    before_curve = term.edited_log_effect.copy()
    before_controls_x = controls["x"].copy()
    handle_index = controls["x"].size // 2
    session.move_control_point(
        "x_spline", handle_index, float(controls["log_effect"][handle_index] + 0.35)
    )

    after_controls = session.control_points("x_spline")
    np.testing.assert_allclose(after_controls["x"], before_controls_x)
    assert np.max(np.abs(term.edited_log_effect - before_curve)) > 0.05
    assert after_controls["log_effect"][handle_index] == pytest.approx(
        controls["log_effect"][handle_index] + 0.35,
        abs=0.08,
    )

    session.undo("x_spline")
    np.testing.assert_allclose(term.edited_log_effect, before_curve)

    with pytest.raises(TypeError, match="control handles"):
        session.control_points("region")


def test_spline_control_point_edit_is_local_to_cubic_basis_support(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    controls = session.control_points("x_spline")
    handle_index = controls["x"].size // 2
    basis_index = int(controls["basis_index"][handle_index])
    basis = editor_model._specs["x_spline"]._raw_basis_matrix(term.x)
    support = np.abs(basis[:, basis_index]) > 1e-12
    assert np.any(support)
    assert np.any(~support)

    before = term.edited_log_effect.copy()
    session.move_control_point(
        "x_spline",
        handle_index,
        float(controls["log_effect"][handle_index] + 0.35),
    )
    delta = term.edited_log_effect - before

    assert np.max(np.abs(delta[support])) > 0.05
    assert np.max(np.abs(delta[~support])) < 1e-8


def test_spline_control_point_count_selects_visible_basis_handles(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    controls = session.control_points("x_spline", n_handles=6)

    assert controls["x"].size == 6
    assert controls["basis_index"].size == 6
    assert np.all(np.diff(controls["basis_index"]) > 0)
    assert np.all(np.diff(controls["x"]) > 0)

    handle_index = 2
    basis_index = int(controls["basis_index"][handle_index])
    basis = editor_model._specs["x_spline"]._raw_basis_matrix(term.x)
    support = np.abs(basis[:, basis_index]) > 1e-12
    before = term.edited_log_effect.copy()

    session.move_control_point(
        "x_spline",
        handle_index,
        float(controls["log_effect"][handle_index] + 0.35),
        n_handles=6,
    )
    delta = term.edited_log_effect - before

    assert np.max(np.abs(delta[support])) > 0.05
    assert np.max(np.abs(delta[~support])) < 1e-8


def test_reset_restores_selection_or_current_term(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    original = term.original_log_effect.copy()

    session.select_indices("x_spline", [2, 3])
    session.shift("x_spline", 0.2)
    session.reset("x_spline")
    np.testing.assert_allclose(term.edited_log_effect[[2, 3]], original[[2, 3]])

    session.clear_selection("x_spline")
    term.edited_log_effect[:] = original + 0.3
    session.reset("x_spline")
    np.testing.assert_allclose(term.edited_log_effect, original)


def test_to_model_returns_copy_and_leaves_source_unchanged(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    original_beta = editor_model.result.beta.copy()
    original_intercept = editor_model.result.intercept

    session.select_x("x_spline", 2.0, 5.0)
    session.shift("x_spline", 0.35)
    edited = session.to_model()

    assert edited is not editor_model
    np.testing.assert_allclose(editor_model.result.beta, original_beta)
    assert editor_model.result.intercept == original_intercept
    assert not np.allclose(edited.result.beta, original_beta)


def test_materialized_model_reuses_one_copy_per_real_edit_epoch(
    editor_model,
    monkeypatch,
):
    from superglm.editor import session as session_module

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    real_apply = session_module.apply.apply_edits_to_model_copy_with_data
    applied_epochs: list[int] = []

    def counted_apply(*args, **kwargs):
        applied_epochs.append(session.edit_epoch)
        return real_apply(*args, **kwargs)

    monkeypatch.setattr(
        session_module.apply,
        "apply_edits_to_model_copy_with_data",
        counted_apply,
    )

    widget = session.widget()
    try:
        assert widget._current_model_for_evidence()[0] is editor_model
        assert applied_epochs == []

        session.select_x("x_spline", 2.0, 5.0)
        session.shift("x_spline", 0.35)
        first = widget._current_model_for_evidence()[0]
        assert widget._current_model_for_evidence()[0] is first
        assert applied_epochs == [1]

        session.shift("x_spline", 0.10)
        second = widget._current_model_for_evidence()[0]
        assert second is not first
        assert applied_epochs == [1, 2]

        session.undo()
        after_undo = widget._current_model_for_evidence()[0]
        assert after_undo is not second
        assert applied_epochs == [1, 2, 3]

        session.redo()
        after_redo = widget._current_model_for_evidence()[0]
        assert after_redo is not after_undo
        assert applied_epochs == [1, 2, 3, 4]

        session.reset("x_spline")
        assert widget._current_model_for_evidence()[0] is editor_model
        assert applied_epochs == [1, 2, 3, 4]
    finally:
        widget.close()


def test_materialized_model_is_invalidated_by_structural_replacement(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_indices("x_spline", [4])
    session.shift("x_spline", 0.2)
    widget = session.widget()
    cached = widget._current_model_for_evidence()[0]

    replacement = editor_model._clone_without_features(set())
    replacement.fit(editor_model._fit_X_ref, editor_model._fit_y_ref)
    session.replace_in_force_model(replacement)

    try:
        assert session._materialized_edit_model is None
        assert session._materialized_edit_epoch is None
        assert widget._current_model_for_evidence()[0] is replacement
        assert widget._current_model_for_evidence()[0] is not cached
    finally:
        widget.close()


def test_materialization_request_copies_terms_and_rejects_stale_publish(
    editor_model,
    editor_frame,
):
    from superglm.editor.apply import materialize_edit_request
    from superglm.editor.evaluation import default_metrics_dataset

    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        validation_data=(X.iloc[300:380], y[300:380], None),
    )
    session.select_indices("x_spline", [4, 5, 6])
    session.shift("x_spline", 0.2)
    dataset = default_metrics_dataset(session)
    request = session.capture_materialization_request()

    assert request is not None
    assert request.base_model is session.model
    assert request.dataset is dataset
    assert request.dataset.X is dataset.X
    assert request.dataset.y is dataset.y
    captured_term = request.terms["x_spline"]
    live_term = session.terms["x_spline"]
    assert captured_term is not live_term
    assert not np.shares_memory(captured_term.edited_log_effect, live_term.edited_log_effect)

    stale_model = materialize_edit_request(request)
    session.shift("x_spline", 0.1)
    assert session.publish_materialized_model(request, stale_model) is False
    assert session.cached_materialized_model(request.edit_epoch) is None

    current_request = session.capture_materialization_request()
    assert current_request is not None
    current_model = materialize_edit_request(current_request)
    assert session.publish_materialized_model(current_request, current_model) is True
    assert (
        session.cached_materialized_model(
            current_request.edit_epoch,
            model_revision=current_request.model_revision,
            base_model=current_request.base_model,
        )
        is current_model
    )
    assert (
        session.cached_materialized_model(
            current_request.edit_epoch,
            model_revision=current_request.model_revision + 1,
            base_model=current_request.base_model,
        )
        is None
    )
    assert (
        session.cached_materialized_model(
            current_request.edit_epoch,
            model_revision=current_request.model_revision,
            base_model=object(),
        )
        is None
    )


def test_to_model_stays_fresh_and_does_not_return_the_materialized_copy(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_indices("x_spline", [4, 5, 6])
    session.shift("x_spline", 0.2)

    widget = session.widget()
    try:
        materialized = widget._current_model_for_evidence()[0]
        first = session.to_model()
        second = session.to_model()

        assert first is not second
        assert first is not materialized
        assert second is not materialized
        assert widget._current_model_for_evidence()[0] is materialized
        np.testing.assert_allclose(first.result.beta, materialized.result.beta)
        np.testing.assert_allclose(second.result.beta, materialized.result.beta)
    finally:
        widget.close()


def test_to_model_without_edits_keeps_transient_state_private(editor_model):
    editor_model.summary()
    editor_model.metrics(
        editor_model._fit_X_ref,
        editor_model._fit_y_ref,
        sample_weight=editor_model._fit_sample_weight_ref,
        offset=editor_model._fit_offset_ref,
    )
    source_objects = {
        name: getattr(editor_model, name)
        for name in (
            "_runtime_canonical_state",
            "_fast_prediction_state",
            "_prediction_plan",
            "_fit_mu",
            "_fit_null_mu",
            "_fit_stats",
            "_fit_metrics_cache",
            "_summary_cache",
        )
    }
    source_mu = editor_model._fit_mu.copy()
    source_null_mu = editor_model._fit_null_mu.copy()
    session = EditorSession.from_model(editor_model, terms=["x_spline"])

    copied = session.to_model()

    for name in (
        "_dm",
        "_fit_X_ref",
        "_fit_y_ref",
        "_fit_sample_weight_ref",
        "_fit_offset_ref",
        "_fit_weights",
        "_fit_offset",
        "_fit_data_guard",
    ):
        assert getattr(copied, name) is getattr(editor_model, name)
    for name, source_value in source_objects.items():
        assert getattr(copied, name) is not source_value
    assert not np.shares_memory(copied._fit_mu, editor_model._fit_mu)
    assert not np.shares_memory(copied._fit_null_mu, editor_model._fit_null_mu)

    copied._runtime_canonical_state["copy_only"] = True
    copied._fast_prediction_state["copy_only"] = True
    copied._prediction_plan["copy_only"] = []
    copied._fit_mu[0] += 1.0
    copied._fit_null_mu[0] += 1.0
    copied._fit_metrics_cache.__dict__["copy_only"] = True
    copied._summary_cache["copy_only"] = True
    for name in source_objects:
        setattr(copied, name, None)

    assert all(getattr(editor_model, name) is value for name, value in source_objects.items())
    assert "copy_only" not in editor_model._runtime_canonical_state
    assert "copy_only" not in editor_model._fast_prediction_state
    assert "copy_only" not in editor_model._prediction_plan
    assert "copy_only" not in editor_model._fit_metrics_cache.__dict__
    assert "copy_only" not in editor_model._summary_cache
    np.testing.assert_array_equal(editor_model._fit_mu, source_mu)
    np.testing.assert_array_equal(editor_model._fit_null_mu, source_null_mu)


def test_materialized_model_shares_only_row_scale_fit_inputs():
    rng = np.random.default_rng(20260804)
    n = 120
    X = pd.DataFrame({"x": rng.normal(size=n), "z": rng.normal(size=n)})
    sample_weight = rng.uniform(0.5, 2.0, size=n)
    offset = rng.normal(0.0, 0.03, size=n)
    y = (
        0.4
        + 0.3 * X["x"].to_numpy()
        - 0.2 * X["z"].to_numpy()
        + 0.1 * X["x"].to_numpy() * X["z"].to_numpy()
        + offset
        + rng.normal(0.0, 0.04, size=n)
    )
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.1,
        features={"x": Numeric(), "z": Numeric()},
        interactions=[("x", "z")],
    )
    model.fit(X, y, sample_weight=sample_weight, offset=offset)
    model.summary()
    model.metrics(
        model._fit_X_ref,
        model._fit_y_ref,
        sample_weight=model._fit_sample_weight_ref,
        offset=model._fit_offset_ref,
    )
    source_objects = {
        name: getattr(model, name)
        for name in (
            "_fit_mu",
            "_fit_null_mu",
            "_fit_stats",
            "_fit_metrics_cache",
            "_prediction_plan",
            "_runtime_canonical_state",
            "_fast_prediction_state",
            "_summary_cache",
        )
    }
    source_mu = model._fit_mu.copy()
    source_null_mu = model._fit_null_mu.copy()
    source_beta = model.result.beta.copy()
    source_deviance = model.result.deviance

    session = EditorSession.from_model(model, terms=["x"])
    session.select_indices("x", [0])
    session.shift("x", 0.2)
    widget = session.widget()
    try:
        edited = widget._current_model_for_evidence()[0]
    finally:
        widget.close()

    for name in (
        "_dm",
        "_fit_X_ref",
        "_fit_y_ref",
        "_fit_sample_weight_ref",
        "_fit_offset_ref",
        "_fit_weights",
        "_fit_offset",
        "_fit_data_guard",
    ):
        assert getattr(edited, name) is getattr(model, name)

    assert edited._result is not model._result
    assert edited._solver_result is not model._solver_result
    assert not np.shares_memory(edited._result.beta, model._result.beta)
    assert not np.shares_memory(edited._solver_result.beta, model._solver_result.beta)
    assert edited._specs is not model._specs
    assert all(edited._specs[name] is not spec for name, spec in model._specs.items())
    assert edited._interaction_specs is not model._interaction_specs
    assert all(
        edited._interaction_specs[name] is not spec
        for name, spec in model._interaction_specs.items()
    )
    assert edited.penalty is not model.penalty
    assert edited._runtime_canonical_state is not model._runtime_canonical_state
    assert edited._fast_prediction_state is not model._fast_prediction_state

    assert edited._fit_mu is not model._fit_mu
    assert edited._fit_null_mu is not model._fit_null_mu
    assert not np.shares_memory(edited._fit_mu, model._fit_mu)
    assert not np.shares_memory(edited._fit_null_mu, model._fit_null_mu)
    assert edited._fit_stats is not model._fit_stats
    assert edited._prediction_plan is not model._prediction_plan
    assert edited._fit_metrics_cache is None
    assert edited._summary_cache is None

    assert all(getattr(model, name) is value for name, value in source_objects.items())
    np.testing.assert_array_equal(model._fit_mu, source_mu)
    np.testing.assert_array_equal(model._fit_null_mu, source_null_mu)
    np.testing.assert_array_equal(model.result.beta, source_beta)
    assert model.result.deviance == source_deviance


def test_widget_evidence_and_export_reuse_materialized_model(
    editor_model,
    tmp_path,
    monkeypatch,
):
    import joblib

    import superglm.editor.apply as apply_module

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_indices("x_spline", [4, 5, 6])
    session.shift("x_spline", 0.2)
    widget = session.widget()
    materialized_calls = 0
    dumped: list[object] = []
    original_apply = apply_module.apply_edits_to_model_copy_with_data
    original_dump = joblib.dump

    def counted_apply(*args, **kwargs):
        nonlocal materialized_calls
        materialized_calls += 1
        return original_apply(*args, **kwargs)

    def capture_dump(model, target):
        dumped.append(model)
        return original_dump(model, target)

    monkeypatch.setattr(apply_module, "apply_edits_to_model_copy_with_data", counted_apply)
    monkeypatch.setattr("joblib.dump", capture_dump)
    try:
        summary = widget._summary("in_force")
        metrics = widget._metrics("deviance", "in_force")
        report = widget._report("final")
        saved = widget._save_model(path=str(tmp_path / "edited-model.joblib"))
        downloaded, filename = widget._download_model("edited-model.joblib")
    finally:
        widget.close()

    materialized = session._materialized_edit_model
    assert materialized_calls == 1
    assert materialized is not None
    assert summary["available"] is True
    assert metrics["available"] is True
    assert report["summary"]["available"] is True
    assert Path(saved["path"]).exists()
    assert filename == "edited-model.joblib"
    assert downloaded
    assert dumped == [materialized, materialized]


def test_to_model_refreshes_fit_statistics_after_manual_edit(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    original_deviance = editor_model.result.deviance

    session.select_x("x_spline", 2.0, 5.0)
    session.shift("x_spline", 1.25)
    edited = session.to_model()

    y = np.asarray(edited._fit_y_ref, dtype=np.float64)
    mu = edited.predict(edited._fit_X_ref, offset=edited._fit_offset_ref)
    expected_deviance = float(
        np.sum(edited._fit_weights * edited._distribution.deviance_unit(y, mu))
    )

    assert edited.result.deviance == pytest.approx(expected_deviance)
    assert edited.summary()["deviance"]["deviance"] == pytest.approx(expected_deviance)
    assert edited.result.deviance != pytest.approx(original_deviance)


def test_to_model_publishes_synchronized_immutable_result_revision(editor_model):
    source_revision = editor_model._fit_revision
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_x("x_spline", 2.0, 5.0)
    session.shift("x_spline", 0.2)

    edited = session.to_model()

    assert edited._fit_revision == source_revision + 1
    assert edited._fit_state.revision == edited._fit_revision
    assert edited._fit_state.repair_revision == editor_model._fit_state.repair_revision + 1
    for name, projected in edited._fit_state.projections.items():
        assert getattr(edited, name) is projected, name
    with pytest.raises(AttributeError, match="published"):
        edited.result.intercept = 0.0
    with pytest.raises(AttributeError, match="published"):
        edited.result.beta = edited.result.beta.copy()
    with pytest.raises(ValueError):
        edited.result.beta.setflags(write=True)


def test_editor_coefficient_revision_invalidates_accepted_reml_mode(tmp_path, monkeypatch):
    from superglm._debug import get_debug_level, set_debug_level

    previous = get_debug_level()
    monkeypatch.setenv("SUPERGLM_DEBUG_DIR", str(tmp_path))
    set_debug_level(2)
    try:
        x = np.linspace(0.0, 1.0, 140)
        X = pd.DataFrame({"x": x})
        y = 0.2 + np.sin(4.0 * x) + 0.01 * np.cos(17.0 * x)
        model = SuperGLM(
            family="gaussian",
            selection_penalty=0.0,
            features={"x": Spline(n_knots=7)},
        ).fit_reml(X, y, max_reml_iter=2)
        assert model._solver_result.state_id is not None
        assert model._reml_result.objective is not None
        session = EditorSession.from_model(model, terms=["x"])
        session.select_x("x", 0.25, 0.75)
        session.shift("x", 0.2)

        edited = session.to_model()

        assert edited._solver_result.state_id is None
        assert edited._solver_result.evaluation_id is None
        assert edited._solver_result.log_det_H is None
        assert edited._solver_result.reml_hessian_rank is None
        assert not edited._solver_result.converged
        assert edited._solver_result.termination_reason == "coefficients_revised"
        assert edited._reml_result.objective is None
        assert not edited._reml_result.converged
        assert edited._reml_result.termination_reason == "coefficients_revised"
    finally:
        set_debug_level(previous)


def test_editor_scoring_revision_owns_weights_and_offset_and_invalidates_cache():
    x = np.linspace(-1.0, 1.0, 120)
    X = pd.DataFrame({"x": x})
    y = 0.4 + 0.7 * x + 0.03 * np.sin(9.0 * x)
    weights = np.linspace(0.5, 1.5, len(x))
    offset = np.linspace(-0.1, 0.1, len(x))
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Numeric()},
    ).fit(X, y)

    edited = apply_edits_to_model_copy_with_data(
        model,
        {},
        X=X,
        y=y,
        sample_weight=weights,
        offset=offset,
    )
    cached = edited.metrics(X, y, sample_weight=weights, offset=offset)
    fitted_weight = float(edited._fit_weights[0])
    fitted_offset = float(edited._fit_offset[0])

    assert not np.shares_memory(edited._fit_weights, weights)
    assert not np.shares_memory(edited._fit_offset, offset)
    assert not edited._fit_weights.flags.writeable
    assert not edited._fit_offset.flags.writeable

    weights[0] *= 20.0
    offset[0] += 3.0
    refreshed = edited.metrics(X, y, sample_weight=weights, offset=offset)

    assert edited._fit_weights[0] == pytest.approx(fitted_weight)
    assert edited._fit_offset[0] == pytest.approx(fitted_offset)
    assert refreshed is not cached


def test_to_model_weight_only_override_reuses_retained_offset():
    rng = np.random.default_rng(20260814)
    n = 140
    X = pd.DataFrame({"x": rng.normal(size=n)})
    train_weight = rng.uniform(0.5, 1.5, size=n)
    train_offset = rng.normal(0.0, 0.08, size=n)
    y = 0.4 + 0.3 * X["x"].to_numpy() + train_offset + rng.normal(0.0, 0.04, size=n)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Numeric()},
    ).fit(X, y, sample_weight=train_weight, offset=train_offset)
    override_weight = np.linspace(0.75, 1.75, n)
    session = EditorSession.from_model(model, terms=["x"])
    session.select_indices("x", [0])
    session.shift("x", 0.2)

    edited = session.to_model(sample_weight=override_weight)

    mu = edited.predict(X, offset=train_offset)
    expected = float(np.sum(override_weight * edited._distribution.deviance_unit(y, mu)))
    assert edited.result.deviance == pytest.approx(expected)
    np.testing.assert_array_equal(edited._fit_offset, model._fit_offset)
    assert edited._fit_offset_ref is model._fit_offset_ref
    np.testing.assert_array_equal(edited._fit_weights, override_weight)


def test_to_model_offset_only_override_reuses_retained_weights():
    rng = np.random.default_rng(20260815)
    n = 140
    X = pd.DataFrame({"x": rng.normal(size=n)})
    train_weight = rng.uniform(0.5, 1.5, size=n)
    train_offset = rng.normal(0.0, 0.08, size=n)
    y = 0.4 + 0.3 * X["x"].to_numpy() + train_offset + rng.normal(0.0, 0.04, size=n)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Numeric()},
    ).fit(X, y, sample_weight=train_weight, offset=train_offset)
    override_offset = train_offset + np.linspace(-0.02, 0.02, n)
    session = EditorSession.from_model(model, terms=["x"])
    session.select_indices("x", [0])
    session.shift("x", 0.2)

    edited = session.to_model(offset=override_offset)

    mu = edited.predict(X, offset=override_offset)
    expected = float(np.sum(train_weight * edited._distribution.deviance_unit(y, mu)))
    assert edited.result.deviance == pytest.approx(expected)
    np.testing.assert_array_equal(edited._fit_weights, model._fit_weights)
    assert edited._fit_sample_weight_ref is model._fit_sample_weight_ref
    np.testing.assert_array_equal(edited._fit_offset, override_offset)


def test_to_model_rejects_compact_coefficient_edits_without_scoring_data():
    rng = np.random.default_rng(20260816)
    X = pd.DataFrame({"x": rng.normal(size=120)})
    y = 0.4 + 0.3 * X["x"].to_numpy() + rng.normal(0.0, 0.04, size=len(X))
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"x": Numeric()},
    ).fit(X, y)
    session = EditorSession.from_model(model, terms=["x"])
    session.select_indices("x", [0])
    session.shift("x", 0.2)
    installed_dict = model.__dict__
    installed_result = model.result

    with pytest.raises(RuntimeError, match="coefficient edits require scoring data"):
        session.to_model()

    assert model.__dict__ is installed_dict
    assert model.result is installed_result


def test_to_model_compact_edit_with_scoring_data_releases_row_scale_state():
    rng = np.random.default_rng(20260817)
    n = 130
    X = pd.DataFrame({"x": rng.normal(size=n)})
    weights = rng.uniform(0.5, 1.5, size=n)
    offset = rng.normal(0.0, 0.05, size=n)
    y = 0.4 + 0.3 * X["x"].to_numpy() + offset + rng.normal(0.0, 0.04, size=n)
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"x": Numeric()},
    ).fit(X, y, sample_weight=weights, offset=offset)
    session = EditorSession.from_model(model, terms=["x"])
    session.select_indices("x", [0])
    session.shift("x", 0.2)

    edited = session.to_model(
        X=X,
        y=y,
        sample_weight=weights,
        offset=offset,
    )

    mu = edited.predict(X, offset=offset)
    expected = float(np.sum(weights * edited._distribution.deviance_unit(y, mu)))
    assert edited.result.deviance == pytest.approx(expected)
    assert edited._fit_state.retained is False
    for name in (
        "_dm",
        "_fit_X_ref",
        "_fit_y_ref",
        "_fit_sample_weight_ref",
        "_fit_offset_ref",
        "_fit_weights",
        "_fit_offset",
        "_fit_mu",
        "_fit_null_mu",
        "_fit_data_guard",
    ):
        assert getattr(edited, name) is None, name


def test_to_model_noop_copy_does_not_start_fitted_state_revision(editor_model, monkeypatch):
    def unexpected_revision(*args, **kwargs):
        raise AssertionError("no-op editor copy must not start a fitted-state revision")

    monkeypatch.setattr(
        "superglm.editor.apply.FittedStateRevision.start",
        unexpected_revision,
    )

    copied = EditorSession.from_model(editor_model, terms=["x_spline"]).to_model()

    assert copied is not editor_model
    assert copied.result is not editor_model.result
    assert copied._fit_state is not editor_model._fit_state
    assert copied._fit_revision == editor_model._fit_revision
    for name, projected in copied._fit_state.projections.items():
        assert getattr(copied, name) is projected, name


def test_stale_summary_routes_varying_interactions_as_curve_groups():
    rng = np.random.default_rng(20260726)
    n = 260
    x = rng.uniform(-1.0, 1.0, n)
    cat = rng.choice(["A", "B", "C", "D"], n)
    X = pd.DataFrame({"x": x, "cat": cat})
    y = (
        0.2
        + 0.3 * x
        + 0.15 * x**2 * (cat == "B")
        - 0.10 * x**2 * (cat == "C")
        + rng.normal(0.0, 0.05, n)
    )
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Polynomial(degree=2), "cat": Categorical(base="first")},
        interactions=[("x", "cat")],
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["x"])
    session.select_indices("x", [0])
    session.shift("x", 0.05)

    edited = session.to_model()
    rows = edited.summary()._coef_rows
    row_by_name = {row.name: row for row in rows}

    assert "x:cat[B]" in row_by_name
    assert row_by_name["x:cat[B]"].is_spline is True
    assert row_by_name["x:cat[B]"].n_params == 2
    assert "x:cat[B][B]" not in row_by_name


def test_to_model_can_refresh_fit_statistics_from_explicit_data():
    rng = np.random.default_rng(20260710)
    n = 180
    X = pd.DataFrame({"x": rng.normal(size=n)})
    sample_weight = rng.uniform(0.5, 2.0, size=n)
    offset = rng.normal(0.0, 0.05, size=n)
    y = 0.4 + 0.3 * X["x"].to_numpy() + offset + rng.normal(0.0, 0.04, size=n)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X, y, sample_weight=sample_weight, offset=offset)
    session = EditorSession.from_model(model, terms=["x"])
    session.select_indices("x", [0])
    session.shift("x", 0.2)

    eval_offset = offset + 0.01
    edited = session.to_model(X=X, y=y, sample_weight=sample_weight, offset=eval_offset)
    mu = edited.predict(X, offset=eval_offset)
    expected_deviance = float(np.sum(sample_weight * edited._distribution.deviance_unit(y, mu)))

    assert edited.result.deviance == pytest.approx(expected_deviance)
    assert edited._fit_stats.pearson_chi2 == pytest.approx(expected_deviance)


def test_to_model_refreshes_fit_statistics_from_explicit_data_without_manual_edits():
    rng = np.random.default_rng(20260725)
    n_train = 160
    n_val = 60
    X_train = pd.DataFrame({"x": rng.normal(size=n_train)})
    y_train = 0.4 + 0.3 * X_train["x"].to_numpy() + rng.normal(0.0, 0.04, size=n_train)
    X_val = pd.DataFrame({"x": rng.normal(size=n_val)})
    y_val = 0.4 + 0.3 * X_val["x"].to_numpy() + rng.normal(0.0, 0.12, size=n_val)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X_train, y_train)
    original_deviance = model.result.deviance
    session = EditorSession.from_model(model, terms=["x"])

    edited = session.to_model(X=X_val, y=y_val)
    mu = edited.predict(X_val)
    expected_deviance = float(np.sum(edited._distribution.deviance_unit(y_val, mu)))

    assert edited.result.deviance == pytest.approx(expected_deviance)
    assert edited.summary()["deviance"]["deviance"] == pytest.approx(expected_deviance)
    assert edited._fit_stats.pearson_chi2 == pytest.approx(expected_deviance)
    assert edited.result.deviance != pytest.approx(original_deviance)


def test_to_model_explicit_data_omits_retained_offset_and_weights_when_not_supplied():
    rng = np.random.default_rng(20260720)
    n_train = 180
    n_val = 70
    X_train = pd.DataFrame({"x": rng.normal(size=n_train)})
    train_weight = rng.uniform(0.5, 2.0, size=n_train)
    train_offset = rng.normal(0.0, 0.05, size=n_train)
    y_train = (
        0.4 + 0.3 * X_train["x"].to_numpy() + train_offset + rng.normal(0.0, 0.04, size=n_train)
    )
    X_val = pd.DataFrame({"x": rng.normal(size=n_val)})
    y_val = 0.4 + 0.3 * X_val["x"].to_numpy() + rng.normal(0.0, 0.04, size=n_val)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X_train, y_train, sample_weight=train_weight, offset=train_offset)
    session = EditorSession.from_model(model, terms=["x"])
    session.select_indices("x", [0])
    session.shift("x", 0.2)

    edited = session.to_model(X=X_val, y=y_val)
    mu = edited.predict(X_val)
    expected_deviance = float(np.sum(edited._distribution.deviance_unit(y_val, mu)))

    assert edited.result.deviance == pytest.approx(expected_deviance)
    assert edited._fit_offset is None
    assert edited._fit_offset_ref is None
    np.testing.assert_allclose(edited._fit_weights, np.ones(n_val))
    assert edited._fit_sample_weight_ref is None


def test_to_model_rejects_partial_explicit_scoring_data(editor_model, editor_frame):
    X, _ = editor_frame
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_x("x_spline", 2.0, 5.0)
    session.shift("x_spline", 0.2)

    with pytest.raises(ValueError, match="Explicit scoring data requires both X and y"):
        session.to_model(X=X)


@pytest.mark.parametrize(
    ("invalid_weight", "message"),
    [
        (-0.1, "nonnegative"),
        (np.nan, "finite"),
    ],
)
def test_to_model_rejects_invalid_explicit_scoring_weights(
    editor_model,
    editor_frame,
    invalid_weight,
    message,
):
    X, y = editor_frame
    weights = np.ones(len(y), dtype=np.float64)
    weights[0] = invalid_weight
    original_dict = editor_model.__dict__

    with pytest.raises(ValueError, match=message):
        apply_edits_to_model_copy_with_data(
            editor_model,
            {},
            X=X,
            y=y,
            sample_weight=weights,
        )

    assert editor_model.__dict__ is original_dict


def test_x_domain_handles_constant_continuous_grid():
    from superglm.editor.payloads import _x_domain

    assert _x_domain([3.0, 3.0, 3.0]) == [2.5, 3.5]
    assert _x_domain([0.0, 0.0]) == [-0.5, 0.5]


def test_jsonable_converts_numpy_boolean_scalars():
    import json

    from superglm.editor.io import jsonable

    payload = jsonable({"ok": np.bool_(True), "bad": np.bool_(False)})

    assert payload == {"ok": True, "bad": False}
    assert type(payload["ok"]) is bool
    assert type(payload["bad"]) is bool
    assert json.dumps(payload, allow_nan=False) == '{"ok": true, "bad": false}'


def test_weighted_average_falls_back_when_selected_exposure_is_zero(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    term = session.terms["region"]
    term.weights = np.array([1.0, 0.0, 0.0])
    term.edited_log_effect = np.array([0.0, 0.2, 0.8])

    session.select_levels("region", ["B", "C"])
    session.weighted_average("region")

    np.testing.assert_allclose(term.edited_log_effect[[1, 2]], 0.5)


def test_in_force_summary_uses_session_train_data_without_retained_fit_state():
    from superglm.editor.summaries import summary_payload

    rng = np.random.default_rng(20260717)
    n = 180
    X = pd.DataFrame({"x": rng.normal(size=n)})
    y = 0.4 + 0.3 * X["x"].to_numpy() + rng.normal(0.0, 0.04, size=n)
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["x"], train_data=(X, y, None))
    session.select_indices("x", [0])
    session.shift("x", 0.2)

    widget = SimpleNamespace(session=session, _in_force_info=None)
    payload = summary_payload(widget, "in_force")

    assert payload["available"] is True
    assert payload["compact"]["model"]["deviance"] is not None


@pytest.mark.parametrize(
    ("phi_method", "expected_status"),
    [
        ("mle", "not computed"),
        ("pearson", "unavailable for Pearson plug-in"),
    ],
)
def test_in_force_summary_never_computes_tweedie_ci(
    editor_model,
    phi_method,
    expected_status,
):
    from superglm.editor.summaries import summary_payload

    def unexpected_ci(*args, **kwargs):
        raise AssertionError("editor summary must not evaluate a Tweedie profile CI")

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.model._tweedie_profile_result = SimpleNamespace(
        p_hat=1.55,
        phi_hat=0.8,
        nll=11.0,
        method="brent",
        phi_method=phi_method,
        density_exact=True,
        _ci_cache={0.05: (1.4, 1.7)} if phi_method == "pearson" else {},
        ci=unexpected_ci,
        ci_details=unexpected_ci,
    )
    session.model._summary_cache = None
    widget = SimpleNamespace(session=session, _in_force_info=None)

    payload = summary_payload(widget, "in_force")

    assert payload["available"] is True
    assert payload["compact"]["model"]["tweedie_p"] == 1.55
    assert payload["compact"]["model"]["tweedie_p_ci"] is None
    assert payload["compact"]["model"]["tweedie_p_ci_status"] == expected_status


def test_in_force_summary_uses_validation_data_without_retained_fit_state():
    from superglm.editor.summaries import summary_payload

    rng = np.random.default_rng(20260730)
    n_train = 180
    n_val = 70
    X_train = pd.DataFrame({"x": rng.normal(size=n_train)})
    y_train = 0.4 + 0.3 * X_train["x"].to_numpy() + rng.normal(0.0, 0.04, size=n_train)
    X_val = pd.DataFrame({"x": rng.normal(size=n_val)})
    y_val = 0.4 + 0.3 * X_val["x"].to_numpy() + rng.normal(0.0, 0.04, size=n_val)
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X_train, y_train)
    session = EditorSession.from_model(model, terms=["x"], validation_data=(X_val, y_val, None))
    session.select_indices("x", [0])
    session.shift("x", 0.2)

    widget = SimpleNamespace(session=session, _in_force_info=None)
    payload = summary_payload(widget, "in_force")

    assert payload["available"] is True
    assert payload["compact"]["model"]["deviance"] is not None
    assert payload["compact"]["model"]["n_obs"] == n_val


def test_unedited_in_force_summary_preserves_retained_inference_without_fit_state():
    from superglm.editor.summaries import summary_payload

    rng = np.random.default_rng(20260723)
    n = 180
    X = pd.DataFrame({"x": rng.normal(size=n)})
    y = 0.4 + 0.3 * X["x"].to_numpy() + rng.normal(0.0, 0.04, size=n)
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["x"])

    widget = SimpleNamespace(session=session, _in_force_info=None)
    payload = summary_payload(widget, "in_force")

    assert payload["available"] is True
    assert payload["compact"]["model"]["deviance"] is not None


def test_to_model_mean_centering_applies_equivalent_native_curve_edit(editor_model):
    native_session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        n_points=80,
        centering="native",
        with_se=False,
    )
    mean_session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        n_points=80,
        centering="mean",
        with_se=False,
    )
    indices = [22, 23, 24, 25]
    native_session.select_indices("x_spline", indices)
    mean_session.select_indices("x_spline", indices)
    native_session.shift("x_spline", 0.35)
    mean_session.shift("x_spline", 0.35)

    native_model = native_session.to_model()
    mean_model = mean_session.to_model()

    X_ref = editor_model._fit_X_ref
    np.testing.assert_allclose(
        mean_model.predict(X_ref),
        native_model.predict(X_ref),
        rtol=1e-10,
        atol=1e-10,
    )


def test_edited_offset_mean_centering_applies_equivalent_native_curve_edit(editor_model):
    native_session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        n_points=80,
        centering="native",
        with_se=False,
    )
    mean_session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        n_points=80,
        centering="mean",
        with_se=False,
    )
    indices = [22, 23, 24, 25]
    native_session.select_indices("x_spline", indices)
    mean_session.select_indices("x_spline", indices)
    native_session.shift("x_spline", 0.35)
    mean_session.shift("x_spline", 0.35)

    X_ref = editor_model._fit_X_ref
    np.testing.assert_allclose(
        mean_session.edited_offset("x_spline", X=X_ref),
        native_session.edited_offset("x_spline", X=X_ref),
        rtol=1e-10,
        atol=1e-10,
    )


def test_offset_labels_use_native_values_for_mean_centered_terms(editor_model):
    from superglm.editor.summaries import offset_label_payload
    from superglm.editor.terms import native_log_effect_values

    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        n_points=40,
        centering="mean",
        with_se=False,
    )
    session.select_indices("x_spline", [10, 11, 12])
    session.shift("x_spline", 0.35)

    label = offset_label_payload(session, ["x_spline"])[0]
    actual = np.asarray([row["log_offset"] for row in label["values"]], dtype=np.float64)
    expected = native_log_effect_values(session.terms["x_spline"])

    np.testing.assert_allclose(actual, expected)


def test_to_model_marks_edited_copy_inference_stale(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_x("x_spline", 2.0, 5.0)
    session.shift("x_spline", 0.35)

    edited = session.to_model()

    assert edited._editor_inference_stale is True
    assert edited._editor_edits["terms"] == ["x_spline"]
    summary = edited.summary()
    assert summary["standard_errors"] == {"inference_stale": True}
    assert "Editor edits applied" in str(summary)
    spline_row = next(row for row in summary._coef_rows if row.name == "x_spline")
    assert spline_row.wald_p is None
    assert summary._basis_detail == {}
    with pytest.warns(UserWarning, match="Editor coefficient edits"):
        inference = edited.term_inference("x_spline", with_se=True)
    assert inference.se_log_relativity is None
    assert inference.ci_lower is None
    with pytest.warns(UserWarning, match="Editor coefficient edits"):
        relativities = edited.relativities(with_se=True)
    assert "se_log_relativity" not in relativities["x_spline"].columns


def test_stale_editor_summary_skips_coef_row_inference(editor_model, monkeypatch):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_x("x_spline", 2.0, 5.0)
    session.shift("x_spline", 0.35)
    edited = session.to_model()

    def fail_build_coef_rows(*args, **kwargs):
        raise AssertionError("stale editor summary should not build inference rows")

    monkeypatch.setattr("superglm.inference.coef_tables.build_coef_rows", fail_build_coef_rows)

    summary = edited.summary()

    assert summary["standard_errors"] == {"inference_stale": True}
    spline_row = next(row for row in summary._coef_rows if row.name == "x_spline")
    assert spline_row.wald_p is None


def test_session_uses_supplied_train_data_for_weights_without_retained_fit_state():
    rng = np.random.default_rng(20260724)
    X = pd.DataFrame(
        {
            "region": ["A", "A", "B", "B", "B", "C"],
            "x": rng.normal(size=6),
        }
    )
    sample_weight = np.array([1.0, 2.0, 3.0, 4.0, 5.0, 6.0])
    y = np.array([0.2, 0.3, 0.9, 1.0, 1.1, -0.1])
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"region": Categorical(base="first"), "x": Numeric()},
    )
    model.fit(X, y, sample_weight=sample_weight)
    assert getattr(model, "_fit_X_ref", None) is None

    session = EditorSession.from_model(
        model,
        terms=["region"],
        train_data=(X, y, sample_weight),
    )

    expected = np.array([3.0, 12.0, 6.0])
    np.testing.assert_allclose(session.terms["region"].weights, expected)


def test_replacement_reuses_supplied_train_data_for_weights_without_retained_fit_state():
    rng = np.random.default_rng(20260724)
    X = pd.DataFrame(
        {
            "region": ["A", "A", "B", "B", "B", "C"],
            "x": rng.normal(size=6),
        }
    )
    sample_weight = np.array([1.0, 2.0, 3.0, 4.0, 5.0, 6.0])
    y = np.array([0.2, 0.3, 0.9, 1.0, 1.1, -0.1])
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"region": Categorical(base="first"), "x": Numeric()},
    )
    model.fit(X, y, sample_weight=sample_weight)
    session = EditorSession.from_model(
        model,
        terms=["region"],
        train_data=(X, y, sample_weight),
    )

    session.select_levels("region", ["B", "C"])
    session.replace_with_collapsed_levels("region", method="fit")

    expected = np.array([3.0, 12.0, 6.0])
    np.testing.assert_allclose(session.terms["region"].weights, expected)


def test_collapsed_level_member_edits_apply_to_whole_group(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    session.select_levels("region", ["B", "C"])
    session.replace_with_collapsed_levels("region", method="fit")

    term = session.terms["region"]
    b_idx = term.levels.index("B")
    c_idx = term.levels.index("C")
    original = term.edited_log_effect.copy()

    session.select_levels("region", ["B"])
    np.testing.assert_array_equal(session.selection("region"), np.array([b_idx]))

    session.shift("region", 0.2)
    assert term.edited_log_effect[b_idx] == pytest.approx(original[b_idx] + 0.2)
    assert term.edited_log_effect[c_idx] == pytest.approx(original[c_idx] + 0.2)

    session.set_values("region", [b_idx], [1.25])
    assert term.edited_log_effect[b_idx] == pytest.approx(1.25)
    assert term.edited_log_effect[c_idx] == pytest.approx(1.25)


def test_collapsed_level_member_conflicting_values_are_rejected(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    session.select_levels("region", ["B", "C"])
    session.replace_with_collapsed_levels("region", method="fit")
    term = session.terms["region"]
    b_idx = term.levels.index("B")
    c_idx = term.levels.index("C")

    with pytest.raises(ValueError, match="Conflicting values"):
        session.set_values("region", [b_idx, c_idx], [0.1, 0.2])


def test_edited_offset_factor_and_refit_with_fixed_offset(editor_model, editor_frame):
    X, _ = editor_frame
    session = EditorSession.from_model(editor_model, terms=["region", "x_spline"])
    session.select_levels("region", ["B"])
    session.shift("region", 0.4)

    offset = session.edited_offset("region", X=X)
    factor = session.edited_offset_factor("region", X=X)
    assert np.all(factor > 0)
    assert np.mean(offset[X["region"] == "B"]) > np.mean(offset[X["region"] == "A"])

    refit = session.refit_with_edited_offset(terms=["region"], method="fit")

    assert "region" not in refit.features
    assert "x_spline" in refit.features
    np.testing.assert_allclose(refit._fit_offset, offset)
    assert refit._editor_offset["terms"] == ["region"]
    refit_summary = refit.summary()
    assert "Editor offset refit" in str(refit_summary)
    inference = refit.term_inference("x_spline", with_se=True)
    assert inference.se_log_relativity is not None


def test_offset_refit_rejects_terms_used_by_interactions():
    rng = np.random.default_rng(20260731)
    n = 180
    x = rng.normal(size=n)
    region = rng.choice(["A", "B", "C"], n)
    X = pd.DataFrame({"x": x, "region": region})
    y = 0.3 + 0.4 * x + 0.2 * (region == "B") * x + rng.normal(0.0, 0.04, n)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Numeric(), "region": Categorical(base="first")},
        interactions=[("x", "region")],
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["x", "region"], train_data=(X, y, None))
    session.select_indices("x", [0])
    session.shift("x", 0.2)

    with pytest.raises(ValueError, match="used by interaction"):
        session.refit_with_edited_offset("x", method="fit")


def test_refit_explicit_data_omits_session_train_weights_and_offsets_when_not_supplied():
    rng = np.random.default_rng(20260721)
    n_train = 180
    n_val = 70
    X_train = pd.DataFrame({"x": rng.normal(size=n_train), "z": rng.normal(size=n_train)})
    train_weight = rng.uniform(0.5, 2.0, size=n_train)
    train_offset = rng.normal(0.0, 0.05, size=n_train)
    y_train = (
        0.4
        + 0.3 * X_train["x"].to_numpy()
        - 0.2 * X_train["z"].to_numpy()
        + train_offset
        + rng.normal(0.0, 0.04, n_train)
    )
    X_val = pd.DataFrame({"x": rng.normal(size=n_val), "z": rng.normal(size=n_val)})
    y_val = (
        0.4
        + 0.3 * X_val["x"].to_numpy()
        - 0.2 * X_val["z"].to_numpy()
        + rng.normal(0.0, 0.04, n_val)
    )
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Numeric(), "z": Numeric()},
    )
    model.fit(X_train, y_train, sample_weight=train_weight, offset=train_offset)
    session = EditorSession.from_model(
        model,
        terms=["x", "z"],
        train_data=(X_train, y_train, train_weight, train_offset),
    )
    session.select_indices("x", [0])
    session.shift("x", 0.2)

    refit = session.refit_with_edited_offset("x", X=X_val, y=y_val, method="fit")

    expected_offset = session.edited_offset("x", X=X_val)
    np.testing.assert_allclose(refit._fit_offset, expected_offset)
    np.testing.assert_allclose(refit._fit_weights, np.ones(n_val))
    assert refit._fit_sample_weight_ref is None


def test_refit_rejects_partial_explicit_data(editor_model, editor_frame):
    X, _ = editor_frame
    session = EditorSession.from_model(editor_model, terms=["region", "x_spline"])
    session.select_levels("region", ["B"])
    session.shift("region", 0.4)

    with pytest.raises(ValueError, match="Explicit refit data requires both X and y"):
        session.refit_with_edited_offset("region", X=X)


def test_collapse_selected_categorical_levels_refits_copied_model(editor_model, editor_frame):
    X, _ = editor_frame
    session = EditorSession.from_model(editor_model, terms=["region", "x_spline"])
    session.select_levels("region", ["B", "C"])

    refit = session.refit_with_collapsed_levels("region", method="fit")

    assert refit is not editor_model
    assert set(refit.features) == set(editor_model.features)
    assert getattr(editor_model.features["region"], "_grouping", None) is None
    grouping = refit.features["region"]._grouping
    assert grouping.original_to_group["A"] == "A"
    assert grouping.original_to_group["B"] == "B+C"
    assert grouping.original_to_group["C"] == "B+C"
    assert refit._editor_level_collapse == {
        "format": "superglm.editor.level_collapse.v1",
        "term": "region",
        "group_label": "B+C",
        "levels": ["B", "C"],
        "method": "fit",
        "message": "Selected categorical levels were collapsed and the full model was refit.",
    }
    assert refit._predict_eta_exact(X).shape == editor_model._predict_eta_exact(X).shape
    inference = refit.term_inference("region", with_se=True)
    assert inference.se_log_relativity is not None


def test_collapse_selected_ordered_categorical_levels_requires_contiguous_selection(editor_model):
    session = EditorSession.from_model(editor_model, terms=["band"])
    session.select_levels("band", ["low", "high"])

    with pytest.raises(ValueError, match="contiguous"):
        session.refit_with_collapsed_levels("band", method="fit")

    session.select_levels("band", ["medium", "high"])
    refit = session.refit_with_collapsed_levels("band", method="fit")

    grouping = refit.features["band"]._grouping
    assert grouping.original_to_group["low"] == "low"
    assert grouping.original_to_group["medium"] == "medium+high"
    assert grouping.original_to_group["high"] == "medium+high"


def test_auto_level_refits_use_reml_when_source_model_was_reml_fit():
    rng = np.random.default_rng(20260711)
    n = 700
    age = rng.uniform(18.0, 80.0, n)
    territory = rng.choice(["T01", "T02", "T03", "T04"], n, p=[0.10, 0.45, 0.10, 0.35])
    territory_effect = {"T01": 0.08, "T02": 0.0, "T03": 0.10, "T04": -0.06}
    X = pd.DataFrame({"age": age, "territory": territory})
    y = (
        0.8
        + 0.18 * np.sin(age / 8.0)
        + np.array([territory_effect[level] for level in territory])
        + rng.normal(0.0, 0.05, n)
    )
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={
            "age": Spline(k=8),
            "territory": Categorical(base="most_exposed"),
        },
    )
    model.fit_reml(X, y)
    session = EditorSession.from_model(model, terms=["age", "territory"], train_data=(X, y, None))

    session.select_levels("territory", ["T01", "T03"])
    collapsed = session.replace_with_collapsed_levels("territory", method="auto")

    assert collapsed._last_fit_meta["method"] == "fit_reml"
    assert collapsed._editor_level_collapse["method"] == "fit_reml"
    assert collapsed._reml_lambdas
    assert "age" in collapsed._reml_lambdas


def test_collapse_levels_replaces_in_force_model_and_clears_manual_edits(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline", "region"])
    session.select_indices("x_spline", [10, 11, 12])
    session.shift("x_spline", 0.4)
    assert session.edited_terms() == ["x_spline"]

    session.select_levels("region", ["B", "C"])
    refit = session.replace_with_collapsed_levels("region", method="fit")

    assert session.reference_model is editor_model
    assert session.model is refit
    assert session.edited_terms() == []
    assert session.history == []
    assert session.selection("x_spline").size == 0
    np.testing.assert_allclose(
        session.terms["x_spline"].edited_log_effect,
        session.terms["x_spline"].original_log_effect,
    )
    grouping = session.model.features["region"]._grouping
    assert grouping.original_to_group["B"] == "B+C"
    assert grouping.original_to_group["C"] == "B+C"


def test_collapse_levels_rejects_terms_used_by_interactions():
    rng = np.random.default_rng(20260726)
    n = 180
    x = rng.normal(size=n)
    region = rng.choice(["A", "B", "C"], n)
    X = pd.DataFrame({"x": x, "region": region})
    y = 0.3 + 0.4 * x + 0.2 * (region == "B") * x + rng.normal(0.0, 0.04, n)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Numeric(), "region": Categorical(base="first")},
        interactions=[("x", "region")],
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["region"], train_data=(X, y, None))
    session.select_levels("region", ["B", "C"])

    with pytest.raises(ValueError, match="used by interaction"):
        session.replace_with_collapsed_levels("region", method="fit")


def test_collapse_levels_refits_numeric_categorical_levels():
    rng = np.random.default_rng(20260712)
    code = rng.choice([1, 2, 3, 4], 320, p=[0.35, 0.30, 0.20, 0.15])
    X = pd.DataFrame({"code": code})
    effects = {1: 0.0, 2: 0.05, 3: -0.20, 4: 0.15}
    y = 0.7 + np.array([effects[value] for value in code]) + rng.normal(0.0, 0.04, code.size)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"code": Categorical(base="most_exposed")},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["code"], train_data=(X, y, None))

    session.select_levels("code", ["1", "2"])
    collapsed = session.replace_with_collapsed_levels("code", method="fit")

    grouping = collapsed.features["code"]._grouping
    assert grouping.original_to_group["1"] == "1+2"
    assert grouping.original_to_group["2"] == "1+2"
    assert np.isfinite(collapsed.predict(X)).all()


def test_to_model_applies_manual_edit_after_level_collapse(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    session.select_levels("region", ["B", "C"])
    session.replace_with_collapsed_levels("region", method="fit")

    session.select_levels("region", ["B"])
    session.shift("region", 0.2)
    edited = session.to_model()

    mu = edited.predict(editor_model._fit_X_ref)
    assert np.isfinite(mu).all()


def test_uncollapse_levels_restores_previous_in_force_model(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region", "band"])
    original_model = session.model
    session.select_levels("region", ["B", "C"])

    collapsed = session.replace_with_collapsed_levels("region", method="fit")

    assert session.model is collapsed
    assert session.can_uncollapse_levels() is True
    assert session.model.features["region"]._grouping.original_to_group["B"] == "B+C"

    restored = session.uncollapse_levels()

    assert restored is original_model
    assert session.model is original_model
    assert session.can_uncollapse_levels() is False
    assert getattr(session.model.features["region"], "_grouping", None) is None
    assert session.history == []
    assert session.redo_stack == []
    assert session.selection("region").size == 0


def test_uncollapse_levels_rolls_back_one_collapse_at_a_time(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region", "band"])
    original_model = session.model

    session.select_levels("region", ["B", "C"])
    first_collapse = session.replace_with_collapsed_levels("region", method="fit")
    session.select_levels("band", ["medium", "high"])
    second_collapse = session.replace_with_collapsed_levels("band", method="fit")
    assert session.model is second_collapse

    restored_once = session.uncollapse_levels()

    assert restored_once is first_collapse
    assert session.model is first_collapse
    assert session.can_uncollapse_levels() is True
    assert session.model.features["region"]._grouping.original_to_group["B"] == "B+C"
    assert getattr(session.model.features["band"], "_grouping", None) is None

    restored_twice = session.uncollapse_levels()

    assert restored_twice is original_model
    assert session.model is original_model
    assert session.can_uncollapse_levels() is False


def test_ordered_categorical_ungroup_rejects_non_contiguous_remainder():
    rng = np.random.default_rng(20260713)
    levels = ["A", "B", "C", "D"]
    band = rng.choice(levels, 400, p=[0.25, 0.30, 0.25, 0.20])
    X = pd.DataFrame({"band": band})
    effects = {"A": -0.15, "B": -0.05, "C": 0.10, "D": 0.25}
    y = 0.7 + np.array([effects[value] for value in band]) + rng.normal(0.0, 0.04, band.size)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={
            "band": OrderedCategorical(
                order=levels, basis=Spline(kind="ps", n_knots=3), base="first"
            )
        },
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["band"], train_data=(X, y, None))

    session.select_levels("band", ["A", "B", "C"])
    session.replace_with_collapsed_levels("band", method="fit")
    session.select_levels("band", ["B"])

    with pytest.raises(ValueError, match="non-contiguous ordered group"):
        session.replace_with_ungrouped_levels("band", method="fit")


def test_ordered_categorical_collapse_checks_fitted_order_after_display_reorder():
    rng = np.random.default_rng(20260714)
    levels = ["A", "B", "C", "D"]
    band = rng.choice(levels, 400, p=[0.25, 0.30, 0.25, 0.20])
    X = pd.DataFrame({"band": band})
    effects = {"A": -0.15, "B": -0.05, "C": 0.10, "D": 0.25}
    y = 0.7 + np.array([effects[value] for value in band]) + rng.normal(0.0, 0.04, band.size)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={
            "band": OrderedCategorical(
                order=levels, basis=Spline(kind="ps", n_knots=3), base="first"
            )
        },
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["band"], train_data=(X, y, None))

    session._apply_level_order("band", np.asarray([0, 2, 1, 3], dtype=np.intp), persist=False)
    assert session.terms["band"].levels == ["A", "C", "B", "D"]

    session.select_levels("band", ["A", "C"])
    with pytest.raises(ValueError, match="contiguous in fitted order"):
        session.replace_with_collapsed_levels("band", method="fit")


def test_repeated_categorical_collapses_create_distinct_level_groups():
    rng = np.random.default_rng(20260705)
    levels = [f"T{i:02d}" for i in range(1, 11)]
    probs = np.array([0.04, 0.26, 0.04, 0.16, 0.12, 0.10, 0.09, 0.03, 0.13, 0.03])
    territory = rng.choice(levels, 900, p=probs / probs.sum())
    effects = {
        "T01": 0.12,
        "T02": 0.10,
        "T03": 0.13,
        "T04": -0.08,
        "T05": 0.20,
        "T06": -0.12,
        "T07": 0.02,
        "T08": -0.18,
        "T09": -0.20,
        "T10": -0.17,
    }
    X = pd.DataFrame({"territory": territory})
    y = 0.5 + np.array([effects[value] for value in territory]) + rng.normal(0.0, 0.05, 900)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"territory": Categorical(base="most_exposed")},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["territory"], train_data=(X, y, None))

    session.select_levels("territory", ["T01", "T03"])
    session.replace_with_collapsed_levels("territory", method="fit")
    session.select_levels("territory", ["T08", "T10"])
    session.replace_with_collapsed_levels("territory", method="fit")

    grouping = session.model.features["territory"]._grouping
    assert grouping.original_to_group["T01"] == "T01+T03"
    assert grouping.original_to_group["T03"] == "T01+T03"
    assert grouping.original_to_group["T08"] == "T08+T10"
    assert grouping.original_to_group["T10"] == "T08+T10"
    from superglm.editor.payloads import session_payload

    groups = session_payload(session)["territory"]["level_groups"]
    assert groups == [
        {"label": "T01+T03", "indices": [0, 2], "levels": ["T01", "T03"]},
        {"label": "T08+T10", "indices": [7, 9], "levels": ["T08", "T10"]},
    ]


def test_compact_summary_shows_collapsed_reference_level_group():
    from superglm.editor.summaries import summary_payload

    rng = np.random.default_rng(20260709)
    levels = [f"T{i:02d}" for i in range(1, 11)]
    territory = np.repeat(levels, 40)
    X = pd.DataFrame({"territory": territory})
    effects = {
        "T01": 0.02,
        "T02": 0.03,
        "T03": 0.12,
        "T04": 0.13,
        "T05": 0.11,
        "T06": -0.04,
        "T07": -0.03,
        "T08": -0.05,
        "T09": -0.15,
        "T10": 0.05,
    }
    y = 0.5 + np.array([effects[value] for value in territory]) + rng.normal(0.0, 0.02, len(X))
    sample_weight = np.ones(len(X), dtype=np.float64)
    sample_weight[np.isin(territory, ["T03", "T04", "T05"])] = 3.0
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"territory": Categorical(base="most_exposed")},
    )
    model.fit(X, y, sample_weight=sample_weight)
    session = EditorSession.from_model(
        model,
        terms=["territory"],
        train_data=(X, y, sample_weight),
    )

    session.select_levels("territory", ["T03", "T04", "T05"])
    session.replace_with_collapsed_levels("territory", method="fit")
    session.select_levels("territory", ["T06", "T07", "T08"])
    session.replace_with_collapsed_levels("territory", method="fit")
    assert session.model.features["territory"]._base_level == "T03+T04+T05"

    widget = session.widget()
    try:
        payload = summary_payload(widget, "in_force")
        grouped_payload = summary_payload(widget, "in_force", level_display="grouped")
    finally:
        widget.close()

    rows = payload["compact"]["rows"]
    rows_by_name = {row["name"]: row for row in rows}
    for level in ["T03", "T04", "T05"]:
        reference_row = rows_by_name[f"territory[{level}]"]
        assert reference_row["kind"] == "reference"
        assert reference_row["level_group"] == "G1"
        assert reference_row["coef"] == 0.0
        assert reference_row["se_label"] == "ref"
    shared_rows = [rows_by_name[f"territory[{level}]"] for level in ["T06", "T07", "T08"]]
    assert {row["level_group"] for row in shared_rows} == {"G2"}
    assert len({row["coef"] for row in shared_rows}) == 1
    assert len({row["se"] for row in shared_rows}) == 1
    assert payload["level_display"] == "expanded"
    assert payload["compact"]["level_display"] == "expanded"
    assert payload["compact"]["has_level_groups"] is True

    grouped_rows = grouped_payload["compact"]["rows"]
    grouped_reference = next(row for row in grouped_rows if row["level_group"] == "G1")
    assert grouped_reference["name"] == "territory"
    assert grouped_reference["kind"] == "reference"
    assert grouped_payload["compact"]["level_groups"] == [
        {
            "feature": "territory",
            "group_id": "G1",
            "members": ["T03", "T04", "T05"],
        },
        {
            "feature": "territory",
            "group_id": "G2",
            "members": ["T06", "T07", "T08"],
        },
    ]


def test_compact_summary_shows_regular_reference_level(editor_model):
    from superglm.editor.summaries import summary_payload

    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    try:
        payload = summary_payload(widget, "in_force")
    finally:
        widget.close()

    rows = payload["compact"]["rows"]
    reference_row = next(row for row in rows if row["name"] == "region[A]")
    assert reference_row["kind"] == "reference"
    assert reference_row["coef"] == 0.0
    assert reference_row["se_label"] == "ref"
    assert reference_row["sig_class"] == "sig-reference"


def test_compact_summary_carries_tweedie_ci_status_and_cached_interval():
    from superglm.editor.summaries import _compact_summary_payload

    summary = SimpleNamespace(
        _info={
            "family": "Tweedie",
            "tweedie_p": 1.55,
            "tweedie_p_ci": (1.4, 1.7),
            "tweedie_p_ci_status": "available",
            "tweedie_p_method": "Profile MLE (Brent)",
        },
        _coef_rows=[],
    )

    payload = _compact_summary_payload(summary, "in_force", [], [])

    assert payload["model"]["tweedie_p_ci"] == [1.4, 1.7]
    assert payload["model"]["tweedie_p_ci_status"] == "available"
    assert payload["model"]["tweedie_p_method"] == "Profile MLE (Brent)"


def test_collapse_selected_levels_across_existing_groups_preserves_remainders():
    rng = np.random.default_rng(20260707)
    levels = [f"T{i:02d}" for i in range(1, 7)]
    territory = rng.choice(levels, 720, p=[0.10, 0.25, 0.08, 0.09, 0.24, 0.24])
    effects = {"T01": 0.10, "T02": 0.11, "T03": 0.12, "T04": -0.18, "T05": -0.17, "T06": -0.16}
    X = pd.DataFrame({"territory": territory})
    y = 0.5 + np.array([effects[value] for value in territory]) + rng.normal(0.0, 0.05, 720)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"territory": Categorical(base="most_exposed")},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["territory"], train_data=(X, y, None))

    session.select_levels("territory", ["T01", "T02", "T03"])
    session.replace_with_collapsed_levels("territory", method="fit")
    session.select_levels("territory", ["T04", "T05", "T06"])
    session.replace_with_collapsed_levels("territory", method="fit")
    session.select_levels("territory", ["T03", "T04"])
    refit = session.replace_with_collapsed_levels("territory", method="fit")

    grouping = refit.features["territory"]._grouping
    assert grouping.original_to_group["T01"] == "T01+T02"
    assert grouping.original_to_group["T02"] == "T01+T02"
    assert grouping.original_to_group["T03"] == "T03+T04"
    assert grouping.original_to_group["T04"] == "T03+T04"
    assert grouping.original_to_group["T05"] == "T05+T06"
    assert grouping.original_to_group["T06"] == "T05+T06"


def _pinned_categorical_model():
    """A fit whose declared universe carries one level the data never shows."""
    rng = np.random.default_rng(20260811)
    levels = ["T01", "T02", "T03"]
    territory = rng.choice(levels, 600, p=[0.40, 0.35, 0.25])
    effects = {"T01": 0.10, "T02": -0.05, "T03": 0.12}
    X = pd.DataFrame({"territory": territory})
    y = 0.5 + np.array([effects[value] for value in territory]) + rng.normal(0.0, 0.05, 600)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"territory": Categorical(base="first", levels=[*levels, "GHOST"])},
    )
    with pytest.warns(UserWarning, match="pinned to base"):
        model.fit(X, y)
    spec = model._specs["territory"]
    assert spec._pinned_levels == ["GHOST"], "precondition: exactly one level is pinned"
    assert spec._non_base == ["T02", "T03"], "precondition: the pin owns no column"
    return model, X, y


@pytest.mark.parametrize("level", ["T02", "GHOST"])
def test_editing_a_categorical_carrying_a_pinned_level_is_refused_by_name(level):
    # A pinned level is a declared level with no effective training rows.
    # `reconstruct()` reports it at relativity 1.0 -- correctly, it contributes
    # nothing -- and the editor draws a handle from exactly that table, so it
    # looks as editable as any level sitting at the base. It is not: `_non_base`
    # excludes it, so `_apply_categorical_term` writes a block with no slot for
    # it and the edit vanishes with no error at all, which the user reads back
    # as applied. Parametrized over both kinds of selection because the term is
    # patched as one block: editing a FITTED level is refused too, so the
    # refusal is the term's, not the selection's.
    model = _pinned_categorical_model()[0]
    session = EditorSession.from_model(model, terms=["territory"])
    session.select_levels("territory", [level])
    session.shift("territory", 0.3)

    with pytest.raises(ValueError) as excinfo:
        session.to_model()

    message = str(excinfo.value)
    assert "GHOST" in message, "the refusal must name the pinned level"
    assert "territory" in message
    assert "no fitted coefficient" in message
    # ...and both remedies, because either one makes the term editable again.
    assert "levels=" in message
    assert "refit" in message.lower()


def test_a_pinned_term_does_not_block_editing_a_different_term():
    # Only CHANGED terms are applied, so the refusal has to be the pinned
    # TERM's, not the model's: a model carrying one pinned categorical must
    # still be editable everywhere else, or the pin locks the whole session.
    rng = np.random.default_rng(20260811)
    territory = rng.choice(["T01", "T02", "T03"], 600, p=[0.40, 0.35, 0.25])
    x = rng.normal(0.0, 1.0, 600)
    X = pd.DataFrame({"territory": territory, "x": x})
    y = 0.5 + 0.3 * x + rng.normal(0.0, 0.05, 600)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={
            "territory": Categorical(base="first", levels=["T01", "T02", "T03", "GHOST"]),
            "x": Numeric(),
        },
    )
    with pytest.warns(UserWarning, match="pinned to base"):
        model.fit(X, y)

    session = EditorSession.from_model(model, terms=["territory", "x"], train_data=(X, y, None))
    session.select_indices("x", np.arange(session.terms["x"].size))
    session.shift("x", 0.2)
    edited = session.to_model()

    assert edited._specs["territory"]._pinned_levels == ["GHOST"]
    # ...and the untouched pinned term came through unchanged.
    np.testing.assert_allclose(
        edited._specs["territory"].reconstruct(np.zeros(2))["relativities"]["GHOST"], 1.0
    )


def test_a_categorical_with_no_pinned_level_still_edits():
    # The guard must not be reachable from an ordinary fit: every level here
    # has rows, so nothing is pinned and the edit applies as before.
    unpinned = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"territory": Categorical(base="first")},
    )
    X = pd.DataFrame({"territory": np.repeat(["T01", "T02", "T03"], 200)})
    rng = np.random.default_rng(20260811)
    y = 0.5 + rng.normal(0.0, 0.05, len(X))
    unpinned.fit(X, y)
    assert unpinned._specs["territory"]._pinned_levels == []

    session = EditorSession.from_model(unpinned, terms=["territory"])
    session.select_levels("territory", ["T02"])
    session.shift("territory", 0.3)
    edited = session.to_model()

    before = np.asarray(unpinned._predict_eta_exact(X))
    delta = np.asarray(edited._predict_eta_exact(X)) - before
    is_t02 = (X["territory"] == "T02").to_numpy()
    assert delta[is_t02] == pytest.approx(0.3, abs=1e-10)
    assert delta[~is_t02] == pytest.approx(0.0, abs=1e-10)


def test_regrouping_split_base_group_chooses_valid_remaining_base():
    rng = np.random.default_rng(20260722)
    levels = [f"T{i:02d}" for i in range(1, 7)]
    territory = rng.choice(levels, 720, p=[0.22, 0.20, 0.08, 0.09, 0.20, 0.21])
    effects = {"T01": 0.10, "T02": 0.11, "T03": 0.12, "T04": -0.18, "T05": -0.17, "T06": -0.16}
    X = pd.DataFrame({"territory": territory})
    y = 0.5 + np.array([effects[value] for value in territory]) + rng.normal(0.0, 0.05, 720)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"territory": Categorical(base="T01")},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["territory"], train_data=(X, y, None))

    session.select_levels("territory", ["T01", "T02", "T03"])
    first = session.replace_with_collapsed_levels("territory", method="fit")
    assert first.features["territory"].base == "T01+T02+T03"

    session.select_levels("territory", ["T03", "T04"])
    refit = session.replace_with_collapsed_levels("territory", method="fit")

    grouping = refit.features["territory"]._grouping
    assert grouping.original_to_group["T01"] == "T01+T02"
    assert grouping.original_to_group["T02"] == "T01+T02"
    assert grouping.original_to_group["T03"] == "T03+T04"
    assert grouping.original_to_group["T04"] == "T03+T04"
    assert refit.features["territory"].base == "T01+T02"
    assert refit.features["territory"]._base_level == "T01+T02"


def test_ungroup_selected_levels_removes_subset_from_collapsed_group():
    rng = np.random.default_rng(20260706)
    levels = [f"T{i:02d}" for i in range(1, 7)]
    territory = rng.choice(levels, 600, p=[0.08, 0.35, 0.08, 0.22, 0.17, 0.10])
    effects = {"T01": 0.11, "T02": 0.10, "T03": 0.13, "T04": -0.08, "T05": -0.12, "T06": 0.02}
    X = pd.DataFrame({"territory": territory})
    y = 0.5 + np.array([effects[value] for value in territory]) + rng.normal(0.0, 0.05, 600)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"territory": Categorical(base="most_exposed")},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["territory"], train_data=(X, y, None))

    session.select_levels("territory", ["T01", "T02", "T03"])
    session.replace_with_collapsed_levels("territory", method="fit")
    session.select_levels("territory", ["T02"])
    refit = session.replace_with_ungrouped_levels("territory", method="fit")

    grouping = refit.features["territory"]._grouping
    assert grouping.original_to_group["T01"] == "T01+T03"
    assert grouping.original_to_group["T03"] == "T01+T03"
    assert grouping.original_to_group["T02"] == "T02"
    from superglm.editor.payloads import session_payload

    groups = session_payload(session)["territory"]["level_groups"]
    assert groups == [{"label": "T01+T03", "indices": [0, 2], "levels": ["T01", "T03"]}]


def test_ungroup_preserves_symbolic_base_policy_to_avoid_display_uplift():
    from superglm.editor.payloads import session_payload

    rng = np.random.default_rng(20260708)
    levels = [f"T{i:02d}" for i in range(1, 7)]
    territory = rng.choice(levels, 900, p=[0.05, 0.34, 0.07, 0.18, 0.31, 0.05])
    effects = {"T01": 0.10, "T02": 0.12, "T03": -0.16, "T04": 0.00, "T05": 0.22, "T06": -0.08}
    X = pd.DataFrame({"territory": territory})
    y = 0.8 + np.array([effects[value] for value in territory]) + rng.normal(0.0, 0.05, 900)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"territory": Categorical(base="most_exposed")},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["territory"], train_data=(X, y, None))

    session.select_levels("territory", ["T01", "T03"])
    session.replace_with_collapsed_levels("territory", method="fit")
    session.select_levels("territory", ["T03"])
    refit = session.replace_with_ungrouped_levels("territory", method="fit")

    assert refit.features["territory"].base == "most_exposed"
    assert refit.features["territory"]._base_level == model.features["territory"]._base_level
    payload = session_payload(session)["territory"]
    assert payload["impact"]["weighted_mean_relativity"] == pytest.approx(1.0, abs=0.03)
    assert payload["impact"]["max_abs_link_delta"] < 0.08


def test_ungroup_last_collapsed_levels_removes_identity_grouping():
    from superglm.editor.payloads import session_payload

    rng = np.random.default_rng(20260709)
    levels = [f"T{i:02d}" for i in range(1, 5)]
    territory = rng.choice(levels, 500, p=[0.20, 0.35, 0.25, 0.20])
    X = pd.DataFrame({"territory": territory})
    y = 0.6 + (territory == "T03") * 0.2 + rng.normal(0.0, 0.05, 500)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"territory": Categorical(base="most_exposed")},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["territory"], train_data=(X, y, None))

    session.select_levels("territory", ["T01", "T04"])
    session.replace_with_collapsed_levels("territory", method="fit")
    session.select_levels("territory", ["T01", "T04"])
    restored = session.replace_with_ungrouped_levels("territory", method="fit")

    assert restored is model
    assert session.model is model
    assert getattr(restored.features["territory"], "_grouping", None) is None
    payload = session_payload(session)["territory"]
    assert payload["level_groups"] == []
    np.testing.assert_allclose(payload["y"], payload["original_y"])
    assert payload["impact"]["weighted_mean_relativity"] == pytest.approx(1.0)


def test_ungroup_pre_collapsed_model_refits_without_history():
    rng = np.random.default_rng(20260724)
    levels = [f"T{i:02d}" for i in range(1, 5)]
    territory = rng.choice(levels, 500, p=[0.20, 0.35, 0.25, 0.20])
    X = pd.DataFrame({"territory": territory})
    y = 0.6 + (territory == "T03") * 0.2 + rng.normal(0.0, 0.05, 500)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"territory": Categorical(base="most_exposed")},
    )
    model.fit(X, y)
    setup = EditorSession.from_model(model, terms=["territory"], train_data=(X, y, None))
    setup.select_levels("territory", ["T01", "T04"])
    collapsed = setup.replace_with_collapsed_levels("territory", method="fit")

    session = EditorSession.from_model(collapsed, terms=["territory"], train_data=(X, y, None))
    session.select_levels("territory", ["T01", "T04"])
    refit = session.replace_with_ungrouped_levels("territory", method="fit")

    assert refit is not collapsed
    assert getattr(refit.features["territory"], "_grouping", None) is None
    assert getattr(session.model.features["territory"], "_grouping", None) is None


def test_ordered_integer_ungroup_pre_collapsed_model_refits_without_history():
    rng = np.random.default_rng(20260803)
    levels = [1, 2, 3, 4]
    band = rng.choice(levels, 500, p=[0.20, 0.35, 0.25, 0.20])
    X = pd.DataFrame({"band": band})
    effects = {1: -0.15, 2: -0.05, 3: 0.10, 4: 0.25}
    y = 0.7 + np.array([effects[value] for value in band]) + rng.normal(0.0, 0.04, band.size)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={
            "band": OrderedCategorical(
                order=levels, basis=Spline(kind="ps", n_knots=3), base="first"
            )
        },
    )
    model.fit(X, y)
    setup = EditorSession.from_model(model, terms=["band"], train_data=(X, y, None))
    setup.select_levels("band", ["1", "2"])
    collapsed = setup.replace_with_collapsed_levels("band", method="fit")

    session = EditorSession.from_model(collapsed, terms=["band"], train_data=(X, y, None))
    session.select_levels("band", ["1", "2"])
    refit = session.replace_with_ungrouped_levels("band", method="fit")

    assert refit is not collapsed
    assert getattr(refit.features["band"], "_grouping", None) is None
    assert getattr(session.model.features["band"], "_grouping", None) is None
    assert np.isfinite(refit.predict(X)).all()


def test_ungroup_last_collapsed_levels_restores_pre_collapse_in_force_model(editor_model):
    profiled_model = editor_model._clone_without_features(set())
    profiled_model.fit(editor_model._fit_X_ref, editor_model._fit_y_ref)
    profiled_model._editor_profile_marker = "kept"
    session = EditorSession.from_model(editor_model, terms=["region"])
    session.replace_in_force_model(profiled_model)

    session.select_levels("region", ["B", "C"])
    session.replace_with_collapsed_levels("region", method="fit")
    session.select_levels("region", ["B", "C"])

    restored = session.replace_with_ungrouped_levels("region", method="fit")

    assert restored is profiled_model
    assert session.model is profiled_model
    assert session.model._editor_profile_marker == "kept"
    assert session.can_uncollapse_levels() is False


def test_final_ungroup_after_partial_ungroup_does_not_restore_stale_history():
    rng = np.random.default_rng(20260801)
    levels = [f"T{i:02d}" for i in range(1, 5)]
    territory = rng.choice(levels, 520, p=[0.25, 0.25, 0.25, 0.25])
    effects = {"T01": 0.12, "T02": 0.10, "T03": -0.08, "T04": -0.10}
    X = pd.DataFrame({"territory": territory})
    y = 0.5 + np.array([effects[value] for value in territory]) + rng.normal(0.0, 0.05, 520)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"territory": Categorical(base="most_exposed")},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["territory"], train_data=(X, y, None))

    session.select_levels("territory", ["T01", "T02"])
    session.replace_with_collapsed_levels("territory", method="fit")
    session.select_levels("territory", ["T03", "T04"])
    session.replace_with_collapsed_levels("territory", method="fit")

    session.select_levels("territory", ["T01", "T02"])
    partial = session.replace_with_ungrouped_levels("territory", method="fit")
    assert partial.features["territory"]._grouping.original_to_group["T03"] == "T03+T04"

    session.select_levels("territory", ["T03", "T04"])
    final = session.replace_with_ungrouped_levels("territory", method="fit")

    assert final is not partial
    assert getattr(final.features["territory"], "_grouping", None) is None
    assert getattr(session.model.features["territory"], "_grouping", None) is None
    assert session.can_uncollapse_levels() is False


def test_reorder_categorical_levels_is_display_only(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    term = session.terms["region"]
    original_values = dict(zip(term.levels, term.edited_log_effect, strict=False))

    session.select_levels("region", ["C"])
    session.reorder_levels("region", target_index=1)

    assert term.levels == ["A", "C", "B"]
    assert term.x.tolist() == [0.0, 1.0, 2.0]
    assert session.selection("region").tolist() == [1]
    assert dict(zip(term.levels, term.edited_log_effect, strict=False)) == pytest.approx(
        original_values
    )
    assert session.edited_terms() == []


def test_reorder_ordered_categorical_levels_is_rejected(editor_model):
    session = EditorSession.from_model(editor_model, terms=["band"])
    session.select_levels("band", ["medium"])

    with pytest.raises(TypeError, match="cannot be display-reordered"):
        session.reorder_levels("band", target_index=0)

    session._level_orders["band"] = ["medium", "low", "high"]
    session.reset_level_order("band")
    assert session.level_order_changed("band") is False
    assert "band" not in session._level_orders


def test_reorder_categorical_levels_keeps_native_metadata_aligned(editor_model):
    from superglm.editor.terms import native_log_effect_values

    session = EditorSession.from_model(
        editor_model,
        terms=["region"],
        centering="mean",
        with_se=False,
    )
    term = session.terms["region"]
    original_native = dict(zip(term.levels, native_log_effect_values(term), strict=False))

    session.select_levels("region", ["C"])
    session.reorder_levels("region", target_index=1)

    reordered_native = dict(zip(term.levels, native_log_effect_values(term), strict=False))
    assert reordered_native == pytest.approx(original_native)


def test_reorder_multiple_categorical_levels_moves_selection_as_contiguous_block():
    values = np.array([0.1, 0.2, 0.3, 0.4], dtype=np.float64)
    session = EditorSession(
        model=None,
        terms={
            "territory": EditableTerm(
                name="territory",
                kind="categorical",
                levels=["A", "B", "C", "D"],
                x=np.arange(4, dtype=np.float64),
                original_log_effect=values.copy(),
                edited_log_effect=values.copy(),
            )
        },
    )
    session.select_levels("territory", ["A", "C"])

    session.reorder_levels("territory", target_index=2)

    assert session.terms["territory"].levels == ["B", "A", "C", "D"]
    assert session.selection("territory").tolist() == [1, 2]


def test_reorder_multiple_categorical_levels_can_move_to_far_right():
    values = np.array([0.1, 0.2, 0.3, 0.4, 0.5], dtype=np.float64)
    session = EditorSession(
        model=None,
        terms={
            "territory": EditableTerm(
                name="territory",
                kind="categorical",
                levels=["A", "B", "C", "D", "E"],
                x=np.arange(5, dtype=np.float64),
                original_log_effect=values.copy(),
                edited_log_effect=values.copy(),
            )
        },
    )
    session.select_levels("territory", ["A", "C"])

    session.reorder_levels("territory", target_index=5)

    assert session.terms["territory"].levels == ["B", "D", "E", "A", "C"]
    assert session.selection("territory").tolist() == [3, 4]


def test_save_load_preserves_display_only_level_order(editor_model, tmp_path):
    session = EditorSession.from_model(editor_model, terms=["region"])
    session.select_levels("region", ["C"])
    session.reorder_levels("region", target_index=1)
    path = tmp_path / "editor-session.json"

    session.save(path)
    loaded = EditorSession.load(path, model=editor_model)

    assert loaded.terms["region"].levels == ["A", "C", "B"]
    assert loaded.selection("region").tolist() == [1]
    assert loaded.level_order_changed("region")


def test_profile_options_forward_search_fit_mode():
    """A client asking for the decoupled search must not be silently coupled."""
    from superglm.editor.server import _profile_options

    options = _profile_options(
        {
            "parameter": "tweedie_p",
            "fit_mode": "reml",
            "search_fit_mode": "fit",
            "junk": "dropped",
        }
    )

    assert options["search_fit_mode"] == "fit"
    assert options["fit_mode"] == "reml"
    assert "junk" not in options


def test_profile_options_filter_tweedie_only_keys_for_theta():
    """search_fit_mode is Tweedie-only: estimate_nb_theta has no decoupled
    search, so a client that retains options while switching parameters must
    not crash the theta profile with an unexpected keyword. Unknown or empty
    parameters strip it too -- fail safe, never fail loud downstream."""
    from superglm.editor.server import _profile_options

    for parameter in ("nb2_theta", "theta", "nb2", ""):
        options = _profile_options(
            {"parameter": parameter, "fit_mode": "reml", "search_fit_mode": "fit"}
        )
        assert "search_fit_mode" not in options
        assert options["fit_mode"] == "reml"

    for parameter in ("tweedie", "tweedie_p", "p"):
        options = _profile_options(
            {"parameter": parameter, "fit_mode": "reml", "search_fit_mode": "fit"}
        )
        assert options["search_fit_mode"] == "fit"


def test_reprofile_distribution_parameter_dispatches_to_model(
    editor_model, editor_frame, monkeypatch
):
    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(X, y, None),
    )
    calls: list[dict[str, object]] = []

    replaced: list[object] = []

    def fake_estimate_p(self, X_arg, y_arg, sample_weight=None, offset=None, **kwargs):
        calls.append(
            {
                "parameter": "p",
                "model": self,
                "X": X_arg,
                "y": y_arg,
                "sample_weight": sample_weight,
                "offset": offset,
                "kwargs": kwargs,
            }
        )
        return "p-result"

    def fake_estimate_theta(self, X_arg, y_arg, sample_weight=None, offset=None, **kwargs):
        calls.append(
            {
                "parameter": "theta",
                "model": self,
                "X": X_arg,
                "y": y_arg,
                "sample_weight": sample_weight,
                "offset": offset,
                "kwargs": kwargs,
            }
        )
        return "theta-result"

    def fake_replace(model):
        replaced.append(model)
        return session

    monkeypatch.setattr(type(editor_model), "estimate_p", fake_estimate_p)
    monkeypatch.setattr(type(editor_model), "estimate_theta", fake_estimate_theta)
    monkeypatch.setattr(session, "replace_in_force_model", fake_replace)

    p_result = session.reprofile_distribution(
        "tweedie_p",
        method="grid",
        n_grid=7,
        phi_method="pearson",
    )
    theta_result = session.reprofile_distribution("nb2_theta", theta_bounds=(0.2, 20.0))

    assert p_result == "p-result"
    assert theta_result == "theta-result"
    assert calls[0]["parameter"] == "p"
    assert calls[0]["model"] is not editor_model
    assert calls[0]["X"] is X
    assert calls[0]["y"] is y
    assert calls[0]["kwargs"] == {
        "method": "grid",
        "n_grid": 7,
        "phi_method": "pearson",
        "fit_mode": "inherit",
    }
    assert calls[1]["parameter"] == "theta"
    assert calls[1]["model"] is not editor_model
    assert calls[1]["kwargs"] == {"theta_bounds": (0.2, 20.0), "fit_mode": "inherit"}
    assert replaced == [calls[0]["model"], calls[1]["model"]]


def test_reprofile_distribution_uses_cloned_in_force_model():
    rng = np.random.default_rng(20260711)
    n = 120
    x = rng.uniform(0.0, 1.0, n)
    X = pd.DataFrame({"x": x})
    mu = np.exp(0.2 + 0.4 * x)
    y = generate_tweedie_cpg(n, mu=mu, phi=0.5, p=1.45, rng=rng)
    model = SuperGLM(
        family=families.tweedie(p=1.3),
        selection_penalty=0.0,
        spline_penalty=0.1,
        features={"x": Spline(n_knots=5)},
    )
    model.fit(X, y)
    original_beta = model.result.beta.copy()
    original_intercept = model.result.intercept
    session = EditorSession.from_model(model, terms=["x"], train_data=(X, y, None))

    session.reprofile_distribution(
        "tweedie_p",
        fit_mode="fit",
        phi_method="pearson",
        method="grid",
        grid=np.array([1.25, 1.45]),
    )

    assert session.model is not model
    assert session.reference_model is model
    assert model.family.p == 1.3
    np.testing.assert_allclose(model.result.beta, original_beta)
    assert model.result.intercept == original_intercept


def test_reprofile_distribution_clears_collapse_history(editor_model, editor_frame, monkeypatch):
    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["region"],
        train_data=(X, y, None),
    )
    session.select_levels("region", ["B", "C"])
    session.replace_with_collapsed_levels("region", method="fit")
    assert session.can_uncollapse_levels()

    def fake_estimate_p(self, X_arg, y_arg, sample_weight=None, offset=None, **kwargs):
        self.fit(X_arg, y_arg, sample_weight=sample_weight, offset=offset)
        self._editor_profile_marker = "profiled"
        return "p-result"

    monkeypatch.setattr(type(editor_model), "estimate_p", fake_estimate_p)

    result = session.reprofile_distribution(
        "tweedie_p",
        method="grid",
        grid=np.array([1.45]),
        phi_method="pearson",
    )

    assert result == "p-result"
    assert session.model._editor_profile_marker == "profiled"
    assert session.can_uncollapse_levels() is False


def test_reprofile_distribution_rejects_pending_manual_edits(editor_model, editor_frame):
    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(X, y, None),
    )
    session.select_indices("x_spline", [0, 1])
    session.shift("x_spline", 0.1)

    with pytest.raises(RuntimeError, match="manual coefficient edits"):
        session.reprofile_distribution(
            "tweedie_p",
            method="grid",
            grid=np.array([1.25, 1.45]),
            phi_method="pearson",
        )


def test_tweedie_profile_trace_can_include_candidate_fit_curves():
    rng = np.random.default_rng(20260706)
    n = 120
    x = rng.uniform(0.0, 1.0, n)
    X = pd.DataFrame({"x": x})
    mu = np.exp(0.2 + 0.4 * np.sin(2 * np.pi * x))
    y = generate_tweedie_cpg(n, mu=mu, phi=0.5, p=1.45, rng=rng)
    model = SuperGLM(
        family=families.tweedie(p=1.3),
        selection_penalty=0.0,
        spline_penalty=0.1,
        features={"x": Spline(n_knots=5)},
    )

    result = model.estimate_p(
        X,
        y,
        fit_mode="fit",
        phi_method="pearson",
        method="grid",
        grid=np.array([1.25, 1.45, 1.65]),
        trace_iterations=True,
    )

    assert "fit_trace" in result.search_trace.columns
    traces = result.search_trace["fit_trace"].tolist()
    assert all(isinstance(trace, list) for trace in traces)
    assert any(len(trace) >= 1 for trace in traces)
    first_trace = next(trace for trace in traces if trace)
    assert {"iteration", "loss"}.issubset(first_trace[0])


def test_tweedie_reml_profile_trace_can_include_candidate_objective_curves():
    rng = np.random.default_rng(20260707)
    n = 120
    x = rng.uniform(0.0, 1.0, n)
    X = pd.DataFrame({"x": x})
    mu = np.exp(0.2 + 0.4 * np.sin(2 * np.pi * x))
    y = generate_tweedie_cpg(n, mu=mu, phi=0.5, p=1.45, rng=rng)
    model = SuperGLM(
        family=families.tweedie(p=1.3),
        selection_penalty=0.0,
        spline_penalty=0.1,
        features={"x": Spline(n_knots=5)},
    )
    model.fit_reml(X, y, max_reml_iter=5)

    result = model.estimate_p(
        X,
        y,
        fit_mode="inherit",
        phi_method="pearson",
        method="grid",
        grid=np.array([1.25, 1.45]),
        trace_iterations=True,
    )

    traces = result.search_trace["fit_trace"].tolist()
    assert all(isinstance(trace, list) for trace in traces)
    assert any(len(trace) >= 1 for trace in traces)
    assert set(result.search_trace["fit_trace_kind"]) == {"REML objective"}


def test_widget_profile_distribution_forwards_options_and_returns_trace(
    editor_model, editor_frame, monkeypatch
):
    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(X, y, None),
    )
    widget = session.widget()
    calls: list[dict[str, object]] = []
    summary_calls: list[dict[str, str]] = []

    class FakeTrace:
        def to_dict(self, orient):
            assert orient == "records"
            return [{"step": 0, "p": 1.42, "phi": 0.3, "nll": 0.12, "source": "brent"}]

    class FakeResult:
        p_hat = 1.42
        phi_hat = 0.3
        nll = 0.12
        method = "brent"
        phi_method = "mle"
        density_exact = True
        _ci_cache = {}
        search_trace = FakeTrace()

        def ci(self, alpha=0.05):
            raise AssertionError("editor payload must not evaluate a Tweedie profile CI")

        ci_details = ci

    def fake_reprofile(parameter, **kwargs):
        calls.append({"parameter": parameter, "kwargs": kwargs})
        return FakeResult()

    def fake_summary(_widget, source, *, level_display="expanded"):
        summary_calls.append({"source": source, "level_display": level_display})
        return {
            "available": True,
            "source": source,
            "level_display": level_display,
            "compact": {"model": {}},
        }

    monkeypatch.setattr(session, "reprofile_distribution", fake_reprofile)
    monkeypatch.setattr("superglm.editor.widget.summary_payload", fake_summary)

    try:
        payload = widget._profile_distribution(
            "tweedie_p",
            level_display="grouped",
            method="brent",
            phi_method="mle",
            xatol=0.002,
        )
    finally:
        widget.close()

    assert calls == [
        {
            "parameter": "tweedie_p",
            "kwargs": {"method": "brent", "phi_method": "mle", "xatol": 0.002},
        }
    ]
    assert summary_calls == [{"source": "in_force", "level_display": "grouped"}]
    assert payload["level_display"] == "grouped"
    assert payload["profile_trace"] == [
        {"step": 0, "p": 1.42, "phi": 0.3, "nll": 0.12, "source": "brent"}
    ]
    assert payload["profile_estimate"] == {
        "parameter": "p",
        "label": "p_hat",
        "value": 1.42,
        "ci_low": None,
        "ci_high": None,
        "ci_status": "not computed",
        "objective": 0.12,
        "objective_label": "loss",
        "lower_is_better": True,
    }


@pytest.mark.parametrize(
    ("phi_method", "ci_cache", "expected_low", "expected_high", "expected_status"),
    [
        ("mle", {}, None, None, "not computed"),
        ("mle", {0.05: (1.31, 1.53)}, 1.31, 1.53, "available"),
        (
            "pearson",
            {0.05: (1.31, 1.53)},
            None,
            None,
            "unavailable for Pearson plug-in",
        ),
    ],
)
def test_tweedie_editor_profile_payload_only_reads_valid_cached_mle_ci(
    phi_method,
    ci_cache,
    expected_low,
    expected_high,
    expected_status,
):
    from superglm.editor.widget import _profile_estimate_payload

    def unexpected_ci(*args, **kwargs):
        raise AssertionError("editor payload must not evaluate a Tweedie profile CI")

    result = SimpleNamespace(
        p_hat=1.42,
        phi_hat=0.3,
        nll=0.12,
        method="brent",
        phi_method=phi_method,
        density_exact=True,
        _ci_cache=dict(ci_cache),
        ci=unexpected_ci,
        ci_details=unexpected_ci,
    )

    payload = _profile_estimate_payload(result, "tweedie_p")

    assert payload["ci_low"] == expected_low
    assert payload["ci_high"] == expected_high
    assert payload["ci_status"] == expected_status


def test_tweedie_editor_profile_payload_tolerates_legacy_missing_ci_state():
    from superglm.editor.widget import _profile_estimate_payload

    payload = _profile_estimate_payload(
        SimpleNamespace(p_hat=1.42, phi_hat=0.3, nll=0.12),
        "tweedie_p",
    )

    assert payload["ci_low"] is None
    assert payload["ci_high"] is None
    assert payload["ci_status"] == "not computed"


def test_widget_profile_distribution_job_reports_live_trace(
    editor_model, editor_frame, monkeypatch
):
    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(X, y, None),
    )
    widget = session.widget()

    class FakeTrace:
        def to_dict(self, orient):
            assert orient == "records"
            return [{"step": 1, "p": 1.5, "phi": 0.2, "nll": 0.1, "source": "final"}]

    class FakeResult:
        search_trace = FakeTrace()

    def fake_reprofile(parameter, **kwargs):
        trace_callback = kwargs["trace_callback"]
        trace_callback({"step": 0, "p": 1.3, "phi": 0.22, "nll": 0.14, "source": "brent"})
        return FakeResult()

    def fake_summary(_widget, source, *, level_display="expanded"):
        return {
            "available": True,
            "source": source,
            "level_display": level_display,
            "compact": {"model": {}},
        }

    monkeypatch.setattr(session, "reprofile_distribution", fake_reprofile)
    monkeypatch.setattr("superglm.editor.widget.summary_payload", fake_summary)

    try:
        started = widget._start_profile_distribution_job(
            "tweedie_p",
            level_display="grouped",
            method="brent",
            phi_method="mle",
            xatol=0.001,
        )
        job_id = started["job_id"]
        status = widget._profile_distribution_status(job_id, wait=True)
    finally:
        widget.close()

    assert started["status"] in {"running", "complete"}
    assert status["status"] == "complete"
    assert status["options"] == {"method": "brent", "phi_method": "mle", "xatol": 0.001}
    assert status["trace"] == [
        {"step": 0, "p": 1.3, "phi": 0.22, "nll": 0.14, "source": "brent"},
        {"step": 1, "p": 1.5, "phi": 0.2, "nll": 0.1, "source": "final"},
    ]
    assert status["result"]["source"] == "in_force"
    assert status["result"]["level_display"] == "grouped"


def test_widget_profile_distribution_job_reports_finalizing_phase(
    editor_model, editor_frame, monkeypatch
):
    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(X, y, None),
    )
    widget = session.widget()
    summary_started = threading.Event()
    release_summary = threading.Event()

    class FakeTrace:
        def to_dict(self, orient):
            assert orient == "records"
            return [{"step": 1, "p": 1.5, "phi": 0.2, "nll": 0.1, "source": "final"}]

    class FakeResult:
        search_trace = FakeTrace()

    def fake_reprofile(parameter, **kwargs):
        trace_callback = kwargs["trace_callback"]
        trace_callback({"step": 0, "p": 1.3, "phi": 0.22, "nll": 0.14, "source": "brent"})
        return FakeResult()

    def blocking_summary(_widget, source, *, level_display="expanded"):
        assert level_display == "expanded"
        summary_started.set()
        assert release_summary.wait(timeout=2.0)
        return {"available": True, "source": source, "compact": {"model": {}}}

    monkeypatch.setattr(session, "reprofile_distribution", fake_reprofile)
    monkeypatch.setattr("superglm.editor.widget.summary_payload", blocking_summary)

    try:
        started = widget._start_profile_distribution_job(
            "tweedie_p",
            method="brent",
            phi_method="mle",
            xatol=0.001,
        )
        job_id = started["job_id"]
        assert summary_started.wait(timeout=2.0)
        status = widget._profile_distribution_status(job_id)
        release_summary.set()
        complete = widget._profile_distribution_status(job_id, wait=True)
    finally:
        release_summary.set()
        widget.close()

    assert status["status"] == "running"
    assert status["phase"] == "finalizing"
    assert status["trace"] == [{"step": 0, "p": 1.3, "phi": 0.22, "nll": 0.14, "source": "brent"}]
    assert complete["status"] == "complete"
    assert complete["phase"] == "complete"


def test_widget_profile_distribution_job_reports_best_parameter_and_refit_phase(
    editor_model, editor_frame, monkeypatch
):
    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(X, y, None),
    )
    widget = session.widget()
    refit_started = threading.Event()
    release_refit = threading.Event()

    class FakeTrace:
        def to_dict(self, orient):
            assert orient == "records"
            return [{"step": 1, "p": 1.5, "phi": 0.2, "nll": 0.1, "source": "final"}]

    class FakeResult:
        p_hat = 1.5
        phi_hat = 0.2
        nll = 0.1
        method = "brent"
        phi_method = "mle"
        density_exact = True
        _ci_cache = {0.05: (1.4, 1.6)}
        search_trace = FakeTrace()

        def ci(self, alpha=0.05):
            raise AssertionError("editor payload must not evaluate a Tweedie profile CI")

        ci_details = ci

    def fake_reprofile(parameter, **kwargs):
        kwargs["trace_callback"]({"step": 0, "p": 1.3, "phi": 0.22, "nll": 0.14, "source": "brent"})
        kwargs["progress_callback"](
            "best_found",
            {
                "profile_estimate": {
                    "parameter": "p",
                    "label": "p_hat",
                    "value": 1.5,
                    "objective": 0.1,
                    "objective_label": "loss",
                    "lower_is_better": True,
                }
            },
        )
        kwargs["progress_callback"]("final_refit")
        refit_started.set()
        assert release_refit.wait(timeout=2.0)
        return FakeResult()

    monkeypatch.setattr(session, "reprofile_distribution", fake_reprofile)
    monkeypatch.setattr(
        "superglm.editor.widget.summary_payload",
        lambda _widget, source, *, level_display="expanded": {
            "available": True,
            "source": source,
            "level_display": level_display,
            "compact": {"model": {}},
        },
    )

    try:
        started = widget._start_profile_distribution_job(
            "tweedie_p",
            method="brent",
            phi_method="mle",
            xatol=0.001,
        )
        job_id = started["job_id"]
        assert refit_started.wait(timeout=2.0)
        status = widget._profile_distribution_status(job_id)
        release_refit.set()
        complete = widget._profile_distribution_status(job_id, wait=True)
    finally:
        release_refit.set()
        widget.close()

    assert status["status"] == "running"
    assert status["phase"] == "final_refit"
    assert status["profile_estimate"]["value"] == 1.5
    assert status["profile_estimate"]["ci_low"] is None
    assert status["profile_estimate"]["ci_status"] == "not computed"
    assert complete["status"] == "complete"
    assert complete["phase"] == "complete"
    assert complete["result"]["profile_estimate"]["ci_low"] == 1.4
    assert complete["result"]["profile_estimate"]["ci_high"] == 1.6
    assert complete["result"]["profile_estimate"]["ci_status"] == "available"


def test_reprofile_distribution_parameter_rejects_unknown_parameter(editor_model, editor_frame):
    X, y = editor_frame
    session = EditorSession.from_model(editor_model, terms=["x_spline"], train_data=(X, y, None))

    with pytest.raises(ValueError, match="parameter must be"):
        session.reprofile_distribution("dispersion")


def test_reordered_categorical_levels_persist_across_collapse_refit(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])

    session.select_levels("region", ["C"])
    session.reorder_levels("region", target_index=1)
    assert session.terms["region"].levels == ["A", "C", "B"]

    session.select_levels("region", ["C", "B"])
    session.replace_with_collapsed_levels("region", method="fit")

    assert session.terms["region"].levels == ["A", "C", "B"]


def test_reordered_categorical_levels_persist_after_ungroup_restores_reference_model(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    session.select_levels("region", ["C"])
    session.reorder_levels("region", target_index=1)
    assert session.terms["region"].levels == ["A", "C", "B"]

    session.select_levels("region", ["C", "B"])
    session.replace_with_collapsed_levels("region", method="fit")
    session.select_levels("region", ["C", "B"])
    session.replace_with_ungrouped_levels("region", method="fit")

    assert session.model is session.reference_model
    assert session.terms["region"].levels == ["A", "C", "B"]


def test_reset_level_order_restores_model_order_after_visual_reorder(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    session.select_levels("region", ["C"])
    session.reorder_levels("region", target_index=1)

    session.reset_level_order("region")

    assert session.terms["region"].levels == ["A", "B", "C"]
    assert session.terms["region"].x.tolist() == [0.0, 1.0, 2.0]
    assert session.selection("region").tolist() == [2]


def test_session_payload_compares_in_force_against_reference_model_after_collapse(editor_model):
    from superglm.editor.payloads import session_payload

    session = EditorSession.from_model(editor_model, terms=["region"])
    session.select_levels("region", ["B", "C"])
    session.replace_with_collapsed_levels("region", method="fit")

    payload = session_payload(session)
    region = payload["region"]
    reference = editor_model.term_inference("region", with_se=True)
    in_force = session.model.term_inference("region", with_se=True)
    export_df = session.model.plot_data("region")["terms"][0]["effect"]

    np.testing.assert_allclose(region["original_y"], reference.relativity)
    np.testing.assert_allclose(region["y"], in_force.relativity)
    expected_mean_relativity = np.average(
        np.asarray(region["y"]) / np.asarray(region["original_y"]),
        weights=np.asarray(region["weights"]),
    )
    assert region["impact"]["weighted_mean_relativity"] == pytest.approx(expected_mean_relativity)
    assert expected_mean_relativity != pytest.approx(1.0)
    assert region["levels"] == ["A", "B", "C"]
    assert region["y"][1] == pytest.approx(region["y"][2])
    assert region["level_groups"] == [{"label": "B+C", "indices": [1, 2], "levels": ["B", "C"]}]
    assert export_df["level"].tolist() == ["A", "B", "C"]
    assert export_df.loc[export_df["level"] == "B", "relativity"].item() == pytest.approx(
        export_df.loc[export_df["level"] == "C", "relativity"].item()
    )


def test_session_payload_omits_previous_y_without_term_history(editor_model):
    from superglm.editor.payloads import session_payload

    session = EditorSession.from_model(editor_model, terms=["x_spline"])

    payload = session_payload(session)["x_spline"]

    assert payload["previous_y"] is None


def test_session_payload_previous_y_uses_latest_edit_for_that_term(editor_model):
    from superglm.editor.payloads import session_payload

    session = EditorSession.from_model(editor_model, terms=["x_spline", "x_num"])
    spline = session.terms["x_spline"]
    previous_spline = spline.edited_log_effect.copy()

    session.select_indices("x_spline", [0, 1])
    session.shift("x_spline", 0.05)
    session.select_indices("x_num", [0])
    session.shift("x_num", 0.10)

    payload = session_payload(session)

    assert payload["x_spline"]["previous_y"] == pytest.approx(np.exp(previous_spline).tolist())
    assert payload["x_num"]["previous_y"] is not None


def test_history_payload_includes_stable_linear_history(editor_model):
    from superglm.editor.payloads import history_payload

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_indices("x_spline", [0, 1])
    session.shift("x_spline", 0.05)

    first = history_payload(session)["active"][0]
    second = history_payload(session)["active"][0]

    assert first["hash"] == second["hash"]
    assert len(first["hash"]) == 7
    assert first["term"] == "x_spline"
    assert first["operation"] == "shift"
    assert first["n_points"] == 2
    assert first["is_head"] is True


def test_history_payload_reports_redo_stack_after_undo(editor_model):
    from superglm.editor.payloads import history_payload

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_indices("x_spline", [0])
    session.shift("x_spline", 0.05)
    session.undo()

    history = history_payload(session)

    assert history["active"] == []
    assert len(history["redo"]) == 1
    assert history["redo"][0]["operation"] == "shift"
    assert history["redo"][0]["is_head"] is False


def test_reset_clears_term_history_without_creating_reset_edit(editor_model):
    from superglm.editor.payloads import history_payload, session_payload

    session = EditorSession.from_model(editor_model, terms=["x_spline", "x_num"])
    session.select_indices("x_spline", [0, 1])
    session.shift("x_spline", 0.05)
    session.select_indices("x_num", [0])
    session.shift("x_num", 0.10)
    session.undo("x_spline")

    session.reset("x_spline")

    history = history_payload(session)
    assert [record["term"] for record in history["active"]] == ["x_num"]
    assert history["redo"] == []
    assert session_payload(session)["x_spline"]["previous_y"] is None


def test_partial_reset_preserves_history_for_unreset_points(editor_model):
    from superglm.editor.payloads import history_payload

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    term = session.terms["x_spline"]
    original = term.original_log_effect.copy()
    session.select_indices("x_spline", [0, 1])
    session.shift("x_spline", 0.05)

    session.select_indices("x_spline", [0])
    session.reset("x_spline")

    history = history_payload(session)
    assert term.edited_log_effect[0] == pytest.approx(original[0])
    assert term.edited_log_effect[1] == pytest.approx(original[1] + 0.05)
    assert history["active"][0]["term"] == "x_spline"
    assert history["active"][0]["n_points"] == 1

    session.undo("x_spline")
    assert term.edited_log_effect[0] == pytest.approx(original[0])
    assert term.edited_log_effect[1] == pytest.approx(original[1])


def test_widget_state_includes_edit_history(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_indices("x_spline", [0])
    session.shift("x_spline", 0.05)
    widget = session.widget()

    try:
        state = widget._state()
    finally:
        widget.close()

    assert state["history"]["active"][0]["term"] == "x_spline"


def test_session_payload_adds_collapsed_group_display_for_ordered_levels(
    editor_model, editor_frame
):
    from superglm.editor.payloads import session_payload

    X, y = editor_frame
    session = EditorSession.from_model(editor_model, terms=["band"], train_data=(X, y, None))
    session.select_levels("band", ["medium", "high"])
    session.replace_with_collapsed_levels("band", method="fit")

    payload = session_payload(session)
    band = payload["band"]

    assert band["levels"] == ["low", "medium", "high"]
    assert band["level_groups"] == [
        {"label": "medium+high", "indices": [1, 2], "levels": ["medium", "high"]}
    ]
    assert band["group_display"]["available"] is True
    assert band["group_display"]["default_mode"] == "collapsed"

    collapsed = band["group_display"]["collapsed"]
    assert collapsed["levels"] == ["low", "medium+high"]
    assert collapsed["source_indices"] == [[0], [1, 2]]
    assert collapsed["source_levels"] == [["low"], ["medium", "high"]]
    assert collapsed["is_group"] == [False, True]
    assert collapsed["x"] == [0.0, 1.0]
    assert collapsed["exposure"]["kind"] == "bars"
    assert collapsed["exposure"]["y"][1] == pytest.approx(band["weights"][1] + band["weights"][2])
    assert collapsed["y"][1] == pytest.approx(band["y"][1])
    assert collapsed["y"][1] == pytest.approx(band["y"][2])


def test_session_payload_group_display_projects_previous_y_for_collapsed_levels(
    editor_model, editor_frame
):
    from superglm.editor.payloads import session_payload

    X, y = editor_frame
    session = EditorSession.from_model(editor_model, terms=["band"], train_data=(X, y, None))
    session.select_levels("band", ["medium", "high"])
    session.replace_with_collapsed_levels("band", method="fit")
    session.select_levels("band", ["medium", "high"])
    session.shift("band", 0.05)

    payload = session_payload(session)["band"]
    collapsed = payload["group_display"]["collapsed"]

    assert collapsed["previous_y"] is not None
    assert len(collapsed["previous_y"]) == len(collapsed["y"])
    assert collapsed["previous_y"][1] == pytest.approx(payload["y"][1] / np.exp(0.05))


def test_session_payload_collapsed_group_display_uses_custom_group_label(
    editor_model, editor_frame
):
    from superglm.editor.payloads import session_payload

    X, y = editor_frame
    session = EditorSession.from_model(editor_model, terms=["band"], train_data=(X, y, None))
    session.select_levels("band", ["medium", "high"])
    session.replace_with_collapsed_levels("band", method="fit", group_label="Upper")

    payload = session_payload(session)["band"]
    collapsed = payload["group_display"]["collapsed"]

    assert collapsed["levels"] == ["low", "Upper"]
    assert collapsed["group_labels"] == ["low", "Upper"]
    assert collapsed["source_levels"] == [["low"], ["medium", "high"]]


def test_session_payload_group_display_defaults_expanded_for_unordered_categorical(
    editor_model, editor_frame
):
    from superglm.editor.payloads import session_payload

    X, y = editor_frame
    session = EditorSession.from_model(editor_model, terms=["region"], train_data=(X, y, None))
    session.select_levels("region", ["B", "C"])
    session.replace_with_collapsed_levels("region", method="fit")

    region = session_payload(session)["region"]

    assert region["group_display"]["available"] is True
    assert region["group_display"]["default_mode"] == "expanded"
    assert region["group_display"]["collapsed"]["levels"] == ["A", "B+C"]
    assert region["group_display"]["collapsed"]["source_indices"] == [[0], [1, 2]]


def test_categorical_level_edit_changes_copied_predictions(editor_model, editor_frame):
    X, _ = editor_frame
    session = EditorSession.from_model(editor_model, terms=["region"])
    session.select_levels("region", ["B"])
    session.shift("region", 0.5)

    edited = session.to_model()

    mask = X["region"] == "B"
    eta_delta = edited._predict_eta_exact(X) - editor_model._predict_eta_exact(X)
    assert np.mean(eta_delta[mask]) > 0.3
    assert np.mean(np.abs(eta_delta[~mask])) < 0.1


def test_session_extracts_all_supported_1d_main_effects(editor_model):
    session = EditorSession.from_model(editor_model)

    assert list(session.terms) == ["x_spline", "x_poly", "x_num", "region", "band"]
    assert session.terms["x_poly"].x is not None
    assert session.terms["x_num"].edited_log_effect.shape == (1,)
    assert session.terms["band"].levels == ["low", "medium", "high"]


def test_session_uses_fit_row_counts_for_term_weights(editor_model, editor_frame):
    X, _ = editor_frame
    session = EditorSession.from_model(editor_model, terms=["region"])

    weights = session.terms["region"].weights
    expected = np.array([(X["region"] == level).sum() for level in ["A", "B", "C"]])

    np.testing.assert_allclose(weights, expected)


def test_session_uses_fit_sample_weight_for_term_weights():
    X = pd.DataFrame({"region": ["A", "A", "B", "B", "C", "C"], "x": np.arange(6)})
    sample_weight = np.array([1.0, 2.0, 0.5, 1.5, 3.0, 4.0])
    y = np.array([0.1, 0.0, 0.4, 0.3, -0.2, -0.1])
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"region": Categorical(base="first"), "x": Numeric()},
    )
    model.fit(X, y, sample_weight=sample_weight)

    session = EditorSession.from_model(model, terms=["region"])

    np.testing.assert_allclose(session.terms["region"].weights, [3.0, 2.0, 7.0])


def test_widget_state_exposes_ci_bands_and_single_point_centering(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_num", "x_spline"])
    widget = session.widget()
    try:
        state = _get_json(f"{widget.url}/state")
        numeric = state["terms"]["x_num"]
        spline = state["terms"]["x_spline"]

        assert numeric["x_domain"][0] < numeric["x"][0] < numeric["x_domain"][1]
        assert numeric["y_label"] == "relativity"
        np.testing.assert_allclose(numeric["y"], np.exp(session.terms["x_num"].edited_log_effect))
        assert spline["ci_lower_y"] is not None
        assert spline["ci_upper_y"] is not None
        assert spline["effective_df"] == pytest.approx(session.terms["x_spline"].metadata["edf"])
        assert len(spline["ci_lower_y"]) == spline["n_points"]
        np.testing.assert_allclose(spline["y"], np.exp(session.terms["x_spline"].edited_log_effect))
        np.testing.assert_allclose(
            spline["original_y"],
            np.exp(session.terms["x_spline"].original_log_effect),
        )
    finally:
        widget.close()


def test_widget_state_exposes_semantic_model_revision(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        initial = _get_json(f"{widget.url}/state")
        _post_json(f"{widget.url}/select", {"term": "x_spline", "indices": [5, 6]})
        selected = _get_json(f"{widget.url}/state")
        changed = _post_json(f"{widget.url}/op", {"operation": "shift_up"})
        assert initial["model_revision"] == selected["model_revision"] == 0
        assert changed["model_revision"] == 1
    finally:
        widget.close()


def test_widget_snapshot_generations_track_chart_mutation_categories(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline", "region"])
    widget = session.widget()
    try:
        snapshots = [widget._state()]
        snapshots.append(widget._select("x_spline", [5]))
        snapshots.append(widget._set_term("region"))

        before_control_revision = session.model_revision
        snapshots.append(widget._set_control_count("x_spline", 6))
        assert snapshots[-1]["model_revision"] == before_control_revision
        controls = session.control_points("x_spline", n_handles=6)
        handle_index = int(controls["x"].size // 2)
        handle_value = float(np.exp(controls["log_effect"][handle_index] + 0.05))
        snapshots.append(widget._control("x_spline", handle_index, handle_value))
        snapshots.append(widget._drag("x_spline", [5], delta=0.01))
        snapshots.append(widget._operate("shift_up"))
        snapshots.append(widget._operate("undo"))
        snapshots.append(widget._operate("redo"))
        snapshots.append(widget._operate("reset"))

        snapshots.append(widget._select("region", [1]))
        before_reorder_revision = session.model_revision
        snapshots.append(widget._reorder_levels("region", target_index=0))
        assert snapshots[-1]["model_revision"] == before_reorder_revision
        snapshots.append(widget._operate("reset_order", "region"))
        snapshots.append(widget._operate("select_all", "region"))
        snapshots.append(widget._state())
    finally:
        widget.close()

    assert [state["state_generation"] for state in snapshots] == list(range(1, len(snapshots) + 1))
    assert [state["chart_generation"] for state in snapshots] == [
        0,
        0,
        0,
        1,
        2,
        3,
        4,
        5,
        6,
        7,
        7,
        8,
        9,
        9,
        9,
    ]


def test_widget_structural_mutations_advance_chart_generation_once(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    try:
        selected = widget._select("region", [1, 2])
        collapsed = widget._collapse_levels("region", method="fit")["state"]
        uncollapsed = widget._uncollapse_levels()["state"]
        selected_again = widget._select("region", [1, 2])
        collapsed_again = widget._collapse_levels("region", method="fit")["state"]
        ungrouped = widget._ungroup_levels("region", method="fit")["state"]
    finally:
        widget.close()

    assert [
        state["state_generation"]
        for state in (
            selected,
            collapsed,
            uncollapsed,
            selected_again,
            collapsed_again,
            ungrouped,
        )
    ] == [1, 2, 3, 4, 5, 6]
    assert [
        state["chart_generation"]
        for state in (
            selected,
            collapsed,
            uncollapsed,
            selected_again,
            collapsed_again,
            ungrouped,
        )
    ] == [0, 1, 2, 2, 3, 4]


@pytest.mark.parametrize("operation", ["collapse", "ungroup", "uncollapse"])
def test_invalid_level_display_cannot_commit_structural_mutations(editor_model, operation):
    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    try:
        widget._select("region", [1, 2])
        if operation != "collapse":
            widget._collapse_levels("region", method="fit")

        before_model = session.model
        before_revision = session.model_revision
        before_history = [id(model) for model in session.collapse_history]
        before_chart_generation = widget._chart_generation

        with pytest.raises(ValueError, match="level_display"):
            if operation == "collapse":
                widget._collapse_levels("region", method="fit", level_display="invalid")
            elif operation == "ungroup":
                widget._ungroup_levels("region", method="fit", level_display="invalid")
            else:
                widget._uncollapse_levels(level_display="invalid")

        assert session.model is before_model
        assert session.model_revision == before_revision
        assert [id(model) for model in session.collapse_history] == before_history
        assert widget._chart_generation == before_chart_generation
    finally:
        widget.close()


def test_invalid_level_display_is_rejected_before_other_editor_mutations(
    editor_model,
    monkeypatch,
):
    import superglm.editor.widget as widget_module

    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    profile_called = False
    thread_started = False

    def unexpected_profile(*_args, **_kwargs):
        nonlocal profile_called
        profile_called = True
        raise AssertionError("profile mutation ran before level_display validation")

    def unexpected_thread_start(_thread):
        nonlocal thread_started
        thread_started = True

    monkeypatch.setattr(session, "reprofile_distribution", unexpected_profile)
    monkeypatch.setattr(widget_module.threading.Thread, "start", unexpected_thread_start)
    try:
        with pytest.raises(ValueError, match="level_display"):
            widget._refit_offset(level_display="invalid")
        with pytest.raises(ValueError, match="level_display"):
            widget._profile_distribution("tweedie_p", level_display="invalid")
        with pytest.raises(ValueError, match="level_display"):
            widget._start_profile_distribution_job("tweedie_p", level_display="invalid")

        assert profile_called is False
        assert thread_started is False
        assert widget._offset_refit_model is None
        assert widget._profile_jobs == {}
    finally:
        widget.close()


def test_widget_final_ungroup_keeps_collapse_metadata_history_aligned(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    try:
        widget._select("region", [1, 2])
        widget._collapse_levels("region", method="fit")
        widget._select("region", [1, 2])

        state = widget._ungroup_levels("region", method="fit")["state"]
    finally:
        widget.close()

    assert session.collapse_history == []
    assert widget._collapse_info_history == []
    assert widget._in_force_info is None
    assert state["can_uncollapse_levels"] is False
    assert state["last_collapse"] is None


def test_widget_composite_selection_serializes_state_once(editor_model, monkeypatch):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    state_calls = 0
    original_state = widget._state

    def counted_state():
        nonlocal state_calls
        state_calls += 1
        return original_state()

    try:
        monkeypatch.setattr(widget, "_state", counted_state)

        payload = widget._select("x_spline", [3, 4])

        assert state_calls == 1
        assert payload["selected_term"] == "x_spline"
        assert payload["selection"]["x_spline"] == [3, 4]
    finally:
        widget.close()


def test_widget_state_exposes_density_for_continuous_and_bars_for_levels(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline", "region"])
    widget = session.widget()
    try:
        state = _get_json(f"{widget.url}/state")
        spline_exposure = state["terms"]["x_spline"]["exposure"]
        region_exposure = state["terms"]["region"]["exposure"]

        assert spline_exposure["kind"] == "density"
        assert len(spline_exposure["x"]) == state["terms"]["x_spline"]["n_points"]
        assert len(spline_exposure["y"]) == state["terms"]["x_spline"]["n_points"]
        assert max(spline_exposure["y"]) > min(spline_exposure["y"])
        assert region_exposure["kind"] == "bars"
        assert len(region_exposure["y"]) == state["terms"]["region"]["n_points"]
    finally:
        widget.close()


def test_widget_state_caps_continuous_handles_for_large_grids(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline", "region"], n_points=1200)
    widget = session.widget()
    try:
        state = _get_json(f"{widget.url}/state")
        spline = state["terms"]["x_spline"]
        region = state["terms"]["region"]

        assert spline["n_points"] == 1200
        assert 100 < len(spline["handle_indices"]) < spline["n_points"]
        assert spline["handle_indices"][0] == 0
        assert spline["handle_indices"][-1] == spline["n_points"] - 1
        assert region["handle_indices"] == list(range(region["n_points"]))
    finally:
        widget.close()


def test_numeric_polynomial_and_ordered_step_edits_apply(editor_model, editor_frame):
    X, _ = editor_frame
    session = EditorSession.from_model(editor_model, terms=["x_num", "x_poly", "band"])

    session.select_indices("x_num", [0])
    session.shift("x_num", 0.1)
    session.select_x("x_poly", -1.0, 1.0)
    session.shift("x_poly", -0.15)
    session.select_levels("band", ["high"])
    session.shift("band", 0.25)

    edited = session.to_model()

    eta_delta = edited._predict_eta_exact(X) - editor_model._predict_eta_exact(X)
    assert np.max(np.abs(eta_delta)) > 0.1
    assert np.mean(eta_delta[X["band"] == "high"]) > np.mean(eta_delta[X["band"] == "low"])


def test_ordered_categorical_spline_level_edit_applies_to_model_copy():
    rng = np.random.default_rng(1234)
    levels = ["young", "mid", "senior", "older"]
    n = 360
    X = pd.DataFrame(
        {
            "age_band": rng.choice(levels, n, p=[0.25, 0.35, 0.25, 0.15]),
            "x": rng.normal(size=n),
        }
    )
    level_effect = {"young": -0.1, "mid": 0.0, "senior": 0.2, "older": 0.35}
    y = (
        np.array([level_effect[level] for level in X["age_band"]])
        + 0.1 * X["x"].to_numpy()
        + rng.normal(0.0, 0.05, n)
    )
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={
            "age_band": OrderedCategorical(order=levels, basis=Spline(kind="ps", n_knots=3)),
            "x": Numeric(),
        },
    )
    model.fit(X, y)

    session = EditorSession.from_model(model, terms=["age_band"])
    session.select_levels("age_band", ["older"])
    session.shift("age_band", 0.3)
    edited = session.to_model()

    eta_delta = edited._predict_eta_exact(X) - model._predict_eta_exact(X)
    assert np.mean(eta_delta[X["age_band"] == "older"]) > 0.1


def test_ordered_categorical_spline_integer_level_edit_applies_to_model_copy():
    rng = np.random.default_rng(20260802)
    levels = [1, 2, 3, 4]
    n = 360
    X = pd.DataFrame(
        {
            "age_band": rng.choice(levels, n, p=[0.25, 0.35, 0.25, 0.15]),
            "x": rng.normal(size=n),
        }
    )
    level_effect = {1: -0.1, 2: 0.0, 3: 0.2, 4: 0.35}
    y = (
        np.array([level_effect[level] for level in X["age_band"]])
        + 0.1 * X["x"].to_numpy()
        + rng.normal(0.0, 0.05, n)
    )
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={
            "age_band": OrderedCategorical(order=levels, basis=Spline(kind="ps", n_knots=3)),
            "x": Numeric(),
        },
    )
    model.fit(X, y)

    session = EditorSession.from_model(model, terms=["age_band"])
    session.select_levels("age_band", ["4"])
    session.shift("age_band", 0.3)
    edited = session.to_model()

    eta_delta = edited._predict_eta_exact(X) - model._predict_eta_exact(X)
    assert np.mean(eta_delta[X["age_band"] == 4]) > 0.1


def test_ordered_categorical_spline_edited_summary_reports_level_rows():
    rng = np.random.default_rng(20260713)
    levels = ["young", "mid", "senior", "older"]
    n = 360
    X = pd.DataFrame({"age_band": rng.choice(levels, n, p=[0.25, 0.35, 0.25, 0.15])})
    level_effect = {"young": -0.1, "mid": 0.0, "senior": 0.2, "older": 0.35}
    y = np.array([level_effect[level] for level in X["age_band"]]) + rng.normal(0.0, 0.05, n)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={
            "age_band": OrderedCategorical(
                order=levels,
                basis=Spline(kind="ps", n_knots=3),
            )
        },
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["age_band"])
    session.select_levels("age_band", ["older"])
    session.shift("age_band", 0.2)

    edited = session.to_model()
    summary = edited.summary()
    row_names = [row.name for row in summary._coef_rows]

    assert all(f"age_band[{level}]" in row_names for level in levels)
    smooth_row = next(row for row in summary._coef_rows if row.name == "age_band")
    assert smooth_row.is_spline
    assert smooth_row.wald_p is None
    assert not any("[bs" in name or "[basis" in name for name in row_names)


def test_save_load_roundtrip(editor_model, tmp_path):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_indices("x_spline", [1, 2])
    session.shift("x_spline", 0.2)
    path = tmp_path / "edits.json"

    session.save(path)
    loaded = EditorSession.load(path, model=editor_model)

    np.testing.assert_allclose(
        loaded.terms["x_spline"].edited_log_effect,
        session.terms["x_spline"].edited_log_effect,
    )
    assert loaded.history[-1].operation == "shift"


def test_load_rejects_same_shape_artifact_from_different_baseline(
    editor_model, editor_frame, tmp_path
):
    X, y = editor_frame
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_indices("x_spline", [1, 2])
    session.shift("x_spline", 0.2)
    path = tmp_path / "edits.json"
    session.save(path)

    refit = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        spline_penalty=0.1,
        features={
            "x_spline": Spline(n_knots=8),
            "x_poly": Polynomial(degree=2),
            "x_num": Numeric(),
            "region": Categorical(base="first"),
            "band": OrderedCategorical(
                order=["low", "medium", "high"], basis=Spline(kind="ps", n_knots=2), base="first"
            ),
        },
    )
    shifted_y = y + 0.35 * np.sin(np.asarray(X["x_spline"], dtype=np.float64))
    refit.fit(X, shifted_y)

    assert (
        refit.term_inference("x_spline", with_se=False).x.shape == session.terms["x_spline"].x.shape
    )
    with pytest.raises(ValueError, match="Loaded baseline"):
        EditorSession.load(path, model=refit)


def test_widget_import_is_lazy():
    import sys

    import superglm.editor

    assert "anywidget" not in sys.modules
    assert "ipympl" not in sys.modules
    assert "ipywidgets" not in sys.modules
    assert hasattr(superglm.editor, "EditorSession")


def test_editor_asset_reader_rejects_path_traversal(monkeypatch):
    from superglm.editor import assets

    assert assets.app_asset_content_type("styles.css").startswith("text/css")

    def fail_resource_lookup(_package):
        raise AssertionError("invalid asset paths must be rejected before resource lookup")

    monkeypatch.setattr(assets, "files", fail_resource_lookup)
    for path in [
        "",
        ".",
        "..",
        "../widget.py",
        r"..\widget.py",
        r"nested\..\asset.js",
        "/styles.css",
        "styles.css/",
        "nested//asset.js",
    ]:
        with pytest.raises(FileNotFoundError):
            assets.read_app_asset(path)


def test_widget_renders_plain_iframe_app(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])

    widget = session.widget()
    try:
        html = widget._repr_html_()

        assert "<iframe" in html
        assert "127.0.0.1" in html
        assert "token=" in html
        assert not hasattr(widget, "_model_module")
        assert widget.selected_term == "x_spline"
        assert widget.terms["x_spline"]["n_points"] == session.terms["x_spline"].size

        payload = _get_json(f"{widget.url}/state")
        assert payload["selected_term"] == "x_spline"
        assert payload["terms"]["x_spline"]["n_points"] == session.terms["x_spline"].size
    finally:
        widget.close()


def test_widget_http_rejects_missing_editor_token(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        with pytest.raises(urllib.error.HTTPError) as excinfo:
            urllib.request.urlopen(f"{widget.url}/state", timeout=5)
        assert excinfo.value.code == 403
        assert _get_json(f"{widget.url}/state")["selected_term"] == "x_spline"
    finally:
        widget.close()


def test_widget_favicon_does_not_emit_browser_404(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        request = urllib.request.Request(f"{widget.url}/favicon.ico", method="GET")
        try:
            with urllib.request.urlopen(request, timeout=5) as response:
                assert response.status == 204
        except urllib.error.HTTPError as exc:  # pragma: no cover - assertion path
            pytest.fail(f"favicon request returned HTTP {exc.code}")
    finally:
        widget.close()


def test_widget_http_selection_and_average_updates_session(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        term = session.terms["x_spline"]
        indices = [8, 9, 10, 11]
        before = term.edited_log_effect[indices].copy()

        _post_json(f"{widget.url}/select", {"term": "x_spline", "indices": indices})
        assert session.selection("x_spline").tolist() == indices

        _post_json(f"{widget.url}/op", {"operation": "average"})
        expected = np.log(np.average(np.exp(before), weights=term.weights[indices]))
        np.testing.assert_allclose(term.edited_log_effect[indices], expected)

        term.edited_log_effect[indices] = np.array([0.0, 1.0, -0.5, 2.0])
        _post_json(f"{widget.url}/op", {"operation": "linearise"})
        linear_target = np.interp(
            term.x[indices],
            [term.x[indices[0]], term.x[indices[-1]]],
            [0.0, 2.0],
        )
        expected_linear = 0.5 * np.array([0.0, 1.0, -0.5, 2.0]) + 0.5 * linear_target
        np.testing.assert_allclose(term.edited_log_effect[indices], expected_linear)

        term.edited_log_effect[indices] = np.array([0.2, 1.0, -0.4, 0.8])
        _post_json(f"{widget.url}/op", {"operation": "level_left"})
        np.testing.assert_allclose(term.edited_log_effect[indices], 0.2)

        term.edited_log_effect[indices] = np.array([0.2, 1.0, -0.4, 0.8])
        _post_json(f"{widget.url}/op", {"operation": "level_right"})
        np.testing.assert_allclose(term.edited_log_effect[indices], 0.8)

        term.edited_log_effect[indices] = np.array([0.2, 1.0, -0.4, 0.8])
        _post_json(f"{widget.url}/op", {"operation": "snap_highest"})
        np.testing.assert_allclose(term.edited_log_effect[indices], 1.0)

        term.edited_log_effect[indices] = np.array([0.2, 1.0, -0.4, 0.8])
        _post_json(f"{widget.url}/op", {"operation": "snap_lowest"})
        np.testing.assert_allclose(term.edited_log_effect[indices], -0.4)
    finally:
        widget.close()


def test_widget_average_falls_back_when_selected_exposure_is_zero(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        term = session.terms["x_spline"]
        indices = [8, 9, 10]
        term.weights[indices] = 0.0
        term.edited_log_effect[indices] = np.log(np.array([1.0, 1.4, 2.0]))

        _post_json(f"{widget.url}/select", {"term": "x_spline", "indices": indices})
        _post_json(f"{widget.url}/op", {"operation": "average"})

        expected = np.log(np.mean([1.0, 1.4, 2.0]))
        np.testing.assert_allclose(term.edited_log_effect[indices], expected)
    finally:
        widget.close()


def test_widget_average_expands_collapsed_level_groups_before_weighting(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    session.select_levels("region", ["B", "C"])
    session.replace_with_collapsed_levels("region", method="fit")
    widget = session.widget()
    try:
        term = session.terms["region"]
        term.weights = np.array([1.0, 1.0, 100.0])
        term.edited_log_effect = np.log(np.array([1.0, 3.0, 3.0]))

        _post_json(f"{widget.url}/select", {"term": "region", "indices": [0, 1]})
        _post_json(f"{widget.url}/op", {"operation": "average"})

        expected = np.log(np.average([1.0, 3.0, 3.0], weights=[1.0, 1.0, 100.0]))
        np.testing.assert_allclose(term.edited_log_effect[[0, 1, 2]], expected)
    finally:
        widget.close()


def test_widget_http_collapse_levels_refit_updates_summary_source(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    try:
        _post_json(f"{widget.url}/select", {"term": "region", "indices": [1, 2]})
        payload = _post_json(
            f"{widget.url}/collapse_levels",
            {"term": "region", "method": "fit", "level_display": "grouped"},
        )

        assert set(payload) == {"state", "summary", "timing"}
        assert payload["summary"]["available"] is True
        assert payload["summary"]["source"] == "in_force"
        assert payload["summary"]["level_display"] == "grouped"
        assert payload["summary"]["label"] == "In-force edit model"
        assert payload["state"]["model_revision"] == session.model_revision == 1
        assert payload["state"]["terms"]["region"]["y"][1] == pytest.approx(
            payload["state"]["terms"]["region"]["y"][2]
        )
        assert payload["timing"]["fit_ms"] >= 0.0
        assert payload["timing"]["summary_ms"] >= 0.0
        assert payload["timing"]["state_ms"] >= 0.0
        assert widget.session.model is not editor_model
        grouping = widget.session.model.features["region"]._grouping
        assert grouping.original_to_group["B"] == "B+C"
        assert grouping.original_to_group["C"] == "B+C"
        assert "model_source" not in payload["state"]
        assert "model_sources" not in payload["state"]
        assert "Current editable model" in payload["summary"]["note"]
    finally:
        widget.close()


def test_widget_http_ungroup_levels_returns_transition_envelope(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    try:
        _post_json(f"{widget.url}/select", {"term": "region", "indices": [1, 2]})
        _post_json(f"{widget.url}/collapse_levels", {"term": "region", "method": "fit"})

        payload = _post_json(
            f"{widget.url}/ungroup_levels",
            {"term": "region", "method": "fit", "level_display": "grouped"},
        )

        assert set(payload) == {"state", "summary", "timing"}
        assert payload["summary"]["available"] is True
        assert payload["summary"]["source"] == "in_force"
        assert payload["summary"]["level_display"] == "grouped"
        assert payload["state"]["model_revision"] == session.model_revision == 2
        assert payload["timing"]["operation"] == "ungroup_levels"
        assert payload["timing"]["fit_ms"] >= 0.0
        assert payload["timing"]["summary_ms"] >= 0.0
        assert payload["timing"]["state_ms"] >= 0.0
    finally:
        widget.close()


def _structural_alias_fixture(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    session.select_levels("region", ["B", "C"])
    widget._collapse_levels("region", method="fit")
    current_info = {
        "term": "region",
        "method": "fit",
        "details": {"groups": [{"label": "B+C", "levels": ["B", "C"]}]},
    }
    collapsed_info = {
        "term": "region",
        "method": "fit",
        "details": {"groups": [{"label": "B+C", "levels": ["B", "C"]}]},
    }
    widget._in_force_info = current_info
    widget._collapsed_refit_info = collapsed_info
    widget._collapse_info_history.append(dict(current_info))
    envelope = widget._structural_transition(
        "collapse_levels",
        operation_start=0.0,
        fit_start=0.0,
        fit_end=0.0,
    )
    return session, widget, envelope, current_info, collapsed_info


def test_structural_transition_envelope_mutation_cannot_change_live_state(editor_model):
    session, widget, envelope, current_info, collapsed_info = _structural_alias_fixture(
        editor_model
    )
    expected_levels = list(session.terms["region"].levels)
    expected_current = json.loads(json.dumps(current_info))
    expected_collapsed = json.loads(json.dumps(collapsed_info))
    try:
        envelope["state"]["terms"]["region"]["levels"][0] = "envelope-level"
        envelope["summary"]["collapse"]["term"] = "envelope-term"
        envelope["summary"]["collapse"]["details"]["groups"][0]["levels"][0] = (
            "envelope-summary-nested"
        )
        envelope["state"]["last_collapse"]["term"] = "envelope-last-term"
        envelope["state"]["last_collapse"]["details"]["groups"][0]["levels"][0] = (
            "envelope-last-nested"
        )

        assert session.terms["region"].levels == expected_levels
        assert widget._in_force_info == expected_current
        assert widget._collapsed_refit_info == expected_collapsed
        assert widget._collapse_info_history[-1] == expected_current

        widget._uncollapse_levels()
        assert widget._in_force_info == expected_current
    finally:
        widget.close()


def test_live_state_mutation_cannot_change_structural_transition_envelope(editor_model):
    session, widget, envelope, current_info, collapsed_info = _structural_alias_fixture(
        editor_model
    )
    expected_levels = list(envelope["state"]["terms"]["region"]["levels"])
    expected_summary = json.loads(json.dumps(envelope["summary"]["collapse"]))
    expected_last_collapse = json.loads(json.dumps(envelope["state"]["last_collapse"]))
    try:
        session.terms["region"].levels[0] = "live-level"
        current_info["term"] = "live-term"
        current_info["details"]["groups"][0]["levels"][0] = "live-summary-nested"
        collapsed_info["term"] = "live-last-term"
        collapsed_info["details"]["groups"][0]["levels"][0] = "live-last-nested"
        widget._collapse_info_history[-1]["details"]["groups"][0]["levels"][1] = (
            "live-history-nested"
        )

        assert envelope["state"]["terms"]["region"]["levels"] == expected_levels
        assert envelope["summary"]["collapse"] == expected_summary
        assert envelope["state"]["last_collapse"] == expected_last_collapse
    finally:
        widget.close()


def test_structural_transition_holds_widget_lock_through_snapshot_detachment(
    editor_model,
    monkeypatch,
):
    import superglm.editor.widget as widget_module

    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    detachment_entered = threading.Event()
    release_detachment = threading.Event()
    mutation_attempted = threading.Event()
    mutation_done = threading.Event()
    transition_errors = []
    mutation_errors = []
    original_jsonable = widget_module.jsonable
    detachment_calls = 0

    def blocking_jsonable(value):
        nonlocal detachment_calls
        detachment_calls += 1
        if detachment_calls == 2:
            detachment_entered.set()
            if not release_detachment.wait(timeout=2):
                raise TimeoutError("snapshot detachment was not released")
        return original_jsonable(value)

    def run_transition():
        try:
            widget._structural_transition(
                "collapse_levels",
                operation_start=0.0,
                fit_start=0.0,
                fit_end=0.0,
            )
        except BaseException as exc:  # pragma: no cover - asserted below
            transition_errors.append(exc)

    def run_mutation():
        try:
            mutation_attempted.set()
            widget._select("region", [0])
        except BaseException as exc:  # pragma: no cover - asserted below
            mutation_errors.append(exc)
        finally:
            mutation_done.set()

    transition_thread = threading.Thread(target=run_transition)
    mutation_thread = threading.Thread(target=run_mutation)
    try:
        monkeypatch.setattr(widget_module, "jsonable", blocking_jsonable)
        transition_thread.start()
        assert detachment_entered.wait(timeout=2), "transition never began JSON-safe detachment"

        mutation_thread.start()
        assert mutation_attempted.wait(timeout=2), "competing mutation was never attempted"
        assert not mutation_done.wait(timeout=0.1), "mutation escaped the transition lock"
    finally:
        release_detachment.set()
        transition_thread.join(timeout=2)
        if mutation_thread.ident is not None:
            mutation_thread.join(timeout=2)
        widget.close()

    assert not transition_thread.is_alive()
    assert not mutation_thread.is_alive()
    assert transition_errors == []
    assert mutation_errors == []
    assert mutation_done.is_set()
    assert detachment_calls == 2


def test_widget_collapse_restores_selection_by_level_when_indices_are_stale(
    editor_model,
    monkeypatch,
):
    import superglm.editor.widget as widget_module

    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    try:
        session.select_indices("region", [1, 2])

        def replace_with_smaller_term(term, *, method="auto"):
            assert term == "region"
            assert method == "fit"
            session.terms["region"] = EditableTerm(
                name="region",
                kind="categorical",
                x=np.array([0.0, 1.0]),
                levels=["A", "B"],
                original_log_effect=np.zeros(2),
                edited_log_effect=np.zeros(2),
                weights=np.ones(2),
                metadata={"term_type": "categorical"},
            )
            session._selection["region"] = np.array([], dtype=np.intp)
            return editor_model

        monkeypatch.setattr(session, "replace_with_collapsed_levels", replace_with_smaller_term)
        monkeypatch.setattr(
            widget_module,
            "summary_payload",
            lambda *_args, **_kwargs: {"available": True, "source": "in_force"},
        )

        payload = widget._collapse_levels("region", method="fit")

        assert payload["summary"]["available"] is True
        assert session.selection("region").tolist() == [1]
    finally:
        widget.close()


def test_widget_collapse_levels_reports_refit_timing(editor_model, monkeypatch):
    import superglm.editor.widget as widget_module

    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    clock = iter([10.00, 10.01, 10.06, 10.07, 10.09, 10.10, 10.13])
    try:
        monkeypatch.setattr(widget_module.time, "perf_counter", lambda: next(clock))
        monkeypatch.setattr(
            session,
            "replace_with_collapsed_levels",
            lambda term, *, method="auto": editor_model,
        )
        monkeypatch.setattr(
            widget_module,
            "summary_payload",
            lambda *_args, **_kwargs: {"available": True, "source": "in_force"},
        )

        payload = widget._collapse_levels("region", method="fit")

        assert payload["timing"]["operation"] == "collapse_levels"
        assert payload["timing"]["fit_ms"] == pytest.approx(50.0)
        assert payload["timing"]["summary_ms"] == pytest.approx(20.0)
        assert payload["timing"]["state_ms"] == pytest.approx(30.0)
        assert payload["timing"]["server_total_ms"] == pytest.approx(130.0)
    finally:
        widget.close()


def test_widget_json_responses_report_serialization_timing(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    request = urllib.request.Request(
        f"{widget.url}/state",
        headers=_editor_token_header(f"{widget.url}/state"),
    )
    try:
        with urllib.request.urlopen(request, timeout=5) as response:
            payload = json.loads(response.read().decode("utf-8"))
            server_timing = response.headers.get("Server-Timing", "")

        assert payload["model_revision"] == session.model_revision
        assert re.fullmatch(r"json;dur=\d+(?:\.\d{3})", server_timing)
    finally:
        widget.close()


def test_widget_http_uncollapse_levels_restores_previous_model(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    try:
        _post_json(f"{widget.url}/select", {"term": "region", "indices": [1, 2]})
        collapse_payload = _post_json(
            f"{widget.url}/collapse_levels", {"term": "region", "method": "fit"}
        )
        collapsed_state = collapse_payload["state"]
        assert collapsed_state["can_uncollapse_levels"] is True
        assert widget.session.model is not editor_model
        assert widget.session.model.features["region"]._grouping.original_to_group["B"] == "B+C"

        payload = _post_json(
            f"{widget.url}/uncollapse_levels",
            {"level_display": "grouped"},
        )

        assert set(payload) == {"state", "summary", "timing"}
        assert payload["summary"]["available"] is True
        assert payload["summary"]["source"] == "in_force"
        assert payload["summary"]["level_display"] == "grouped"
        assert payload["state"]["model_revision"] == session.model_revision == 2
        assert widget.session.model is editor_model
        assert getattr(widget.session.model.features["region"], "_grouping", None) is None
        state = payload["state"]
        assert state["can_uncollapse_levels"] is False
        assert state["last_collapse"] is None
        assert state["terms"]["region"]["y"][1] != pytest.approx(state["terms"]["region"]["y"][2])
        assert payload["timing"]["operation"] == "uncollapse_levels"
        assert payload["timing"]["fit_ms"] >= 0.0
        assert payload["timing"]["summary_ms"] >= 0.0
        assert payload["timing"]["state_ms"] >= 0.0
    finally:
        widget.close()


def test_widget_http_selected_summary_uses_in_force_without_model_selector(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    try:
        state = _get_json(f"{widget.url}/state")
        assert "model_source" not in state
        assert "model_sources" not in state

        summary = _post_json(
            f"{widget.url}/summary",
            {"source": "selected", "level_display": "grouped"},
        )
        assert summary["source"] == "in_force"
        assert summary["level_display"] == "grouped"
        assert summary["compact"]["level_display"] == "grouped"

        original = _post_json(f"{widget.url}/summary", {"source": "original"})
        assert original["source"] == "original"
    finally:
        widget.close()


def test_widget_http_drag_and_reset_updates_session(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        term = session.terms["x_spline"]
        original = term.original_log_effect.copy()

        _post_json(
            f"{widget.url}/drag",
            {"term": "x_spline", "indices": [5], "values": [float(np.exp(original[5]) + 0.25)]},
        )
        assert session.selection("x_spline").tolist() == [5]
        assert np.exp(term.edited_log_effect[5]) == pytest.approx(np.exp(original[5]) + 0.25)

        _post_json(f"{widget.url}/op", {"operation": "reset"})
        assert term.edited_log_effect[5] == pytest.approx(original[5])
        state = _get_json(f"{widget.url}/state")
        assert state["history"]["active"] == []
        assert state["terms"]["x_spline"]["previous_y"] is None
    finally:
        widget.close()


def test_widget_http_control_handle_updates_session(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        term = session.terms["x_spline"]
        state = _post_json(f"{widget.url}/control_count", {"term": "x_spline", "count": 6})
        controls = session.control_points("x_spline", n_handles=6)
        assert len(state["terms"]["x_spline"]["controls"]["x"]) == 6
        assert state["terms"]["x_spline"]["controls"]["count"] == 6
        assert state["terms"]["x_spline"]["controls"]["basis_index"] == (
            controls["basis_index"].astype(int).tolist()
        )
        payload_controls = state["terms"]["x_spline"]["controls"]
        assert "build_basis" in payload_controls
        assert "build_log_effect" in payload_controls
        assert len(payload_controls["build_basis"]) == len(payload_controls["build_log_effect"])
        assert len(payload_controls["build_basis"][0]) == state["terms"]["x_spline"]["n_points"]
        payload_basis = np.asarray(payload_controls["basis"], dtype=np.float64)
        raw_basis = editor_model._specs["x_spline"]._raw_basis_matrix(term.x)
        if hasattr(raw_basis, "toarray"):
            raw_basis = raw_basis.toarray()
        basis_index = int(payload_controls["basis_index"][2])
        assert payload_basis.shape == (6, term.size)
        np.testing.assert_allclose(payload_basis[2], np.asarray(raw_basis)[:, basis_index])

        before = term.edited_log_effect.copy()
        handle_index = int(controls["x"].size // 2)
        target_rel = float(np.exp(controls["log_effect"][handle_index] + 0.25))

        state = _post_json(
            f"{widget.url}/control",
            {"term": "x_spline", "handle_index": handle_index, "value": target_rel},
        )

        assert np.max(np.abs(term.edited_log_effect - before)) > 0.02
        assert len(state["terms"]["x_spline"]["controls"]["x"]) == 6
        assert state["terms"]["x_spline"]["controls"]["x"][handle_index] == pytest.approx(
            controls["x"][handle_index]
        )
        assert state["terms"]["x_spline"]["controls"]["y"][handle_index] == pytest.approx(
            target_rel,
            rel=0.08,
        )
    finally:
        widget.close()


def test_widget_http_select_all_and_impact_updates_session(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        term = session.terms["x_spline"]

        state = _post_json(f"{widget.url}/op", {"operation": "select_all"})
        assert session.selection("x_spline").tolist() == list(range(term.size))
        assert state["terms"]["x_spline"]["impact"]["weighted_mean_relativity"] == pytest.approx(
            1.0
        )

        state = _post_json(
            f"{widget.url}/drag",
            {
                "term": "x_spline",
                "indices": [0],
                "values": [float(np.exp(term.original_log_effect[0]) + 0.2)],
            },
        )
        impact = state["terms"]["x_spline"]["impact"]
        assert impact["selected_weight_share"] > 0.0
        assert impact["weighted_mean_relativity"] > 1.0
    finally:
        widget.close()


def test_widget_http_metrics_recomputes_fit_metric_for_edited_copy(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        baseline = _post_json(f"{widget.url}/metrics", {"metric": "deviance"})
        assert baseline["metric"] == "deviance"
        assert baseline["available"] is True
        assert baseline["original"] == pytest.approx(baseline["edited"])
        assert baseline["metrics"]["original"]["effective_df"] == pytest.approx(
            editor_model.result.effective_df
        )
        assert baseline["metrics"]["edited"]["effective_df"] == pytest.approx(
            editor_model.result.effective_df
        )

        _post_json(f"{widget.url}/select", {"term": "x_spline", "indices": [20, 21, 22]})
        _post_json(f"{widget.url}/op", {"operation": "shift_up"})
        changed = _post_json(f"{widget.url}/metrics", {"metric": "deviance"})

        assert changed["available"] is True
        assert changed["edited"] != pytest.approx(changed["original"])
        assert "metrics" in changed
        assert "aic" in changed["metrics"]["edited"]
        assert "effective_df" in changed["metrics"]["edited"]
    finally:
        widget.close()


def test_http_evidence_echoes_revision_sequence_and_supersedes_stale_requests(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        revision = session.model_revision
        requests = (
            ("metrics", {"metric": "deviance"}),
            ("report", {"report": "validation"}),
            ("summary", {"source": "original"}),
        )
        for sequence, (path, request_payload) in enumerate(requests, start=10):
            payload = _post_json(
                f"{widget.url}/{path}",
                {
                    **request_payload,
                    "model_revision": revision,
                    "request_sequence": sequence,
                },
            )
            assert payload["model_revision"] == revision
            assert payload["request_sequence"] == sequence

        _post_json(f"{widget.url}/select", {"term": "x_spline", "indices": [20, 21]})
        _post_json(f"{widget.url}/op", {"operation": "shift_up"})
        stale = _post_json(
            f"{widget.url}/metrics",
            {
                "metric": "deviance",
                "model_revision": revision,
                "request_sequence": 99,
            },
        )

        assert stale == {
            "status": "superseded",
            "model_revision": revision,
            "request_sequence": 99,
        }
    finally:
        widget.close()


def test_http_evidence_failure_echoes_revision_and_sequence(editor_model, monkeypatch):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()

    def fail(*_args, **_kwargs):
        raise ValueError("metric failure")

    monkeypatch.setattr(widget, "_metrics", fail)
    try:
        data = json.dumps(
            {
                "metric": "deviance",
                "model_revision": session.model_revision,
                "request_sequence": 123,
            }
        ).encode("utf-8")
        request = urllib.request.Request(
            f"{widget.url}/metrics",
            data=data,
            method="POST",
            headers={
                "Content-Type": "application/json",
                **_editor_token_header(f"{widget.url}/metrics"),
            },
        )
        with pytest.raises(urllib.error.HTTPError) as exc_info:
            urllib.request.urlopen(request, timeout=5)
        payload = json.loads(exc_info.value.read().decode("utf-8"))

        assert exc_info.value.code == 400
        assert payload == {
            "model_revision": session.model_revision,
            "request_sequence": 123,
            "error": "metric failure",
        }
    finally:
        widget.close()


def test_live_metrics_and_report_share_validation_scalars(
    editor_model,
    editor_frame,
    monkeypatch,
):
    import superglm.editor.metrics as metrics_module
    import superglm.editor.reports as reports_module

    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(X.iloc[:300], y[:300], None),
        validation_data=(X.iloc[300:380], y[300:380], None),
        test_data=(X.iloc[380:], y[380:], None),
    )
    widget = session.widget()
    calls: list[tuple[int, str]] = []
    original = metrics_module.compute_dataset_metrics

    def counted(model, dataset):
        calls.append((id(model.result), dataset.name))
        return original(model, dataset)

    monkeypatch.setattr(metrics_module, "compute_dataset_metrics", counted)
    monkeypatch.setattr(reports_module, "compute_dataset_metrics", counted)
    try:
        metrics = widget._metrics("deviance", "in_force")
        report = widget._report("validation")
    finally:
        widget.close()

    validation_calls = [call for call in calls if call[1] == "validation"]
    assert len(validation_calls) == 2
    assert metrics["model_revision"] == session.model_revision
    assert report["model_revision"] == session.model_revision


def test_original_metric_request_does_not_poison_current_revision_cache(editor_model):
    from superglm.editor.evaluation import default_metrics_dataset
    from superglm.editor.metrics import compute_dataset_metrics

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_indices("x_spline", [4, 5, 6])
    session.shift("x_spline", 0.35)
    dataset = default_metrics_dataset(session)
    assert dataset is not None
    expected_edited = compute_dataset_metrics(session.to_model(), dataset)["deviance"]
    widget = session.widget()
    try:
        original = widget._metrics("deviance", "original")
        in_force = widget._metrics("deviance", "in_force")
    finally:
        widget.close()

    assert original["edited"] == pytest.approx(original["original"])
    assert in_force["edited"] == pytest.approx(expected_edited)
    assert in_force["edited"] != pytest.approx(in_force["original"])


def test_metric_comparison_payload_assembles_cached_scalars(editor_model, editor_frame):
    from superglm.editor.evaluation import EvaluationDataset
    from superglm.editor.metrics import metric_comparison_payload

    X, y = editor_frame
    dataset = EvaluationDataset("validation", "Validation", X.iloc[:20], y[:20])
    original = {"deviance": 3.0, "aic": 5.0}
    edited = {"deviance": 2.25, "aic": 4.5}

    payload = metric_comparison_payload(
        "deviance",
        dataset,
        original,
        edited,
        model_revision=7,
        request_sequence=11,
    )

    assert payload == {
        "available": True,
        "model_revision": 7,
        "request_sequence": 11,
        "metric": "deviance",
        "label": "Deviance",
        "dataset": "validation",
        "dataset_label": "Validation",
        "n_obs": 20,
        "original": 3.0,
        "edited": 2.25,
        "delta": -0.75,
        "metrics": {"original": original, "edited": edited},
    }


def test_editor_session_accepts_plain_evaluation_tuples_and_cv_report(editor_model, editor_frame):
    from superglm.editor.metrics import metrics_payload
    from superglm.editor.reports import validation_report_payload

    X, y = editor_frame
    X_train, y_train = X.iloc[:300], y[:300]
    X_val, y_val = X.iloc[300:380], y[300:380]
    X_test, y_test = X.iloc[380:], y[380:]
    w_train = np.linspace(0.8, 1.2, len(X_train))
    w_val = np.linspace(1.0, 1.5, len(X_val))
    w_test = np.linspace(0.7, 1.1, len(X_test))
    cv_report = {
        "method": "KFold",
        "folds": 5,
        "rows": [{"candidate": "k=8", "mean_val_deviance": 0.0148, "se_val_deviance": 0.0002}],
    }

    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(X_train, y_train, w_train),
        validation_data=(X_val, y_val, w_val),
        test_data=(X_test, y_test, w_test),
        cv_report=cv_report,
    )

    metric_payload = metrics_payload(session, "deviance")

    assert metric_payload["available"] is True
    assert metric_payload["dataset"] == "validation"
    assert metric_payload["n_obs"] == len(X_val)

    report = validation_report_payload(session)
    assert report["available"] is True
    assert [split["name"] for split in report["splits"]] == ["train", "validation", "test"]
    assert report["splits"][0]["n_obs"] == len(X_train)
    assert report["splits"][1]["n_obs"] == len(X_val)
    assert report["splits"][2]["n_obs"] == len(X_test)
    assert report["cv_report"] == cv_report
    assert report["can_run_cv"] is False


def test_dataset_metrics_use_offset_aware_null_deviance():
    from superglm.editor.evaluation import EvaluationDataset
    from superglm.editor.metrics import compute_dataset_metrics
    from superglm.model.fit_ops import _compute_null_mu

    rng = np.random.default_rng(20260711)
    n = 160
    x = rng.normal(size=n)
    offset = np.linspace(-1.0, 1.0, n)
    sample_weight = rng.uniform(0.7, 1.6, size=n)
    X = pd.DataFrame({"x": x})
    y = 1.2 + 0.25 * x + offset + rng.normal(0.0, 0.04, n)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X, y, sample_weight=sample_weight, offset=offset)

    dataset = EvaluationDataset(
        name="validation",
        label="Validation",
        X=X,
        y=y,
        sample_weight=sample_weight,
        offset=offset,
    )
    metrics = compute_dataset_metrics(model, dataset)

    mu = model.predict(X, offset=offset)
    deviance = float(np.sum(sample_weight * model._distribution.deviance_unit(y, mu)))
    null_mu = _compute_null_mu(
        y,
        sample_weight,
        offset,
        model._distribution,
        model._link,
    )
    null_deviance = float(np.sum(sample_weight * model._distribution.deviance_unit(y, null_mu)))
    assert metrics["explained_deviance"] == pytest.approx(1.0 - deviance / null_deviance)


def test_dataset_metrics_reject_column_vector_weights():
    from superglm.editor.evaluation import EvaluationDataset
    from superglm.editor.metrics import compute_dataset_metrics

    rng = np.random.default_rng(20260712)
    n = 120
    x = rng.normal(size=n)
    offset = np.linspace(-0.5, 0.5, n)
    sample_weight = rng.uniform(0.8, 1.4, size=n)
    X = pd.DataFrame({"x": x})
    y = 0.7 + 0.2 * x + offset + rng.normal(0.0, 0.03, n)
    model = SuperGLM(
        family="gaussian",
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X, y, sample_weight=sample_weight, offset=offset)

    flat = EvaluationDataset(
        name="validation",
        label="Validation",
        X=X,
        y=y,
        sample_weight=sample_weight,
        offset=offset,
    )
    column = EvaluationDataset(
        name="validation",
        label="Validation",
        X=X,
        y=y.reshape(-1, 1),
        sample_weight=sample_weight.reshape(-1, 1),
        offset=offset.reshape(-1, 1),
    )

    flat_metrics = compute_dataset_metrics(model, flat)

    assert np.isfinite(flat_metrics["deviance"])
    with pytest.raises(ValueError, match="sample_weight must be one-dimensional"):
        compute_dataset_metrics(model, column)


def test_widget_http_summary_and_fixed_offset_refit(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline", "region"])
    widget = session.widget()
    try:
        original = _post_json(f"{widget.url}/summary", {"source": "original"})
        assert original["available"] is True
        assert original["source"] == "original"
        assert "SuperGLM Results" in original["html"]
        assert original["compact"]["source"] == "original"
        assert original["compact"]["model"]["family"].lower() == "gaussian"
        assert original["compact"]["model"]["link"].lower() == "identity"
        assert original["compact"]["model"]["method"] == "MLE"
        assert original["compact"]["rows"]
        assert all("sig_class" in row for row in original["compact"]["rows"])

        intercept = next(row for row in original["compact"]["rows"] if row["name"] == "Intercept")
        assert intercept["edf"] == pytest.approx(1.0)
        assert intercept["se"] is not None
        assert intercept["p_value"] is not None
        assert intercept["sig_code"] in {"***", "**", "*", ".", ""}
        assert intercept["sig_class"] in {
            "sig-strong",
            "sig-medium",
            "sig-standard",
            "sig-weak",
            "sig-none",
        }

        spline_row = next(row for row in original["compact"]["rows"] if row["name"] == "x_spline")
        assert spline_row["kind"] == "spline"
        assert spline_row["stat_label"] == "chi2"
        assert spline_row["se"] is None
        assert spline_row["se_label"] == "curve"
        assert spline_row["edf"] is not None
        assert spline_row["p_value"] is not None

        _post_json(f"{widget.url}/select", {"term": "region", "indices": [1]})
        _post_json(f"{widget.url}/op", {"operation": "shift_up"})

        edited = _post_json(f"{widget.url}/summary", {"source": "in_force"})
        assert edited["available"] is True
        assert "Editor edits applied" in edited["html"]
        assert edited["compact"]["source"] == "in_force"
        assert any(row["sig_class"] == "sig-unknown" for row in edited["compact"]["rows"])

        before_refit = _post_json(f"{widget.url}/summary", {"source": "refit"})
        assert before_refit["available"] is False
        assert "No fixed-offset refit" in before_refit["error"]

        refit = _post_json(
            f"{widget.url}/refit_offset",
            {"level_display": "grouped"},
        )
        assert refit["available"] is True
        assert refit["source"] == "refit"
        assert refit["level_display"] == "grouped"
        assert refit["offset_terms"] == ["region"]
        assert "Editor offset refit" in refit["html"]
        assert refit["compact"]["source"] == "refit"
        assert refit["compact"]["offset_terms"] == ["region"]
        assert refit["compact"]["rows"]
        assert all("sig_class" in row for row in refit["compact"]["rows"])
        assert refit["offset_labels"][0]["term"] == "region"
        assert refit["offset_labels"][0]["scale"] == "log edited relativity"
        assert any(item["label"] == "B" for item in refit["offset_labels"][0]["values"])

        after_refit = _post_json(f"{widget.url}/summary", {"source": "refit"})
        assert after_refit["available"] is True
        assert after_refit["offset_terms"] == ["region"]
    finally:
        widget.close()


def test_widget_http_collapse_invalidates_stale_fixed_offset_refit(editor_model):
    session = EditorSession.from_model(editor_model, terms=["region"])
    widget = session.widget()
    try:
        _post_json(f"{widget.url}/select", {"term": "region", "indices": [1]})
        _post_json(f"{widget.url}/op", {"operation": "shift_up"})
        refit = _post_json(f"{widget.url}/refit_offset", {})
        assert refit["available"] is True

        _post_json(f"{widget.url}/select", {"term": "region", "indices": [1, 2]})
        _post_json(f"{widget.url}/collapse_levels", {"term": "region", "method": "fit"})

        stale_refit = _post_json(f"{widget.url}/summary", {"source": "refit"})
        assert stale_refit["available"] is False
        assert "No fixed-offset refit" in stale_refit["error"]
    finally:
        widget.close()


def test_widget_http_reports_are_display_only_evidence_surfaces(editor_model, editor_frame):
    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline", "region"],
        train_data=(X.iloc[:300], y[:300], None),
        validation_data=(X.iloc[300:380], y[300:380], None),
        test_data=(X.iloc[380:], y[380:], None),
        cv_report={
            "method": "KFold",
            "folds": 3,
            "metric": "mean unit deviance",
            "rows": [{"fold": 1, "train_loss": 0.012, "validation_loss": 0.015}],
            "summary": [{"loss": "validation_loss", "mean": 0.015, "std": 0.001}],
            "split_loss": [{"split": "test", "loss": 0.016, "n_obs": 70}],
        },
    )
    widget = session.widget()
    try:
        validation = _post_json(f"{widget.url}/report", {"report": "validation"})
        assert validation["available"] is True
        assert validation["report"] == "validation"
        assert validation["can_run_cv"] is False
        assert [split["name"] for split in validation["splits"]] == [
            "train",
            "validation",
            "test",
        ]
        assert validation["cv_report"]["method"] == "KFold"
        assert validation["cv_report"]["summary"][0]["loss"] == "validation_loss"
        assert validation["cv_report"]["split_loss"][0]["split"] == "test"

        final = _post_json(f"{widget.url}/report", {"report": "final"})
        assert final["available"] is True
        assert final["report"] == "final"
        assert final["summary"]["source"] == "in_force"
        assert [split["name"] for split in final["splits"]] == [
            "train",
            "validation",
            "test",
        ]
    finally:
        widget.close()


def test_final_report_summary_uses_training_fit_statistics_after_edits(
    editor_model,
    editor_frame,
):
    X, y = editor_frame
    train_X = X.iloc[:37].copy()
    train_y = y[:37].copy()
    validation_X = X.iloc[-29:].copy()
    validation_y = y[-29:].copy()
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(train_X, train_y),
        validation_data=(validation_X, validation_y),
    )
    session.select_indices("x_spline", [4, 5, 6])
    session.shift("x_spline", 0.2)
    widget = session.widget()
    try:
        final = widget._report("final")
    finally:
        widget.close()

    assert final["summary"]["compact"]["model"]["n_obs"] == len(train_X)


def test_widget_http_report_sanitizes_non_finite_cv_values(editor_model, editor_frame):
    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(X.iloc[:300], y[:300], None),
        validation_data=(X.iloc[300:380], y[300:380], None),
        cv_report={
            "method": "KFold",
            "rows": [
                {
                    "k": 5,
                    "cv_improvement": float("nan"),
                    "se_val_deviance": np.float64(np.nan),
                }
            ],
        },
    )
    widget = session.widget()
    try:
        data = json.dumps({"report": "validation"}).encode("utf-8")
        request = urllib.request.Request(
            f"{widget.url}/report",
            data=data,
            method="POST",
            headers={
                "Content-Type": "application/json",
                **_editor_token_header(f"{widget.url}/report"),
            },
        )
        with urllib.request.urlopen(request, timeout=5) as response:
            body = response.read().decode("utf-8")

        assert "NaN" not in body
        payload = json.loads(body)
        row = payload["cv_report"]["rows"][0]
        assert row["cv_improvement"] is None
        assert row["se_val_deviance"] is None
    finally:
        widget.close()


def test_widget_http_save_model_writes_edited_joblib(editor_model, tmp_path):
    import joblib

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    original_beta = editor_model.result.beta.copy()
    session.select_indices("x_spline", [4, 5, 6])
    session.shift("x_spline", 0.2)
    widget = session.widget()
    try:
        payload = _post_json(
            f"{widget.url}/save_model",
            {"directory": str(tmp_path), "filename": "edited-model"},
        )
    finally:
        widget.close()

    path = Path(payload["path"])
    assert path == tmp_path / "edited-model.joblib"
    assert path.exists()
    saved_model = joblib.load(path)
    assert saved_model is not editor_model
    np.testing.assert_allclose(editor_model.result.beta, original_beta)
    assert not np.allclose(saved_model.result.beta, original_beta)


def test_save_model_validation_failure_creates_no_file(
    editor_model,
    editor_frame,
    tmp_path,
    monkeypatch,
):
    import superglm.editor.persistence as persistence

    X, y = editor_frame
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        validation_data=(X.iloc[-31:], y[-31:]),
    )
    target = tmp_path / "new-parent" / "edited-model.joblib"
    seen_dataset = None

    def fail_validation(model, *, dataset=None, max_rows=512):
        nonlocal seen_dataset
        seen_dataset = dataset
        raise ValueError("artifact validation failed: forced failure")

    monkeypatch.setattr(persistence, "serialize_validated_model", fail_validation)

    with pytest.raises(ValueError, match="forced failure"):
        session.save_model(target)

    assert not target.exists()
    assert not target.parent.exists()
    assert seen_dataset is not None
    assert seen_dataset.name == "validation"


def test_save_model_uses_session_train_data_without_retained_fit_state(tmp_path):
    import joblib

    rng = np.random.default_rng(20260718)
    n = 180
    X = pd.DataFrame({"x": rng.normal(size=n)})
    y = 0.4 + 0.3 * X["x"].to_numpy() + rng.normal(0.0, 0.04, size=n)
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["x"], train_data=(X, y, None))
    session.select_indices("x", [0])
    session.shift("x", 0.2)

    path = session.save_model(tmp_path / "edited-model.joblib")
    saved_model = joblib.load(path)

    assert saved_model._fit_stats is not None
    assert saved_model.summary()["deviance"]["deviance"] is not None


def test_save_model_uses_validation_data_without_retained_fit_state(tmp_path):
    import joblib

    rng = np.random.default_rng(20260731)
    n_train = 180
    n_val = 70
    X_train = pd.DataFrame({"x": rng.normal(size=n_train)})
    y_train = 0.4 + 0.3 * X_train["x"].to_numpy() + rng.normal(0.0, 0.04, size=n_train)
    X_val = pd.DataFrame({"x": rng.normal(size=n_val)})
    y_val = 0.4 + 0.3 * X_val["x"].to_numpy() + rng.normal(0.0, 0.04, size=n_val)
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X_train, y_train)
    session = EditorSession.from_model(model, terms=["x"], validation_data=(X_val, y_val, None))
    session.select_indices("x", [0])
    session.shift("x", 0.2)

    path = session.save_model(tmp_path / "edited-model.joblib")
    saved_model = joblib.load(path)

    assert saved_model._fit_stats is not None
    assert saved_model._fit_state.retained is False
    for name in (
        "_fit_X_ref",
        "_fit_y_ref",
        "_fit_sample_weight_ref",
        "_fit_offset_ref",
        "_fit_weights",
        "_fit_offset",
        "_fit_mu",
        "_fit_null_mu",
        "_fit_data_guard",
    ):
        assert getattr(saved_model, name) is None, name
    assert saved_model.summary()["deviance"]["deviance"] is not None


def test_widget_http_download_model_returns_joblib_attachment(editor_model):
    import io

    import joblib

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    original_beta = editor_model.result.beta.copy()
    session.select_indices("x_spline", [4, 5, 6])
    session.shift("x_spline", 0.2)
    widget = session.widget()
    try:
        request = urllib.request.Request(
            f"{widget.url}/download_model?filename=edited-model.joblib",
            headers=_editor_token_header(widget.url),
        )
        with urllib.request.urlopen(request, timeout=5) as response:
            payload = response.read()
            content_type = response.headers["content-type"]
            disposition = response.headers["content-disposition"]
    finally:
        widget.close()

    downloaded_model = joblib.load(io.BytesIO(payload))
    assert content_type == "application/octet-stream"
    assert 'filename="edited-model.joblib"' in disposition
    assert downloaded_model is not editor_model
    np.testing.assert_allclose(editor_model.result.beta, original_beta)
    assert not np.allclose(downloaded_model.result.beta, original_beta)


def test_training_export_dataset_prefers_explicit_train_over_retained_and_validation(
    editor_model,
    editor_frame,
):
    from superglm.editor.evaluation import training_export_dataset

    X, y = editor_frame
    explicit_X = X.iloc[:37].copy()
    explicit_y = y[:37].copy()
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(explicit_X, explicit_y),
        validation_data=(X.iloc[-29:], y[-29:]),
    )

    dataset = training_export_dataset(session)

    assert dataset is not None
    assert dataset.name == "train"
    assert dataset.source == "supplied"
    assert dataset.X is explicit_X
    assert dataset.y is explicit_y


def test_widget_http_download_export_joblib_is_validated_and_revision_pinned(
    editor_model,
    editor_frame,
):
    import io

    import joblib

    X, _ = editor_frame
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    session.select_indices("x_spline", [4, 5, 6])
    session.shift("x_spline", 0.2)
    expected_revision = session.model_revision
    widget = session.widget()
    try:
        request = urllib.request.Request(
            f"{widget.url}/download_export?format=model&filename=pricing-model",
            headers=_editor_token_header(widget.url),
        )
        with urllib.request.urlopen(request, timeout=15) as response:
            data = response.read()
            headers = response.headers
    finally:
        widget.close()

    downloaded = joblib.load(io.BytesIO(data))
    assert headers["content-type"] == "application/octet-stream"
    assert 'filename="pricing-model.joblib"' in headers["content-disposition"]
    assert headers["x-superglm-model-revision"] == str(expected_revision)
    assert headers["x-superglm-validation"] == "artifact+predictions"
    predictions = downloaded.predict(X.iloc[:17])
    assert np.all(np.isfinite(predictions))
    assert not np.allclose(predictions, editor_model.predict(X.iloc[:17]))


def test_widget_http_download_export_excel_has_structured_summary(editor_model):
    import io

    from openpyxl import load_workbook

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        request = urllib.request.Request(
            f"{widget.url}/download_export?format=excel&filename=rating-output",
            headers=_editor_token_header(widget.url),
        )
        with urllib.request.urlopen(request, timeout=30) as response:
            data = response.read()
            headers = response.headers
    finally:
        widget.close()

    workbook = load_workbook(io.BytesIO(data), data_only=True)
    summary = workbook["Model Summary"]
    assert headers["content-type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert 'filename="rating-output.xlsx"' in headers["content-disposition"]
    assert headers["x-superglm-model-revision"] == "0"
    assert headers.get("x-superglm-validation") is None
    assert "ModelOverview" in summary.tables
    assert "TermInference" in summary.tables
    assert summary["A4"].value == "Section"
    assert summary["B4"].value == "Metric"
    assert summary["C4"].value == "Value"


def test_excel_export_materializes_edited_model_on_training_split(editor_model, editor_frame):
    import io

    from openpyxl import load_workbook

    X, y = editor_frame
    train_X = X.iloc[:37].copy()
    train_y = y[:37].copy()
    validation_X = X.iloc[-29:].copy()
    validation_y = y[-29:].copy()
    session = EditorSession.from_model(
        editor_model,
        terms=["x_spline"],
        train_data=(train_X, train_y),
        validation_data=(validation_X, validation_y),
    )
    session.select_indices("x_spline", [4, 5, 6])
    session.shift("x_spline", 0.2)
    widget = session.widget()
    try:
        result = widget._export_bytes("xlsx")
    finally:
        widget.close()

    workbook = load_workbook(io.BytesIO(result.data), data_only=True)
    summary = workbook["Model Summary"]
    overview = summary.tables["ModelOverview"]
    rows = list(summary[overview.ref])
    headers = [cell.value for cell in rows[0]]
    records = [dict(zip(headers, (cell.value for cell in row))) for row in rows[1:]]
    observations = next(row["Value"] for row in records if row["Metric"] == "Observations")

    assert observations == len(train_X)


def test_widget_http_download_export_excel_without_training_data_is_400():
    rng = np.random.default_rng(20260802)
    X = pd.DataFrame({"x": rng.normal(size=90)})
    y = 0.2 + 0.4 * X["x"].to_numpy() + rng.normal(0.0, 0.03, size=len(X))
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"x": Numeric()},
    ).fit(X, y)
    session = EditorSession.from_model(
        model,
        terms=["x"],
        validation_data=(X.iloc[:20], y[:20]),
    )
    widget = session.widget()
    try:
        request = urllib.request.Request(
            f"{widget.url}/download_export?format=xlsx",
            headers=_editor_token_header(widget.url),
        )
        with pytest.raises(urllib.error.HTTPError) as error:
            urllib.request.urlopen(request, timeout=10)
        payload = json.loads(error.value.read().decode("utf-8"))
    finally:
        widget.close()

    assert error.value.code == 400
    assert "train_data" in payload["error"]
    assert "retained fit data" in payload["error"]


def test_widget_export_file_writes_joblib_and_excel(editor_model, tmp_path):
    import joblib
    from openpyxl import load_workbook

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        model_payload = _post_json(
            f"{widget.url}/export_file",
            {"format": "joblib", "directory": str(tmp_path), "filename": "pricing-model"},
        )
        excel_payload = _post_json(
            f"{widget.url}/export_file",
            {"format": "xlsx", "directory": str(tmp_path), "filename": "rating-output"},
        )
    finally:
        widget.close()

    saved_model = joblib.load(model_payload["path"])
    workbook = load_workbook(excel_payload["path"], read_only=False)
    assert saved_model.result is not None
    assert model_payload["validation_scope"] == "artifact+predictions"
    assert model_payload["model_revision"] == 0
    assert Path(model_payload["path"]).name == "pricing-model.joblib"
    assert Path(excel_payload["path"]).name == "rating-output.xlsx"
    assert "Model Summary" in workbook.sheetnames


def test_widget_export_file_rejects_invalid_format_and_unsafe_filename(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        with pytest.raises(ValueError, match="Unsupported export format"):
            widget._export_bytes("pdf")
        with pytest.raises(ValueError, match="directory separators"):
            widget._export_bytes("joblib", "../outside.joblib")
        with pytest.raises(ValueError, match="extension"):
            widget._export_bytes("joblib", "model.xlsx")
    finally:
        widget.close()


@pytest.mark.parametrize("filename", ["bad\nname.joblib", "bad\rname.joblib", "bad\x7fname.joblib"])
def test_widget_export_rejects_filename_control_characters(editor_model, filename):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        with pytest.raises(ValueError, match="control characters"):
            widget._export_bytes("joblib", filename)
    finally:
        widget.close()


def test_widget_download_export_encodes_quotes_and_unicode_in_disposition(editor_model):
    filename = 'analyst "mödel".joblib'
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        query = urllib.parse.urlencode({"format": "joblib", "filename": filename})
        request = urllib.request.Request(
            f"{widget.url}/download_export?{query}",
            headers=_editor_token_header(widget.url),
        )
        with urllib.request.urlopen(request, timeout=10) as response:
            disposition = response.headers["content-disposition"]
    finally:
        widget.close()

    assert 'filename="analyst _model_.joblib"' in disposition
    assert "filename*=UTF-8''analyst%20%22m%C3%B6del%22.joblib" in disposition


def test_widget_export_file_discards_superseded_build(
    editor_model,
    tmp_path,
    monkeypatch,
):
    from superglm.editor import widget as widget_module

    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    target = tmp_path / "superseded.joblib"
    original = widget_module.persistence.serialize_validated_model

    def supersede_during_build(*args, **kwargs):
        result = original(*args, **kwargs)
        session._advance_model_revision()
        return result

    monkeypatch.setattr(
        widget_module.persistence,
        "serialize_validated_model",
        supersede_during_build,
    )
    try:
        with pytest.raises(RuntimeError, match="superseded"):
            widget._export_file("joblib", path=str(target))
    finally:
        widget.close()

    assert not target.exists()


def test_widget_http_download_model_uses_session_train_data_without_retained_fit_state():
    import io

    import joblib

    rng = np.random.default_rng(20260719)
    n = 180
    X = pd.DataFrame({"x": rng.normal(size=n)})
    y = 0.4 + 0.3 * X["x"].to_numpy() + rng.normal(0.0, 0.04, size=n)
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X, y)
    session = EditorSession.from_model(model, terms=["x"], train_data=(X, y, None))
    session.select_indices("x", [0])
    session.shift("x", 0.2)
    widget = session.widget()
    try:
        request = urllib.request.Request(
            f"{widget.url}/download_model?filename=edited-model.joblib",
            headers=_editor_token_header(widget.url),
        )
        with urllib.request.urlopen(request, timeout=5) as response:
            payload = response.read()
    finally:
        widget.close()

    downloaded_model = joblib.load(io.BytesIO(payload))
    assert downloaded_model._fit_stats is not None
    assert downloaded_model.summary()["deviance"]["deviance"] is not None


def test_widget_http_download_model_uses_validation_data_without_retained_fit_state():
    import io

    import joblib

    rng = np.random.default_rng(20260801)
    n_train = 180
    n_val = 70
    X_train = pd.DataFrame({"x": rng.normal(size=n_train)})
    y_train = 0.4 + 0.3 * X_train["x"].to_numpy() + rng.normal(0.0, 0.04, size=n_train)
    X_val = pd.DataFrame({"x": rng.normal(size=n_val)})
    y_val = 0.4 + 0.3 * X_val["x"].to_numpy() + rng.normal(0.0, 0.04, size=n_val)
    model = SuperGLM(
        family="gaussian",
        retain_fit_state=False,
        selection_penalty=0.0,
        features={"x": Numeric()},
    )
    model.fit(X_train, y_train)
    session = EditorSession.from_model(model, terms=["x"], validation_data=(X_val, y_val, None))
    session.select_indices("x", [0])
    session.shift("x", 0.2)
    widget = session.widget()
    try:
        request = urllib.request.Request(
            f"{widget.url}/download_model?filename=edited-model.joblib",
            headers=_editor_token_header(widget.url),
        )
        with urllib.request.urlopen(request, timeout=5) as response:
            payload = response.read()
    finally:
        widget.close()

    downloaded_model = joblib.load(io.BytesIO(payload))
    assert downloaded_model._fit_stats is not None
    assert downloaded_model._fit_state.retained is False
    for name in (
        "_fit_X_ref",
        "_fit_y_ref",
        "_fit_sample_weight_ref",
        "_fit_offset_ref",
        "_fit_weights",
        "_fit_offset",
        "_fit_mu",
        "_fit_null_mu",
        "_fit_data_guard",
    ):
        assert getattr(downloaded_model, name) is None, name
    assert downloaded_model.summary()["deviance"]["deviance"] is not None


def test_widget_http_open_directory_launches_current_folder(editor_model, tmp_path, monkeypatch):
    from superglm.editor import widget as widget_module

    opened: list[str] = []

    def fake_open_directory(path):
        opened.append(str(path))
        return Path(path)

    monkeypatch.setattr(widget_module, "open_directory_path", fake_open_directory)
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        payload = _post_json(f"{widget.url}/open_directory", {"path": str(tmp_path)})
    finally:
        widget.close()

    assert payload == {"path": str(tmp_path.resolve())}
    assert opened == [str(tmp_path.resolve())]


def test_open_directory_prefers_dolphin_on_kde(tmp_path, monkeypatch):
    from superglm.editor import native_dialogs

    commands = []

    class FakeProcess:
        def __init__(self, command):
            self.command = command

        def poll(self):
            return None

        def communicate(self, timeout=None):
            return "", ""

    monkeypatch.setattr(native_dialogs.sys, "platform", "linux")
    monkeypatch.setattr(native_dialogs, "_is_wsl", lambda: False)
    monkeypatch.setenv("XDG_CURRENT_DESKTOP", "KDE")
    monkeypatch.setattr(native_dialogs.shutil, "which", lambda command: f"/usr/bin/{command}")
    monkeypatch.setattr(
        native_dialogs.subprocess,
        "Popen",
        lambda command, **_kwargs: commands.append(tuple(command)) or FakeProcess(command),
    )

    opened = native_dialogs.open_directory_path(tmp_path)

    assert opened == tmp_path.resolve()
    assert commands[0] == ("dolphin", "--new-window", str(tmp_path.resolve()))


def test_open_directory_detects_wsl_from_environment(monkeypatch):
    from superglm.editor import native_dialogs

    monkeypatch.setenv("WSL_DISTRO_NAME", "Ubuntu")

    assert native_dialogs._is_wsl()


def test_open_directory_detects_wsl_from_proc_version(monkeypatch):
    from superglm.editor import native_dialogs

    original_read_text = Path.read_text

    def fake_read_text(self, *args, **kwargs):
        if str(self) == "/proc/version":
            return "Linux version 6.6.87.2-microsoft-standard-WSL2"
        return original_read_text(self, *args, **kwargs)

    monkeypatch.delenv("WSL_DISTRO_NAME", raising=False)
    monkeypatch.setattr(native_dialogs.Path, "read_text", fake_read_text)

    assert native_dialogs._is_wsl()


def test_open_directory_uses_windows_explorer_inside_wsl(tmp_path, monkeypatch):
    from superglm.editor import native_dialogs

    commands = []
    windows_path = r"\\wsl.localhost\Ubuntu\home\user\project"

    class FakeProcess:
        def poll(self):
            return None

        def communicate(self, timeout=None):
            return "", ""

    def fake_run(command, **kwargs):
        assert tuple(command) == ("wslpath", "-w", str(tmp_path.resolve()))
        assert kwargs["check"] is True
        assert kwargs["capture_output"] is True
        assert kwargs["text"] is True
        return SimpleNamespace(stdout=f"{windows_path}\n")

    monkeypatch.setattr(native_dialogs.sys, "platform", "linux")
    monkeypatch.setenv("WSL_DISTRO_NAME", "Ubuntu")
    monkeypatch.setattr(native_dialogs.subprocess, "run", fake_run)
    monkeypatch.setattr(
        native_dialogs.subprocess,
        "Popen",
        lambda command, **_kwargs: commands.append(tuple(command)) or FakeProcess(),
    )

    opened = native_dialogs.open_directory_path(tmp_path)

    assert opened == tmp_path.resolve()
    assert commands == [("explorer.exe", windows_path)]


def test_open_directory_falls_back_to_linux_opener_when_wsl_explorer_fails(tmp_path, monkeypatch):
    from superglm.editor import native_dialogs

    commands = []

    class FailedProcess:
        def poll(self):
            return 1

        def communicate(self, timeout=None):
            return "", "explorer failed"

    class RunningProcess:
        def poll(self):
            return None

        def communicate(self, timeout=None):
            return "", ""

    def fake_popen(command, **_kwargs):
        commands.append(tuple(command))
        if command[0] == "explorer.exe":
            return FailedProcess()
        return RunningProcess()

    monkeypatch.setattr(native_dialogs.sys, "platform", "linux")
    monkeypatch.setenv("WSL_DISTRO_NAME", "Ubuntu")
    monkeypatch.setattr(
        native_dialogs.subprocess,
        "run",
        lambda *_args, **_kwargs: SimpleNamespace(stdout="C:\\Users\\me\\project\n"),
    )
    monkeypatch.setenv("XDG_CURRENT_DESKTOP", "")
    monkeypatch.setattr(native_dialogs.shutil, "which", lambda command: "/usr/bin/xdg-open")
    monkeypatch.setattr(native_dialogs.subprocess, "Popen", fake_popen)

    opened = native_dialogs.open_directory_path(tmp_path)

    assert opened == tmp_path.resolve()
    assert commands == [
        ("explorer.exe", "C:\\Users\\me\\project"),
        ("xdg-open", str(tmp_path.resolve())),
    ]


def test_open_directory_raises_when_launcher_exits_immediately(tmp_path, monkeypatch):
    from superglm.editor import native_dialogs

    class FailedProcess:
        def poll(self):
            return 1

        def communicate(self, timeout=None):
            return "", "could not open"

    monkeypatch.setattr(native_dialogs.sys, "platform", "linux")
    monkeypatch.setenv("XDG_CURRENT_DESKTOP", "KDE")
    monkeypatch.setattr(native_dialogs.shutil, "which", lambda command: "/usr/bin/dolphin")
    monkeypatch.setattr(
        native_dialogs.subprocess, "Popen", lambda *_args, **_kwargs: FailedProcess()
    )

    with pytest.raises(RuntimeError, match="could not open"):
        native_dialogs.open_directory_path(tmp_path)


def test_widget_save_directory_defaults_to_cwd(editor_model, tmp_path, monkeypatch):
    child = tmp_path / "models"
    child.mkdir()
    file_path = tmp_path / "existing-model.joblib"
    file_path.write_text("placeholder")
    monkeypatch.chdir(tmp_path)
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        payload = widget._save_directory()
    finally:
        widget.close()

    assert payload["cwd"] == str(tmp_path.resolve())
    assert payload["path"] == str(tmp_path.resolve())
    assert {
        "kind": "directory",
        "name": "models",
        "path": str(child.resolve()),
    } in payload["entries"]
    assert {
        "kind": "file",
        "name": "existing-model.joblib",
        "path": str(file_path.resolve()),
    } in payload["entries"]
    assert [entry["kind"] for entry in payload["entries"]] == ["directory", "file"]


def test_widget_reset_and_undo_target_current_term_after_switching(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline", "region"])
    widget = session.widget()
    try:
        spline = session.terms["x_spline"]
        region = session.terms["region"]
        spline_original = spline.original_log_effect.copy()
        region_original = region.original_log_effect.copy()

        _post_json(f"{widget.url}/select", {"term": "x_spline", "indices": [4, 5]})
        _post_json(f"{widget.url}/op", {"operation": "shift_up"})
        _post_json(f"{widget.url}/select", {"term": "region", "indices": [1]})
        _post_json(f"{widget.url}/op", {"operation": "shift_down"})
        _post_json(f"{widget.url}/term", {"term": "x_spline"})

        _post_json(f"{widget.url}/op", {"operation": "undo"})
        np.testing.assert_allclose(spline.edited_log_effect, spline_original)
        assert region.edited_log_effect[1] == pytest.approx(region_original[1] + np.log(0.95))

        _post_json(f"{widget.url}/op", {"operation": "shift_up"})
        _post_json(f"{widget.url}/term", {"term": "region"})
        _post_json(f"{widget.url}/term", {"term": "x_spline"})
        _post_json(f"{widget.url}/op", {"operation": "reset"})
        np.testing.assert_allclose(spline.edited_log_effect, spline_original)
    finally:
        widget.close()


def test_widget_app_shell_contains_drag_editor(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        with urllib.request.urlopen(widget.url, timeout=5) as response:
            shell = response.read().decode("utf-8")

        module_sources = []
        for asset in ["main.js", "api/client.js", "interactions.js", "summary.js"]:
            request = urllib.request.Request(f"{widget.url}/assets/{asset}", method="GET")
            with urllib.request.urlopen(request, timeout=5) as response:
                module_sources.append(response.read().decode("utf-8"))
        js = "\n".join(module_sources)

        assert '<script type="module" src="/assets/main.js"></script>' in shell
        assert 'role="tablist" aria-label="Editor views"' in shell
        assert 'id="chart"' in shell
        assert 'id="selectionMenu"' in shell
        assert 'id="inspector"' in shell
        assert "X-SuperGLM-Editor-Token" in js
        assert "/drag" in js
        assert "/control" in js
        assert "/metrics" in js
        assert "/summary" in js
    finally:
        widget.close()


def test_editor_frontend_includes_grouped_display_control_and_view_mapping():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    html = (root / "index.html").read_text()
    main_js = (root / "main.js").read_text()
    chart_js = (root / "chart.js").read_text()
    css = (root / "styles.css").read_text()

    assert 'id="groupDisplayMode"' in html
    assert 'id="groupDisplayWrap" class="toolbar-field group-display-toggle"' in html
    assert 'id="groupDisplayWrap" class="toolbar-field group-display-toggle" hidden' not in html
    assert "groupModeByTerm" in main_js
    assert "groupDisplayMode.disabled = !available" in main_js
    assert "groupDisplayWrap.hidden = !available" not in main_js
    assert "resolveDisplayTerm" in chart_js
    assert "source_indices" in chart_js
    assert "displayToSourceIndices" in chart_js
    assert ".group-display-toggle" in css


def test_editor_summary_has_view_only_categorical_levels_control():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    html = (root / "index.html").read_text()
    main_js = (root / "main.js").read_text()

    assert 'id="summaryLevelDisplay"' in html
    assert "<legend>Categorical levels</legend>" in html
    assert html.count('type="radio" name="summaryLevelDisplay"') == 2
    assert 'value="expanded" checked' in html
    assert 'value="grouped"' in html
    assert "<span>Expanded</span>" in html
    assert "<span>Grouped</span>" in html

    binding_start = main_js.index("for (const input of summaryLevelDisplayInputs)")
    binding_end = main_js.index("\n}", binding_start) + 2
    binding_source = main_js[binding_start:binding_end]
    assert "actions.patchView({ summaryLevelDisplay: input.value })" in binding_source
    assert "actions.schedulePanelEvidence(" in binding_source
    assert '"summary"' in binding_source
    assert "{ immediate: true }" in binding_source
    assert "refreshSummaryView" not in binding_source
    assert "runStructuralRefit" not in binding_source
    assert "ungroupTransition" not in binding_source

    assert "function summaryRequestPayload()" in main_js
    assert "level_display: selectSummaryLevelDisplay(store.getState())" in main_js
    assert main_js.count("summaryRequestPayload()") >= 3
    assert "get summaryLevelDisplay()" in main_js
    assert "return selectSummaryLevelDisplay(store.getState())" in main_js
    assert "payload: {" in main_js
    assert "...descriptor.payload," in main_js
    assert "level_display: selectSummaryLevelDisplay(store.getState())" in main_js
    assert 'payload.level_display || "expanded"' in main_js
    assert 'summaryFrame.innerHTML = ""' in main_js


def test_editor_interactions_map_group_display_indices_to_source_indices():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    interactions_js = (root / "interactions.js").read_text()
    chart_js = (root / "chart.js").read_text()

    assert "sourceIndicesForDisplayIndex" in interactions_js
    assert "sourceIndicesForDisplayIndices" in interactions_js
    assert "valuesForSourceIndices" in interactions_js
    assert "if (!scale.displayIsCollapsed) return [index]" in interactions_js
    assert "displayToSourceIndices" in chart_js


def test_editor_collapsed_group_display_does_not_reuse_expanded_group_geometry():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    chart_js = (root / "chart.js").read_text()

    assert "drawCollapsedLevelGroups(svg, view, sx, sy)" in chart_js
    assert "else drawLevelGroups(svg, view, sx, sy)" in chart_js
    assert "term.displayIsGroup" in chart_js


def test_editor_collapsed_group_display_explains_original_projection():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    chart_js = (root / "chart.js").read_text()
    main_js = (root / "main.js").read_text()

    assert "original projection" in chart_js
    assert "collapsedOriginalNote" in main_js
    assert "original line is grouped by exposure-weighted averaging" in main_js


def test_editor_group_display_change_resets_zoom_and_blocks_collapsed_reorder():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    main_js = (root / "main.js").read_text()
    interactions_js = (root / "interactions.js").read_text()

    assert "delete zoomByTerm[term]" in main_js
    assert "scale.displayIsCollapsed" in interactions_js
    assert "if (scale.displayIsCollapsed) return false" in interactions_js


def test_editor_chart_renders_previous_edit_line():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    chart_js = (root / "chart.js").read_text()
    css = (root / "styles.css").read_text()

    assert "previous_y" in chart_js
    assert "previous edit" in chart_js
    assert "previous-edit" in chart_js
    assert ".previous-edit" in css
    assert "#f59e0b" in css


def test_editor_axis_uses_display_only_measured_geometry():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    chart_js = (root / "chart.js").read_text()
    geometry_js = (root / "chart/geometry.js").read_text()

    assert "measureCategoricalLabels" in chart_js
    assert "data-full-label" in chart_js
    assert "axisMeasurementCount" in chart_js
    assert "planCategoricalAxis" in chart_js
    assert "splitLabelGraphemes" in chart_js
    assert "fitMeasuredLabel" in geometry_js
    assert "term.levels[" not in geometry_js
    assert "MAX_LEVEL_LABELS" not in chart_js


def test_editor_chart_points_have_relativity_exposure_tooltips():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    chart_js = (root / "chart.js").read_text()
    css = (root / "styles.css").read_text()

    assert "pointTooltipLines(term, i)" in chart_js
    assert "Relativity: ${fmt(term.y[i])}" in chart_js
    assert "Exposure: ${fmt(exposure)}" in chart_js
    assert "showPointTooltip(svg, circle, pointTooltipLines(term, i))" in chart_js
    assert "hidePointTooltip(svg)" in chart_js
    assert "point-tooltip" in css


def test_editor_structural_refits_show_busy_overlay_and_timing_debug():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    html = (root / "index.html").read_text()
    main_js = (root / "main.js").read_text()
    actions_js = (root / "state/actions.js").read_text()
    timing_js = (root / "state/timing.js").read_text()
    summary_js = (root / "summary.js").read_text()
    css = (root / "styles.css").read_text()

    refit_start = main_js.index("async function runStructuralRefit")
    refit_end = main_js.index("\nfunction setAppBusy", refit_start)
    refit_source = main_js[refit_start:refit_end]
    busy_start = main_js.index("function renderMutationBusy")
    busy_end = main_js.index("\nfunction renderInteractionPreview", busy_start)
    busy_source = main_js[busy_start:busy_end]
    recovery_start = main_js.index("function renderRecovery")
    recovery_end = main_js.index("\nfunction renderMetricsEvidence", recovery_start)
    recovery_source = main_js[recovery_start:recovery_end]
    bindings_start = main_js.index("if (collapseLevels) {", refit_end)
    bindings_end = main_js.index("\nloadState().then", bindings_start)
    bindings_source = main_js[bindings_start:bindings_end]
    collapse_start = summary_js.index("export function collapseTransition")
    collapse_end = summary_js.index("export function ungroupTransition", collapse_start)
    collapse_source = summary_js[collapse_start:collapse_end]
    transition_start = actions_js.index("  async function executeStructuralMutation")
    transition_end = actions_js.index("\n  /**\n   * Refresh one evidence panel", transition_start)
    transition_source = actions_js[transition_start:transition_end]

    assert 'id="appBusyOverlay"' in html
    assert "setAppBusy" in main_js
    assert "actions.executeStructuralMutation" in refit_source
    assert "onRequestSettled" in refit_source
    assert "waitForSecondary" not in refit_source
    request_end_assignment = refit_source.index("requestEnd = performance.now()")
    assert refit_source.index("onRequestSettled") < request_end_assignment
    assert "refreshMetricsView" not in refit_source
    assert "refreshActiveReport" not in refit_source
    assert transition_source.index("commitStructuralTransition") < transition_source.index(
        "await waitForPaint()"
    )
    assert transition_source.index("await waitForPaint()") < transition_source.index(
        "finishStructuralMutation(null)"
    )
    assert transition_source.index("finishStructuralMutation(null)") < transition_source.index(
        "scheduleVisibleEvidence"
    )
    assert "immediate: true" in transition_source
    assert "summaryCommitted: true" in transition_source
    assert "actions.initialize()" not in refit_source
    assert '"/state"' not in refit_source
    assert "setAppBusy" not in refit_source
    assert "refreshSummaryView" not in refit_source
    assert "result.envelope" in refit_source
    assert 'mutation.status === "running"' in busy_source
    assert "setAppBusy(active" in busy_source
    assert "appAlertRetry.hidden" in recovery_source
    assert "!visibleRecovery.retry" in recovery_source
    assert "!recovery.retry" in recovery_source
    assert "appAlertDismiss.disabled = retryInProgress" in recovery_source
    assert "clientTransitionTiming" in main_js
    for field in (
        "client_request_ms",
        "client_commit_ms",
        "client_paint_ms",
        "client_primary_ms",
        "client_total_ms",
    ):
        assert field in timing_js
    assert "client_recovery_ms" not in main_js
    assert "client_recovery_ms" not in timing_js
    assert "onPrimaryCommitted" in refit_source
    assert "onPaintSettled" in refit_source
    assert "createEvidenceTimingTracker" in main_js
    assert 'evidenceTiming.observe("metrics"' in main_js
    assert 'evidenceTiming.observe("summary"' in main_js
    assert 'evidenceTiming.observe("report"' in main_js
    assert "formatEvidenceTimingDetails" in main_js
    assert "formatTimingDetails" in main_js
    assert "Refit completed in" in main_js
    assert 'id="advancedTiming"' in html
    assert "advancedTiming.textContent" in main_js
    assert 'summaryNote.textContent = payload.note || ""' in main_js
    assert "collapseTransition" in main_js[:refit_start]
    assert "ungroupTransition" in main_js[:refit_start]
    assert "uncollapseTransition" in main_js[:refit_start]
    assert "runStructuralRefit(collapseTransition(selectedTerm()))" in bindings_source
    assert "runStructuralRefit(ungroupTransition(selectedTerm()))" in bindings_source
    assert "runStructuralRefit(uncollapseTransition())" in bindings_source
    assert "store.subscribe(selectChartRenderState" in bindings_source
    assert "sameChartRenderState" in bindings_source
    assert "(state) => state.remote.summary" in bindings_source
    assert "if (summary)" in bindings_source
    assert "(state) => state.request.evidence.summary" in bindings_source
    assert "renderStaleSummary" not in main_js
    assert "renderSummaryEvidence" in bindings_source
    assert "state.request.mutation" in bindings_source
    assert 'path: "/collapse_levels"' in collapse_source
    assert 'payload: { term, method: "auto" }' in collapse_source
    assert "requestJSON" not in collapse_source
    assert "app-busy-overlay" in css
    assert "app-shell.is-busy" in css
    assert "busy-spinner" in css


def test_editor_profile_ui_does_not_promise_an_implicit_ci_phase():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    summary_js = (root / "summary.js").read_text()

    assert 'job.phase === "profile_ci"' not in summary_js
    assert '"profile_ci"' not in summary_js


def test_editor_structural_confirmation_gate_precedes_every_refit_side_effect():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    html = (root / "index.html").read_text()
    main_js = (root / "main.js").read_text()
    confirm_js = (root / "views/structural_confirm.js").read_text()
    root_css = (root / "styles.css").read_text()
    dialog_css = (root / "styles/dialogs.css").read_text()

    refit_start = main_js.index("async function runStructuralRefit")
    refit_end = main_js.index("\nfunction setAppBusy", refit_start)
    refit_source = main_js[refit_start:refit_end]
    bindings_start = main_js.index('collapseLevels.addEventListener("click"', refit_end)
    bindings_end = main_js.index("\nstore.subscribe", bindings_start)
    bindings_source = main_js[bindings_start:bindings_end]
    impact_index = refit_source.index("structuralImpact(")
    confirm_index = refit_source.index("structuralConfirm.confirm(impact)")
    contribution_index = refit_source.index("stopContributionBuild()")
    timing_index = refit_source.index("performance.now()")
    mutation_index = refit_source.index("actions.executeStructuralMutation")

    assert main_js.count('from "./views/structural_confirm.js"') == 1
    assert "selectSnapshot" in main_js[:refit_start]
    assert "selectSnapshot(store.getState())" in refit_source
    assert impact_index < confirm_index < contribution_index < timing_index < mutation_index
    assert 'summarySource.value = "selected"' in refit_source
    assert "stopContributionBuild()" not in bindings_source
    assert "summarySource.value" not in bindings_source

    assert html.count('id="structuralConfirmDialog"') == 1
    assert (
        '<dialog id="structuralConfirmDialog" class="structural-confirm" '
        'aria-labelledby="structuralConfirmTitle" '
        'aria-describedby="structuralConfirmMessage">'
    ) in html.replace("\n", " ").replace("  ", " ")
    assert '<form method="dialog">' in html
    assert '<h2 id="structuralConfirmTitle">Confirm structural refit</h2>' in html
    assert '<p id="structuralConfirmMessage"></p>' in html
    assert '<button value="cancel" type="submit">Cancel</button>' in html
    assert "Continue and refit</button>" in html
    assert html.index("/assets/styles/panels.css") < html.index("/assets/styles/dialogs.css")

    assert ".profile-dialog" not in root_css
    assert ".export-dialog" not in root_css
    assert ".profile-dialog" in dialog_css
    assert ".export-dialog" in dialog_css
    assert ".structural-confirm" in dialog_css
    assert ".dialog-actions" in dialog_css
    assert "overflow-wrap" in dialog_css
    assert "@media" in dialog_css
    assert "textContent" in confirm_js
    assert "innerHTML" not in confirm_js


def test_editor_inspector_has_summary_history_advanced_and_help_tabs():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    html = (root / "index.html").read_text()
    main_js = (root / "main.js").read_text()
    css = (root / "styles/panels.css").read_text()
    history_js_path = root / "history.js"

    assert 'aria-label="Model inspector"' in html
    assert ">Summary</button>" in html
    assert ">History</button>" in html
    assert ">Advanced</button>" in html
    assert ">Help</button>" in html
    assert "historyFrame" in html
    assert "advancedTiming" in html
    assert html.count('id="buildDurationWrap"') == 1
    assert 'from "./history.js"' in main_js
    assert ".inspector" in css
    assert ".help-pane" in css
    assert history_js_path.exists()


def test_editor_history_module_renders_hashes_and_redo_stack():
    root = Path(__file__).resolve().parents[1] / "src/superglm/editor/app"
    history_js_path = root / "history.js"

    assert history_js_path.exists()
    source = history_js_path.read_text()

    assert "renderHistory" in source
    assert "Redo stack" in source
    assert "history-hash" in source


def test_widget_serves_editor_app_assets(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        with urllib.request.urlopen(widget.url, timeout=5) as response:
            shell = response.read().decode("utf-8")
        assert '<link rel="stylesheet" href="/assets/styles/tokens.css">' in shell
        assert '<link rel="stylesheet" href="/assets/styles.css">' in shell
        assert '<link rel="stylesheet" href="/assets/styles/shell.css">' in shell
        assert '<link rel="stylesheet" href="/assets/styles/chart.css">' in shell
        assert '<link rel="stylesheet" href="/assets/styles/panels.css">' in shell
        assert '<link rel="stylesheet" href="/assets/styles/dialogs.css">' in shell
        assert '<script type="module" src="/assets/main.js"></script>' in shell
        assert "<style>" not in shell
        assert "<script>\nconst svg" not in shell

        css_request = urllib.request.Request(f"{widget.url}/assets/styles.css", method="GET")
        with urllib.request.urlopen(css_request, timeout=5) as response:
            assert response.headers["Content-Type"].startswith("text/css")
            css = response.read().decode("utf-8")
        assert ".toolbar" in css
        assert ".summary-frame" in css
        assert ".metric-grid" in css

        js_request = urllib.request.Request(f"{widget.url}/assets/main.js", method="GET")
        with urllib.request.urlopen(js_request, timeout=5) as response:
            assert response.headers["Content-Type"].startswith("application/javascript")
            js = response.read().decode("utf-8")
        assert "loadState" in js
        assert "drawChart" in js
        assert "runOffsetRefit" in js

        for asset in [
            "api.js",
            "api/client.js",
            "api/contracts.js",
            "state/store.js",
            "state/selectors.js",
            "state/actions.js",
            "state/timing.js",
            "format.js",
            "chart.js",
            "chart/geometry.js",
            "metrics.js",
            "summary.js",
            "reports.js",
            "interactions.js",
            "views/inspector.js",
            "views/help_drawer.js",
            "views/structural_confirm.js",
        ]:
            request = urllib.request.Request(f"{widget.url}/assets/{asset}", method="GET")
            with urllib.request.urlopen(request, timeout=5) as response:
                assert response.headers["Content-Type"].startswith("application/javascript")
                assert response.read()

        for asset in [
            "styles/tokens.css",
            "styles/shell.css",
            "styles/chart.css",
            "styles/panels.css",
            "styles/dialogs.css",
        ]:
            request = urllib.request.Request(f"{widget.url}/assets/{asset}", method="GET")
            with urllib.request.urlopen(request, timeout=5) as response:
                assert response.headers["Content-Type"].startswith("text/css")
                assert response.read()

        assert 'from "./chart.js"' in js
        assert 'from "./api/client.js"' in js
        assert 'from "./state/actions.js"' in js
        assert 'from "./state/selectors.js"' in js
        assert 'from "./state/store.js"' in js
        assert 'from "./metrics.js"' in js
        assert 'from "./summary.js"' in js
        assert 'from "./interactions.js"' in js
    finally:
        widget.close()


def test_widget_unknown_route_uses_editor_error_shape(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()
    try:
        request = urllib.request.Request(
            f"{widget.url}/missing-route",
            headers=_editor_token_header(f"{widget.url}/missing-route"),
        )
        with pytest.raises(urllib.error.HTTPError) as error:
            urllib.request.urlopen(request, timeout=5)

        assert error.value.code == 404
        payload = json.loads(error.value.read().decode("utf-8"))
        assert payload == {"error": "not found"}
    finally:
        widget.close()


def test_widget_http_unexpected_errors_do_not_expose_exception_text(editor_model):
    session = EditorSession.from_model(editor_model, terms=["x_spline"])
    widget = session.widget()

    def fail_operation(*_args, **_kwargs):
        raise RuntimeError("secret traceback detail")

    widget._operate = fail_operation
    try:
        data = json.dumps({"operation": "reset"}).encode("utf-8")
        request = urllib.request.Request(
            f"{widget.url}/op",
            data=data,
            method="POST",
            headers={
                "Content-Type": "application/json",
                **_editor_token_header(widget.url),
            },
        )
        with pytest.raises(urllib.error.HTTPError) as error:
            urllib.request.urlopen(request, timeout=5)

        assert error.value.code == 500
        payload = json.loads(error.value.read().decode("utf-8"))
        assert payload == {"error": "internal editor error"}
    finally:
        widget.close()


def test_editor_server_declares_fastapi_routes():
    from superglm.editor.server import create_editor_app

    class DummyWidget:
        pass

    app = create_editor_app(DummyWidget())
    routes = {
        (route.path, frozenset(route.methods or set()))
        for route in app.routes
        if getattr(route, "include_in_schema", True)
    }

    assert ("/state", frozenset({"GET"})) in routes
    assert ("/health", frozenset({"GET"})) in routes
    assert ("/term", frozenset({"POST"})) in routes
    assert ("/select", frozenset({"POST"})) in routes
    assert ("/op", frozenset({"POST"})) in routes
    assert ("/drag", frozenset({"POST"})) in routes
    assert ("/control", frozenset({"POST"})) in routes
    assert ("/control_count", frozenset({"POST"})) in routes
    assert ("/metrics", frozenset({"POST"})) in routes
    assert ("/summary", frozenset({"POST"})) in routes
    assert ("/report", frozenset({"POST"})) in routes
    assert ("/save_model", frozenset({"POST"})) in routes
    assert ("/download_model", frozenset({"GET"})) in routes
    assert ("/export_file", frozenset({"POST"})) in routes
    assert ("/download_export", frozenset({"GET"})) in routes
    assert ("/native_save_dialog", frozenset({"POST"})) not in routes
    assert ("/open_directory", frozenset({"POST"})) in routes
    assert ("/save_directory", frozenset({"POST"})) in routes
    assert ("/refit_offset", frozenset({"POST"})) in routes
    assert ("/profile_distribution", frozenset({"POST"})) in routes
    assert ("/profile_distribution/start", frozenset({"POST"})) in routes
    assert ("/profile_distribution/status/{job_id}", frozenset({"GET"})) in routes
    assert ("/collapse_levels", frozenset({"POST"})) in routes
    assert ("/ungroup_levels", frozenset({"POST"})) in routes
    assert ("/reorder_levels", frozenset({"POST"})) in routes
    assert ("/uncollapse_levels", frozenset({"POST"})) in routes
    assert ("/model_source", frozenset({"POST"})) not in routes


def test_summary_html_keeps_spline_significance_out_of_qs_column():
    summary = ModelSummary(
        data={"fit": {}},
        model_info={
            "family": "gaussian",
            "link": "identity",
            "penalty": "ridge",
            "n_obs": 100,
            "effective_df": 4.2,
            "phi": 1.0,
            "log_likelihood": -10.0,
            "deviance": 20.0,
            "null_deviance": 30.0,
            "pseudo_r2": 0.3,
            "aic": 40.0,
            "aicc": 41.0,
            "bic": 45.0,
            "ebic": 46.0,
            "converged": True,
            "n_iter": 3,
        },
        coef_rows=[
            _CoefRow(name="Intercept", coef=0.0, se=1.0, z=0.0, p=1.0, ci_low=-1.0, ci_high=1.0),
            _CoefRow(
                name="age",
                group="age",
                is_spline=True,
                n_params=5,
                active=True,
                wald_chi2=25.0,
                wald_p=1e-4,
                ref_df=3.0,
                edf=3.2,
            ),
        ],
        basis_detail={
            "age": [
                _BasisDetailRow(
                    parent_name="age",
                    basis_index=0,
                    coef=0.2,
                    se=0.05,
                    z=4.0,
                    p=1e-4,
                    ci_low=0.1,
                    ci_high=0.3,
                )
            ]
        },
    )

    html = summary._repr_html_()

    assert '<td style="padding:3px 4px;text-align:left;border:none;">***</td><td' in html
    assert "<td style='padding:1px 6px;'>***</td><td style='padding:1px 6px;'></td>" in html


def _get_json(url: str):
    request = urllib.request.Request(url, headers=_editor_token_header(url))
    with urllib.request.urlopen(request, timeout=5) as response:
        return json.loads(response.read().decode("utf-8"))


def _post_json(url: str, payload: dict):
    data = json.dumps(payload).encode("utf-8")
    headers = {"Content-Type": "application/json"}
    headers.update(_editor_token_header(url))
    request = urllib.request.Request(
        url,
        data=data,
        method="POST",
        headers=headers,
    )
    with urllib.request.urlopen(request, timeout=5) as response:
        return json.loads(response.read().decode("utf-8"))


def _editor_token_header(url: str) -> dict[str, str]:
    from superglm.editor.widget import _LIVE_WIDGETS

    parsed = urllib.parse.urlparse(url)
    origin = f"{parsed.scheme}://{parsed.netloc}"
    for widget in list(_LIVE_WIDGETS):
        if getattr(widget, "url", "") == origin:
            return {"X-SuperGLM-Editor-Token": widget._token}
    return {}
