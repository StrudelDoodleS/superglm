import math
import warnings
from dataclasses import replace
from io import BytesIO
from types import SimpleNamespace

import numpy as np
import pandas as pd
import polars as pl
import pytest
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

from superglm import (
    Categorical,
    FactorSmooth,
    LambdaPolicy,
    Numeric,
    OrderedCategorical,
    Polynomial,
    RandomEffect,
    Spline,
    SuperGLM,
    export_rating_tables,
)
from superglm.editor import EditorSession
from superglm.export.excel import (
    _BASE_RELATIVITY_NUMBER_FORMAT,
    _PIECEWISE_NUMBER_FORMAT,
    write_rating_table_workbook,
)
from superglm.export.rating_tables import RatingTablePayload, build_rating_table_payload
from superglm.export.summary import (
    SummaryExportPayload,
    SummaryOverviewRow,
    SummaryTermRow,
    build_summary_export_payload,
)

EXPECTED_SUMMARY_TERM_HEADERS = [
    "Term",
    "Group",
    "Kind",
    "Estimate",
    "Std Error",
    "Statistic",
    "Statistic Type",
    "P Value",
    "CI Lower",
    "CI Upper",
    "EDF",
    "Lambda",
    "Active",
    "Significance",
    "Warning",
]


def _fit_export_model():
    rng = np.random.default_rng(123)
    n = 300
    X = pd.DataFrame(
        {
            "age": rng.uniform(18, 80, n),
            "region": rng.choice(["A", "B", "C"], n),
            "score": rng.normal(0.0, 1.0, n),
        }
    )
    eta = -1.0 + 0.15 * np.sin(X["age"].to_numpy() / 8.0) + 0.2 * (X["region"] == "B")
    eta = eta + 0.05 * X["score"].to_numpy()
    y = rng.poisson(np.exp(eta)).astype(float)
    w = rng.uniform(0.5, 2.0, n)
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={
            "age": Spline(n_knots=8),
            "region": Categorical(base="first"),
            "score": Numeric(),
        },
    )
    model.fit(X, y, sample_weight=w)
    return model, X, y, w


def _fit_offset_export_model(distinct_terms: int = 2):
    rng = np.random.default_rng(987)
    n = 240
    if distinct_terms == 2:
        term = np.resize(np.array([12.0, 36.0]), n)
    else:
        term = np.linspace(12.0, 48.0, n)
    X = pd.DataFrame({"region": rng.choice(["A", "B"], n)})
    offset = np.log(term / 12.0)
    exposure = rng.uniform(0.5, 2.0, n)
    eta = -1.4 + 0.25 * (X["region"].to_numpy() == "B") + offset
    y = rng.poisson(np.exp(eta)).astype(float)
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={"region": Categorical(base="first")},
    )
    model.fit(X, y, sample_weight=exposure, offset=offset)
    return model, X, y, exposure


def _fit_term_offset_export_model(distinct_terms: int = 2, *, retain_fit_state: bool = True):
    rng = np.random.default_rng(988)
    n = 240
    if distinct_terms == 2:
        term = np.resize(np.array([12.0, 36.0]), n)
    else:
        term = np.linspace(12.0, 48.0, n)
    X = pd.DataFrame({"region": rng.choice(["A", "B"], n)})
    offset = np.log(term / 12.0)
    exposure = rng.uniform(0.5, 2.0, n)
    eta = -1.35 + 0.2 * (X["region"].to_numpy() == "B") + offset
    y = rng.poisson(np.exp(eta)).astype(float)
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        retain_fit_state=retain_fit_state,
        features={"region": Categorical(base="first")},
    )
    model.fit(X, y, sample_weight=exposure, offset=offset)
    return model, X, y, exposure, term, offset


def _overview_values(payload: SummaryExportPayload) -> dict[tuple[str, str], object]:
    return {(row.section, row.metric): row.value for row in payload.overview}


def _workbook_payload(summary: SummaryExportPayload) -> RatingTablePayload:
    return RatingTablePayload(
        base_relativity=1.0,
        selected_n_bins=20,
        main_effects=[],
        interactions=[],
        discretization_impact=pd.DataFrame({"n_bins": [20]}),
        summary=summary,
    )


def _write_workbook(payload: RatingTablePayload, target) -> None:
    write_rating_table_workbook(
        payload,
        target,
        sheet_name="Rating Tables",
        summary_sheet_name="Model Summary",
        impact_sheet_name="Discretization Impact",
    )


def _table_records(ws, table_name: str) -> list[dict[str, object]]:
    min_col, min_row, max_col, max_row = range_boundaries(ws.tables[table_name].ref)
    headers = [ws.cell(row=min_row, column=column).value for column in range(min_col, max_col + 1)]
    return [
        {
            header: ws.cell(row=row, column=column).value
            for header, column in zip(headers, range(min_col, max_col + 1), strict=True)
        }
        for row in range(min_row + 1, max_row + 1)
    ]


def _to_polars(frame: pd.DataFrame) -> pl.DataFrame:
    return pl.DataFrame({name: frame[name].to_numpy() for name in frame.columns})


def _assert_rating_payload_equal(
    actual: RatingTablePayload,
    expected: RatingTablePayload,
) -> None:
    assert actual.base_relativity == pytest.approx(expected.base_relativity)
    assert actual.selected_n_bins == expected.selected_n_bins
    assert [(block.name, block.kind) for block in actual.main_effects] == [
        (block.name, block.kind) for block in expected.main_effects
    ]
    for actual_block, expected_block in zip(
        actual.main_effects,
        expected.main_effects,
        strict=True,
    ):
        pd.testing.assert_frame_equal(actual_block.table, expected_block.table)
    assert [block.name for block in actual.interactions] == [
        block.name for block in expected.interactions
    ]
    for actual_block, expected_block in zip(
        actual.interactions,
        expected.interactions,
        strict=True,
    ):
        pd.testing.assert_frame_equal(actual_block.table, expected_block.table)
    pd.testing.assert_frame_equal(actual.discretization_impact, expected.discretization_impact)
    assert actual.summary == expected.summary


def _workbook_values(path) -> dict[str, list[tuple[object, ...]]]:
    workbook = load_workbook(path, data_only=True)
    return {
        sheet_name: [tuple(cell.value for cell in row) for row in workbook[sheet_name].iter_rows()]
        for sheet_name in workbook.sheetnames
    }


def _fit_ordered_export_model():
    rng = np.random.default_rng(20260712)
    levels = [f"L{i}" for i in range(7)]
    codes = np.tile(np.arange(len(levels)), 80)
    rng.shuffle(codes)
    X = pd.DataFrame({"band": np.asarray(levels, dtype=object)[codes]})
    x_numeric = codes / (len(levels) - 1)
    weights = rng.uniform(0.6, 1.8, len(codes))
    eta = -0.8 + 0.9 * x_numeric + 0.15 * np.sin(2.0 * np.pi * x_numeric)
    y = rng.poisson(np.exp(eta) * weights).astype(float)
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={
            "band": OrderedCategorical(
                order=levels,
                base="first",
                basis=Spline(kind="ps", k=7),
            )
        },
    )
    model.fit(X, y, sample_weight=weights)
    return model, levels


def _fit_selected_away_export_model():
    n = 160
    X = pd.DataFrame(
        {
            "x": np.linspace(-1.0, 1.0, n),
            "cat": np.resize(np.array(["A", "B", "C", "D"], dtype=object), n),
        }
    )
    y = np.ones(n)
    model = SuperGLM(
        family="poisson",
        selection_penalty=10.0,
        features={"x": Numeric(), "cat": Categorical(base="first")},
    )
    model.fit(X, y)
    return model


def _fit_polynomial_categorical_export_model():
    rng = np.random.default_rng(20260713)
    n = 240
    x = rng.uniform(-1.0, 1.0, n)
    cat = rng.choice(["A", "B", "C"], n)
    y = rng.poisson(np.exp(0.2 + 0.3 * x + 0.15 * (cat == "B"))).astype(float)
    X = pd.DataFrame({"x": x, "cat": cat})
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={"x": Polynomial(degree=2), "cat": Categorical(base="first")},
        interactions=[("x", "cat")],
    )
    model.fit(X, y)
    return model


def _fit_large_polynomial_interaction_export_model():
    rng = np.random.default_rng(20260714)
    n = 240
    a = rng.uniform(-1.0, 1.0, n)
    b = rng.uniform(-1.0, 1.0, n)
    y = rng.poisson(np.exp(0.2 + 0.25 * a - 0.15 * b + 0.1 * a * b)).astype(float)
    X = pd.DataFrame({"a": a, "b": b})
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={"a": Polynomial(degree=2), "b": Polynomial(degree=3)},
        interactions=[("a", "b")],
    )
    model.fit(X, y)
    return model


def test_summary_export_preserves_typed_overview_and_intercept_inference():
    model, _, _, _ = _fit_export_model()

    payload = build_summary_export_payload(model)
    overview = _overview_values(payload)
    intercept = next(row for row in payload.terms if row.term == "Intercept")

    assert overview[("Model", "Method")] == "MLE"
    assert isinstance(overview[("Fit", "Observations")], int)
    assert isinstance(overview[("Fit", "Effective DF")], float)
    assert isinstance(overview[("Fit", "Converged")], bool)
    assert isinstance(overview[("Fit", "Iterations")], int)
    assert isinstance(overview[("Information Criteria", "AIC")], float)
    assert intercept.kind == "coefficient"
    assert isinstance(intercept.estimate, float)
    assert isinstance(intercept.p_value, float)
    assert intercept.statistic_type == "z"
    assert intercept.active is True


def test_summary_export_maps_spline_to_one_global_wood_test():
    model, _, _, _ = _fit_export_model()

    payload = build_summary_export_payload(model)
    age_rows = [row for row in payload.terms if row.group == "age"]
    smooth = next(row for row in age_rows if row.kind == "smooth")

    assert len([row for row in age_rows if row.kind == "smooth"]) == 1
    assert smooth.estimate is None
    assert smooth.std_error is None
    assert smooth.ci_lower is None
    assert smooth.ci_upper is None
    assert isinstance(smooth.statistic, float)
    assert smooth.statistic_type == "chi2"
    assert isinstance(smooth.p_value, float)
    assert isinstance(smooth.edf, float)
    assert isinstance(smooth.smoothing_lambda, float)
    assert smooth.active is True
    assert "Wood (2013)" in "\n".join(payload.notes)


@pytest.mark.parametrize("structured_kind", ["random_effect", "factor_smooth"])
def test_summary_export_does_not_label_structured_metadata_as_wald_test(
    structured_kind: str,
):
    rng = np.random.default_rng(20260726)
    n_levels = 8
    repeats = 10
    codes = np.repeat(np.arange(n_levels), repeats)
    groups = np.array([f"g{code}" for code in codes], dtype=object)
    x = np.tile(np.linspace(0.0, 1.0, repeats), n_levels)
    X = pd.DataFrame({"x": x, "group": groups})
    y = np.sin(3.0 * x) + 0.2 * codes + rng.normal(scale=0.05, size=len(x))
    if structured_kind == "random_effect":
        model = SuperGLM(
            family="gaussian",
            features={
                "group": RandomEffect(lambda_policy=LambdaPolicy.fixed(1.0)),
            },
            selection_penalty=0.0,
            direct_solve="structured",
        ).fit_reml(X, y, runtime_validation="skip")
        term_name = "group"
    else:
        policies = {
            "wiggle": LambdaPolicy.fixed(1.0),
            "null_0": LambdaPolicy.fixed(1.0),
            "null_1": LambdaPolicy.fixed(1.0),
        }
        model = SuperGLM(
            family="gaussian",
            interactions=[
                FactorSmooth("x", group="group", k=5, lambda_policy=policies),
            ],
            selection_penalty=0.0,
            direct_solve="structured",
        ).fit_reml(X, y, runtime_validation="skip")
        term_name = "x:group:fs"

    payload = build_summary_export_payload(model)
    row = next(item for item in payload.terms if item.term == term_name)
    notes = "\n".join(payload.notes)

    assert row.kind == "group"
    assert row.estimate is None
    assert row.std_error is None
    assert row.statistic is None
    assert row.statistic_type == ""
    assert row.p_value is None
    assert "Group chi-square p-values are Wald approximations" not in notes


@pytest.mark.parametrize("basis", [None, "fs", "sz"])
def test_rating_table_export_rejects_structured_terms_before_impact(
    basis: str | None,
    monkeypatch: pytest.MonkeyPatch,
):
    rng = np.random.default_rng(20260726)
    n = 160
    x = rng.uniform(-1.0, 1.0, n)
    codes = np.arange(n) % 5
    X = pd.DataFrame(
        {
            "x": x,
            "group": np.array([f"g-{code}" for code in codes], dtype=object),
        }
    )
    y = 0.5 + 0.3 * x + rng.normal(scale=0.15, size=n)
    if basis is None:
        features = {
            "x": Spline(k=6, lambda_policy=LambdaPolicy.fixed(1.0)),
            "group": RandomEffect(lambda_policy=LambdaPolicy.fixed(1.0)),
        }
        interactions = None
        expected = "group"
    else:
        features = {"x": Spline(k=6, lambda_policy=LambdaPolicy.fixed(1.0))}
        interactions = [
            FactorSmooth(
                "x",
                group="group",
                basis=basis,
                k=5,
                lambda_policy=LambdaPolicy.fixed(1.0),
            )
        ]
        expected = f"x:group:{basis}"
    model = SuperGLM(
        family="gaussian",
        features=features,
        interactions=interactions,
        selection_penalty=0.0,
        direct_solve="gram",
    ).fit_reml(X, y, runtime_validation="skip")

    monkeypatch.setattr(
        model,
        "discretization_impact",
        lambda *_args, **_kwargs: pytest.fail("impact analysis ran before export preflight"),
    )
    with pytest.raises(NotImplementedError, match=expected):
        build_rating_table_payload(model, X, y)


def test_rating_table_export_reports_every_unsupported_structured_term():
    rng = np.random.default_rng(32)
    n = 180
    X = pd.DataFrame(
        {
            "x": rng.uniform(-1.0, 1.0, n),
            "group": np.array([f"g-{index % 5}" for index in range(n)], dtype=object),
            "broker": np.array([f"b-{index % 9}" for index in range(n)], dtype=object),
        }
    )
    y = 0.5 + 0.3 * X["x"].to_numpy() + rng.normal(scale=0.15, size=n)
    model = SuperGLM(
        family="gaussian",
        features={
            "x": Spline(k=6, lambda_policy=LambdaPolicy.fixed(1.0)),
            "broker": RandomEffect(lambda_policy=LambdaPolicy.fixed(1.0)),
        },
        interactions=[
            FactorSmooth(
                "x",
                group="group",
                k=5,
                lambda_policy=LambdaPolicy.fixed(1.0),
            )
        ],
        selection_penalty=0.0,
        direct_solve="gram",
    ).fit_reml(X, y, runtime_validation="skip")

    with pytest.raises(NotImplementedError) as exc_info:
        build_rating_table_payload(model, X, y)
    message = str(exc_info.value)
    assert "broker" in message
    assert "x:group:fs" in message


def test_summary_export_keeps_ordered_level_estimates_but_only_one_global_p_value():
    model, levels = _fit_ordered_export_model()

    payload = build_summary_export_payload(model)
    rows = [row for row in payload.terms if row.group == "band"]
    smooth_rows = [row for row in rows if row.kind == "smooth"]
    level_rows = [row for row in rows if row.kind == "level"]

    assert len(smooth_rows) == 1
    assert isinstance(smooth_rows[0].p_value, float)
    assert len(level_rows) == len(levels)
    assert [row.term for row in level_rows] == [f"band[{level}]" for level in levels]
    for row in level_rows:
        assert isinstance(row.estimate, float)
        assert isinstance(row.std_error, float)
        assert isinstance(row.ci_lower, float)
        assert isinstance(row.ci_upper, float)
        assert row.statistic is None
        assert row.statistic_type == ""
        assert row.p_value is None
        assert row.significance == ""


def test_summary_export_keeps_distribution_profile_values_typed():
    model, _, _, _ = _fit_export_model()

    def unexpected_tweedie_ci(*args, **kwargs):
        raise AssertionError("summary export must not evaluate a Tweedie profile CI")

    model._nb_profile_result = SimpleNamespace(
        theta_hat=np.float64(2.75),
        nll=10.0,
        ci=lambda alpha: (np.float64(2.0), np.float64(3.5)),
    )
    model._tweedie_profile_result = SimpleNamespace(
        p_hat=np.float64(1.55),
        phi_hat=np.float64(0.8),
        method="brent",
        phi_method="mle",
        density_exact=True,
        nll=11.0,
        _ci_cache={0.05: (np.float64(1.4), np.float64(1.7))},
        ci=unexpected_tweedie_ci,
        ci_details=unexpected_tweedie_ci,
    )
    model._summary_cache = None

    overview = _overview_values(build_summary_export_payload(model))

    assert overview[("Distribution Profile", "NB2 Theta")] == 2.75
    assert overview[("Distribution Profile", "NB2 Theta CI Lower")] == 2.0
    assert overview[("Distribution Profile", "NB2 Theta CI Upper")] == 3.5
    assert overview[("Distribution Profile", "NB2 Theta Method")] == "Profile (exact)"
    assert overview[("Distribution Profile", "Tweedie p")] == 1.55
    assert overview[("Distribution Profile", "Tweedie p CI Lower")] == 1.4
    assert overview[("Distribution Profile", "Tweedie p CI Upper")] == 1.7
    assert overview[("Distribution Profile", "Tweedie p CI Status")] == "available"
    assert overview[("Distribution Profile", "Tweedie phi")] == 0.8
    assert overview[("Distribution Profile", "Tweedie p Method")] == "Profile MLE (Brent)"


def test_summary_export_ignores_stale_pearson_profile_ci():
    model, _, _, _ = _fit_export_model()

    def unexpected_ci(*args, **kwargs):
        raise AssertionError("summary export must not evaluate a Tweedie profile CI")

    model._tweedie_profile_result = SimpleNamespace(
        p_hat=np.float64(1.55),
        phi_hat=np.float64(0.8),
        method="brent",
        phi_method="pearson",
        density_exact=True,
        nll=11.0,
        _ci_cache={0.05: (np.float64(1.4), np.float64(1.7))},
        ci=unexpected_ci,
        ci_details=unexpected_ci,
    )
    model._summary_cache = None

    overview = _overview_values(build_summary_export_payload(model))

    assert overview[("Distribution Profile", "Tweedie p CI Lower")] is None
    assert overview[("Distribution Profile", "Tweedie p CI Upper")] is None
    assert overview[("Distribution Profile", "Tweedie p CI Status")] == (
        "unavailable for Pearson plug-in"
    )
    assert overview[("Distribution Profile", "Tweedie p Method")] == (
        "Approximate profile (Brent; Pearson plug-in)"
    )


def test_summary_export_marks_uncached_mle_profile_ci_not_computed():
    model, _, _, _ = _fit_export_model()

    def unexpected_ci(*args, **kwargs):
        raise AssertionError("summary export must not evaluate a Tweedie profile CI")

    model._tweedie_profile_result = SimpleNamespace(
        p_hat=np.float64(1.55),
        phi_hat=np.float64(0.8),
        method="brent",
        phi_method="mle",
        density_exact=True,
        nll=11.0,
        _ci_cache={},
        ci=unexpected_ci,
        ci_details=unexpected_ci,
    )
    model._summary_cache = None

    overview = _overview_values(build_summary_export_payload(model))

    assert overview[("Distribution Profile", "Tweedie p CI Lower")] is None
    assert overview[("Distribution Profile", "Tweedie p CI Upper")] is None
    assert overview[("Distribution Profile", "Tweedie p CI Status")] == "not computed"


def test_summary_export_preserves_stale_and_fixed_offset_editor_caveats():
    model, _, _, _ = _fit_export_model()
    model._editor_inference_stale = True
    model._editor_edits = {"terms": ["score"]}
    model._editor_offset = {"terms": ["region"]}
    model._summary_cache = None

    notes = "\n".join(build_summary_export_payload(model).notes)

    assert "suppressed" in notes
    assert "Edited terms: score" in notes
    assert "conditional on those fixed offsets" in notes
    assert "Offset terms: region" in notes


def test_summary_export_marks_selected_away_parametric_and_level_rows_inactive():
    model = _fit_selected_away_export_model()
    assert tuple(model.result.rank_info.selected_group_names) == ()

    payload = build_summary_export_payload(model)

    intercept = next(row for row in payload.terms if row.term == "Intercept")
    selected_away = [row for row in payload.terms if row.term != "Intercept"]

    assert intercept.active is True
    assert {row.kind for row in selected_away} == {"coefficient", "level"}
    assert all(row.estimate == 0.0 for row in selected_away)
    assert all(row.active is False for row in selected_away)


def test_summary_export_does_not_make_nonfinite_selected_away_row_active():
    model = _fit_selected_away_export_model()
    assert tuple(model.result.rank_info.selected_group_names) == ()
    # Deliberately inject an invalid private candidate to exercise the exporter's
    # defensive handling.  Published results correctly reject in-place mutation.
    private_result = model.result._mutable_copy()
    private_result.beta[0] = np.nan
    model._result = private_result
    model._solver_result = private_result
    model._summary_cache = None

    row = next(row for row in build_summary_export_payload(model).terms if row.term == "x")

    assert row.estimate is None
    assert row.active is False


def test_summary_export_labels_polynomial_categorical_tests_as_group_wald():
    payload = build_summary_export_payload(_fit_polynomial_categorical_export_model())

    rows = [row for row in payload.terms if row.group == "x:cat"]
    notes = "\n".join(payload.notes)

    assert len(rows) == 2
    assert all(row.kind == "group" for row in rows)
    assert all(row.estimate is None and row.std_error is None for row in rows)
    assert all(row.statistic_type == "chi2" for row in rows)
    assert all(isinstance(row.statistic, float) for row in rows)
    assert all(isinstance(row.p_value, float) for row in rows)
    assert "Group chi-square p-values are Wald approximations" in notes
    assert "Wood (2013)" not in notes


def test_summary_export_labels_large_polynomial_interaction_as_group_wald():
    payload = build_summary_export_payload(_fit_large_polynomial_interaction_export_model())

    row = next(row for row in payload.terms if row.term == "a:b")
    notes = "\n".join(payload.notes)

    assert row.kind == "group"
    assert row.estimate is None
    assert row.std_error is None
    assert row.statistic_type == "chi2"
    assert isinstance(row.statistic, float)
    assert isinstance(row.p_value, float)
    assert "Group chi-square p-values are Wald approximations" in notes
    assert "Wood (2013)" not in notes


def test_public_export_api_exists(tmp_path):
    model, X, y, w = _fit_export_model()
    output = tmp_path / "rating_tables.xlsx"

    payload = model.rating_table_payload(X, y, sample_weight=w)
    result_path = export_rating_tables(model, output, X, y, sample_weight=w)
    method_path = model.export_rating_tables(
        tmp_path / "rating_tables_method.xlsx", X, y, sample_weight=w
    )

    assert payload.base_relativity > 0.0
    assert result_path == output
    assert output.exists()
    assert method_path.exists()


def test_polars_rating_payload_and_workbook_match_pandas(tmp_path, monkeypatch):
    _, X, y, w = _fit_export_model()
    X = X.assign(segment=np.where(X["score"].to_numpy() >= 0.0, "high", "low"))
    interaction_model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={
            "age": Spline(n_knots=8),
            "region": Categorical(base="first"),
            "score": Numeric(),
            "segment": Categorical(base="first"),
        },
        interactions=[("region", "segment")],
    ).fit(X, y, sample_weight=w)
    X_polars = _to_polars(X)
    expected = build_rating_table_payload(
        interaction_model,
        X,
        y,
        sample_weight=w,
        n_bins=24,
        impact_bins=(20,),
    )

    def unexpected_whole_frame_conversion(*_args, **_kwargs):
        pytest.fail("rating-table export converted the whole Polars frame")

    monkeypatch.setattr(pl.DataFrame, "to_pandas", unexpected_whole_frame_conversion)
    actual = interaction_model.rating_table_payload(
        X_polars,
        y,
        sample_weight=w,
        n_bins=24,
        impact_bins=(20,),
    )

    _assert_rating_payload_equal(actual, expected)
    assert actual.interactions
    pandas_path = tmp_path / "pandas-rating.xlsx"
    polars_path = tmp_path / "polars-rating.xlsx"
    export_rating_tables(
        interaction_model,
        pandas_path,
        X,
        y,
        sample_weight=w,
        n_bins=24,
        impact_bins=(20,),
    )
    interaction_model.export_rating_tables(
        polars_path,
        X_polars,
        y,
        sample_weight=w,
        n_bins=24,
        impact_bins=(20,),
    )
    assert _workbook_values(polars_path) == _workbook_values(pandas_path)


def test_polars_rating_payload_preserves_fitted_offset_and_source_column():
    model, X, y, w, term, offset = _fit_term_offset_export_model()
    X_with_source = X.assign(term_months=term)
    X_polars = _to_polars(X_with_source)
    expected = build_rating_table_payload(
        model,
        X_with_source,
        y,
        sample_weight=w,
        offset=offset,
        offset_source="term_months",
        impact_bins=(),
    )

    actual = build_rating_table_payload(
        model,
        X_polars,
        y,
        sample_weight=w,
        offset=offset,
        offset_source="term_months",
        impact_bins=(),
    )

    _assert_rating_payload_equal(actual, expected)
    offset_block = next(block for block in actual.main_effects if block.name == "term_months")
    assert offset_block.table.columns.tolist() == ["term_months", "Relativity", "Weight"]


def test_polars_fitted_frame_resolves_retained_offset_by_identity():
    pandas_model, X, y, w, _term, offset = _fit_term_offset_export_model()
    X_polars = _to_polars(X)
    polars_model = pandas_model.clone_unfitted().fit(
        X_polars,
        y,
        sample_weight=w,
        offset=offset,
    )

    expected = build_rating_table_payload(
        pandas_model,
        X,
        y,
        sample_weight=w,
        impact_bins=(),
    )
    actual = build_rating_table_payload(
        polars_model,
        X_polars,
        y,
        sample_weight=w,
        impact_bins=(),
    )

    assert polars_model._fit_X_ref is X_polars
    _assert_rating_payload_equal(actual, expected)


def test_default_selected_bins_are_150():
    model, X, y, w = _fit_export_model()
    payload = build_rating_table_payload(model, X, y, sample_weight=w)

    age_block = next(block for block in payload.main_effects if block.name == "age")

    assert payload.selected_n_bins == 150
    assert isinstance(payload.summary, SummaryExportPayload)
    assert not hasattr(payload, "summary_lines")
    assert len(age_block.table) <= 150
    assert {"age", "Relativity", "Weight"} <= set(age_block.table.columns)


def test_default_impact_sweep_bins():
    """The default sweep includes the resolution the workbook was exported at.

    ``impact_bins`` is a comparison ladder and ``n_bins`` is what shipped; with
    the default 150 outside the ladder, every row of the sheet described a
    table the reader does not have, and the 200/250 rows understated their
    error because it falls with resolution.
    """
    model, X, y, w = _fit_export_model()
    payload = build_rating_table_payload(model, X, y, sample_weight=w)
    sheet = payload.discretization_impact

    assert sorted(sheet["n_bins"].unique().tolist()) == [20, 50, 100, 150, 200, 250]
    assert sorted(sheet[sheet["exported"]]["n_bins"].unique().tolist()) == [150]
    assert set(sheet["feature"]) == {"age"}


def test_categorical_and_numeric_blocks_are_exported():
    model, X, y, w = _fit_export_model()
    payload = build_rating_table_payload(model, X, y, sample_weight=w)

    names = [block.name for block in payload.main_effects]
    assert names == ["age", "region", "score"]

    region = next(block for block in payload.main_effects if block.name == "region")
    score = next(block for block in payload.main_effects if block.name == "score")

    assert set(region.table["region"]) == {"A", "B", "C"}
    assert np.isclose(region.table["Weight"].sum(), w.sum())
    assert score.table["score"].tolist() == ["per_unit"]


def test_integer_categorical_block_weights_do_not_warn_on_pandas_integer_keys():
    rng = np.random.default_rng(124)
    n = 120
    X = pd.DataFrame({"region": rng.choice([1, 2, 3], n)})
    eta = -1.0 + 0.2 * (X["region"].to_numpy() == 2)
    y = rng.poisson(np.exp(eta)).astype(float)
    w = rng.uniform(0.5, 2.0, n)
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={"region": Categorical(base="first")},
    )
    model.fit(X, y, sample_weight=w)

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "error",
            message="Series.__getitem__ treating keys as positions is deprecated",
            category=FutureWarning,
        )
        payload = build_rating_table_payload(model, X, y, sample_weight=w)

    region = next(block for block in payload.main_effects if block.name == "region")
    assert np.isclose(region.table["Weight"].sum(), w.sum())


def test_fit_offset_exports_exact_multiplier_block_when_support_is_small():
    model, X, y, w = _fit_offset_export_model(distinct_terms=2)

    payload = build_rating_table_payload(model, X, y, sample_weight=w)

    offset_block = next(
        block for block in payload.main_effects if block.name == "Offset Multiplier"
    )
    assert offset_block.kind == "offset"
    assert offset_block.table.columns.tolist() == ["Offset Multiplier", "Relativity", "Weight"]
    assert sorted(offset_block.table["Offset Multiplier"].tolist()) == [1.0, 3.0]
    np.testing.assert_allclose(
        offset_block.table.sort_values("Offset Multiplier")["Relativity"].to_numpy(),
        [1.0, 3.0],
    )
    assert np.isclose(offset_block.table["Weight"].sum(), w.sum())


def test_fit_offset_exports_binned_multiplier_block_when_support_is_large():
    model, X, y, w = _fit_offset_export_model(distinct_terms=40)

    payload = build_rating_table_payload(model, X, y, sample_weight=w, n_bins=5)

    offset_block = next(
        block for block in payload.main_effects if block.name == "Offset Multiplier"
    )
    assert offset_block.kind == "offset"
    assert len(offset_block.table) == 5
    assert offset_block.table["Offset Multiplier"].str.startswith("[").all()
    assert np.isclose(offset_block.table["Weight"].sum(), w.sum())
    assert offset_block.table["Relativity"].between(1.0, 4.0).all()


def test_fit_offset_source_exports_raw_term_lookup():
    model, X, y, w, term, offset = _fit_term_offset_export_model()

    payload = build_rating_table_payload(
        model,
        X,
        y,
        sample_weight=w,
        offset=offset,
        offset_source=term,
        offset_name="Term",
    )

    offset_block = next(block for block in payload.main_effects if block.name == "Term")
    assert offset_block.kind == "offset"
    assert offset_block.table.columns.tolist() == ["Term", "Relativity", "Weight"]
    table = offset_block.table.sort_values("Term")
    assert table["Term"].tolist() == [12.0, 36.0]
    np.testing.assert_allclose(table["Relativity"].to_numpy(), [1.0, 3.0])
    assert np.isclose(table["Weight"].sum(), w.sum())


def test_fit_offset_source_resolves_string_and_series_names():
    model, X, y, w, term, offset = _fit_term_offset_export_model()
    X_with_term = X.assign(term_months=term)

    from_column = build_rating_table_payload(
        model,
        X_with_term,
        y,
        sample_weight=w,
        offset=offset,
        offset_source="term_months",
    )
    from_series = build_rating_table_payload(
        model,
        X,
        y,
        sample_weight=w,
        offset=offset,
        offset_source=pd.Series(term, name="Policy Term"),
    )

    assert any(block.name == "term_months" for block in from_column.main_effects)
    assert any(block.name == "Policy Term" for block in from_series.main_effects)


def test_fit_offset_source_rejects_reserved_and_blank_names():
    model, X, y, w, term, offset = _fit_term_offset_export_model()

    for bad_name in ["", "   ", "Relativity", "Weight"]:
        with pytest.raises(ValueError, match="offset_name"):
            build_rating_table_payload(
                model,
                X,
                y,
                sample_weight=w,
                offset=offset,
                offset_source=term,
                offset_name=bad_name,
            )


def test_fit_offset_source_requires_name_for_unnamed_array():
    model, X, y, w, term, offset = _fit_term_offset_export_model()

    with pytest.raises(ValueError, match="offset_name is required"):
        build_rating_table_payload(
            model,
            X,
            y,
            sample_weight=w,
            offset=offset,
            offset_source=term,
        )


def test_fit_offset_source_rejects_inconsistent_mapping():
    model, X, y, w, term, offset = _fit_term_offset_export_model()
    bad_offset = offset.copy()
    bad_offset[term == 12.0] = 0.0
    bad_offset[np.flatnonzero(term == 12.0)[0]] = np.log(1.2)

    with pytest.raises(ValueError, match="maps to multiple offset multipliers"):
        build_rating_table_payload(
            model,
            X,
            y,
            sample_weight=w,
            offset=bad_offset,
            offset_source=term,
            offset_name="Term",
        )


def test_fit_offset_source_rejects_high_cardinality_source():
    model, X, y, w, term, offset = _fit_term_offset_export_model(distinct_terms=40)

    with pytest.raises(ValueError, match="exceeding offset_max_exact_levels=20"):
        build_rating_table_payload(
            model,
            X,
            y,
            sample_weight=w,
            offset=offset,
            offset_source=term,
            offset_name="Term",
        )


def test_fit_offset_source_rejects_missing_source_values():
    model, X, y, w, term, offset = _fit_term_offset_export_model()
    source = term.copy()
    source[0] = np.nan

    with pytest.raises(ValueError, match="offset_source cannot contain missing values"):
        build_rating_table_payload(
            model,
            X,
            y,
            sample_weight=w,
            offset=offset,
            offset_source=source,
            offset_name="Term",
        )


def test_fit_offset_source_allows_non_bijective_mapping():
    model, X, y, w, _term, offset = _fit_term_offset_export_model()
    source = np.resize(np.array([-1.0, 1.0]), len(X))
    shared_offset = np.zeros(len(X), dtype=np.float64)

    payload = build_rating_table_payload(
        model,
        X,
        y,
        sample_weight=w,
        offset=shared_offset,
        offset_source=source,
        offset_name="Signed Source",
    )

    offset_block = next(block for block in payload.main_effects if block.name == "Signed Source")
    table = offset_block.table.sort_values("Signed Source")
    assert table["Signed Source"].tolist() == [-1.0, 1.0]
    np.testing.assert_allclose(table["Relativity"].to_numpy(), [1.0, 1.0])


def test_fit_offset_source_ignores_unused_categorical_levels():
    model, X, y, w, term, offset = _fit_term_offset_export_model()
    source = pd.Series(
        pd.Categorical(term, categories=[12.0, 36.0, 60.0]),
        name="Term",
    )

    payload = build_rating_table_payload(
        model,
        X,
        y,
        sample_weight=w,
        offset=offset,
        offset_source=source,
    )

    offset_block = next(block for block in payload.main_effects if block.name == "Term")
    table = offset_block.table.sort_values("Term")
    assert table["Term"].tolist() == [12.0, 36.0]
    np.testing.assert_allclose(table["Relativity"].to_numpy(), [1.0, 3.0])


def test_fit_offset_source_reordered_frame_requires_explicit_offset():
    model, X, y, w, term, _offset = _fit_term_offset_export_model()
    order = np.arange(len(X))[::-1]

    with pytest.raises(ValueError, match="Pass offset="):
        build_rating_table_payload(
            model,
            X.iloc[order].reset_index(drop=True),
            y[order],
            sample_weight=w[order],
            offset_source=term[order],
            offset_name="Term",
        )


def test_fit_offset_source_reordered_frame_uses_aligned_offset():
    model, X, y, w, term, offset = _fit_term_offset_export_model()
    order = np.arange(len(X))[::-1]

    payload = build_rating_table_payload(
        model,
        X.iloc[order].reset_index(drop=True),
        y[order],
        sample_weight=w[order],
        offset=offset[order],
        offset_source=term[order],
        offset_name="Term",
    )

    offset_block = next(block for block in payload.main_effects if block.name == "Term")
    table = offset_block.table.sort_values("Term")
    assert table["Term"].tolist() == [12.0, 36.0]
    np.testing.assert_allclose(table["Relativity"].to_numpy(), [1.0, 3.0])


def test_fit_offset_export_rejects_offset_for_model_fit_without_offset():
    model, X, y, w = _fit_export_model()

    with pytest.raises(ValueError, match="requires a model fitted with an offset"):
        build_rating_table_payload(
            model,
            X,
            y,
            sample_weight=w,
            offset=np.zeros(len(X)),
            offset_source=np.ones(len(X)),
            offset_name="Term",
        )


def test_fit_offset_export_with_released_fit_state_requires_explicit_offset():
    model, X, y, w, term, _offset = _fit_term_offset_export_model(retain_fit_state=False)

    with pytest.raises(ValueError, match="Pass offset="):
        build_rating_table_payload(
            model,
            X,
            y,
            sample_weight=w,
            offset_source=term,
            offset_name="Term",
        )


def test_fit_offset_export_rejects_non_log_link_model():
    rng = np.random.default_rng(989)
    n = 80
    X = pd.DataFrame({"score": rng.normal(size=n)})
    offset = np.resize(np.array([0.0, 1.0]), n)
    y = 1.0 + 0.5 * X["score"].to_numpy() + offset
    model = SuperGLM(
        family="gaussian",
        link="identity",
        selection_penalty=0.0,
        features={"score": Numeric()},
    )
    model.fit(X, y, offset=offset)

    with pytest.raises(ValueError, match="log-link models"):
        build_rating_table_payload(
            model,
            X,
            y,
            offset=offset,
            offset_source=np.exp(offset),
            offset_name="Offset Source",
        )


def test_rating_table_payload_passes_offset_to_discretization_impact(monkeypatch):
    model, X, y, w = _fit_export_model()
    offset = np.full(len(X), np.log(2.0), dtype=np.float64)
    model.fit(X, y, sample_weight=w, offset=offset)
    seen_offsets: list[np.ndarray | None] = []

    def fake_discretization_impact(_X, _y, sample_weight=None, **kwargs):
        seen_offsets.append(kwargs.get("offset"))
        table = pd.DataFrame(
            {
                "bin_from": [18.0],
                "bin_to": [80.0],
                "relativity": [1.0],
                "log_relativity": [0.0],
                "n_obs": [len(_X)],
                "sample_weight": [float(np.sum(sample_weight))],
            }
        )
        return type(
            "FakeDiscretizationResult",
            (),
            {
                "tables": {"age": table},
                # The double stands in for a whole ``DiscretizationResult``, so
                # it carries every mapping the sweep reads; a defaulting
                # ``getattr`` on the reader's side would have hidden a rename
                # here rather than surfacing it.
                "interaction_tables": {},
                "metrics": {
                    "deviance_original": 1.0,
                    "deviance_discretized": 1.0,
                    "deviance_change": 0.0,
                    "deviance_change_pct": 0.0,
                    "mean_abs_prediction_change_pct": 0.0,
                    "max_abs_prediction_change_pct": 0.0,
                    "prediction_correlation": 1.0,
                },
            },
        )()

    monkeypatch.setattr(model, "discretization_impact", fake_discretization_impact)

    build_rating_table_payload(model, X, y, sample_weight=w, offset=offset, n_bins=1)

    # Six ladder rungs plus the exported ``n_bins=1``, which the sweep folds in
    # so the sheet has a row about the table that shipped.
    assert len(seen_offsets) == 7
    for seen in seen_offsets:
        np.testing.assert_allclose(seen, offset)


def test_main_effect_blocks_keep_a_three_column_stride_and_their_number_formats(tmp_path):
    """Characterization test for the main-effects layout, written before it moved.

    ``write_rating_table_workbook`` places main-effect blocks at
    ``start_col = 1 + idx * 3`` and then applies number formats *globally* by
    ``cell.column % 3``.  Both are load-bearing and neither is announced by a
    named constant, so a block wider than three columns would silently overwrite
    its right-hand neighbour and take that neighbour's formats.

    This pins the layout as it stands so the piecewise block added afterwards can
    be shown not to have disturbed it: it must pass identically before and after
    that change.
    """
    stride = 3
    title_row = 5
    header_row = 7

    model, X, y, w = _fit_export_model()
    payload = build_rating_table_payload(model, X, y, sample_weight=w, n_bins=20)
    output = tmp_path / "characterization.xlsx"
    _write_workbook(payload, output)
    ws = load_workbook(output, data_only=True)["Rating Tables"]

    assert [(block.name, block.kind) for block in payload.main_effects] == [
        ("age", "continuous"),
        ("region", "categorical"),
        ("score", "numeric"),
    ]

    for idx, block in enumerate(payload.main_effects):
        start_col = 1 + idx * stride
        assert list(block.table.columns) == [block.name, "Relativity", "Weight"]

        title = ws.cell(row=title_row, column=start_col)
        assert title.value == block.name
        assert title.font.bold

        headers = [
            ws.cell(row=header_row, column=start_col + offset).value
            for offset in range(len(block.table.columns))
        ]
        assert headers == list(block.table.columns)

        # Every data row, not just the first: the format loop walks cells, so a
        # partial reformat is a real failure mode.
        for row in range(header_row + 1, header_row + 1 + len(block.table)):
            assert ws.cell(row=row, column=start_col + 1).number_format == "0.000000"
            assert ws.cell(row=row, column=start_col + 2).number_format == "#,##0.00"

        # Nothing is written in the gap the piecewise note later occupies.
        assert ws.cell(row=title_row + 1, column=start_col).value is None

    # The stride itself, stated as a relation between neighbouring blocks rather
    # than as three hard-coded column letters.
    title_columns = [
        cell.column
        for cell in ws[title_row]
        if cell.value in {block.name for block in payload.main_effects}
    ]
    assert title_columns == [1 + idx * stride for idx in range(len(payload.main_effects))]


def test_excel_workbook_layout(tmp_path):
    model, X, y, w = _fit_export_model()
    output = tmp_path / "tables.xlsx"
    expected_summary = build_summary_export_payload(model)

    model.export_rating_tables(output, X, y, sample_weight=w, n_bins=20)

    wb = load_workbook(output, data_only=True)
    assert wb.sheetnames == ["Rating Tables", "Discretization Impact", "Model Summary"]
    ws = wb["Rating Tables"]
    assert ws["A2"].value == "Base"
    assert isinstance(ws["C2"].value, float)
    assert ws["A5"].value == "age"
    assert ws["A7"].value == "age"
    assert ws["B7"].value == "Relativity"
    assert ws["C7"].value == "Weight"
    assert ws["D5"].value == "region"
    assert ws["D7"].value == "region"
    assert ws["G5"].value == "score"
    assert ws["G7"].value == "score"

    impact_ws = wb["Discretization Impact"]
    headers = [impact_ws.cell(row=1, column=i).value for i in range(1, 12)]
    assert headers[:4] == ["n_bins", "exported", "feature", "actual_bins"]

    summary_ws = wb["Model Summary"]
    assert summary_ws["A1"].value == "Model Summary"
    assert summary_ws["A1"].font.bold
    assert summary_ws["A1"].font.sz == pytest.approx(14.0)
    assert summary_ws["A3"].value == "Fit and model overview"
    assert [summary_ws.cell(row=4, column=column).value for column in range(1, 4)] == [
        "Section",
        "Metric",
        "Value",
    ]
    assert set(summary_ws.tables) == {"ModelOverview", "TermInference"}
    assert summary_ws.freeze_panes == "A5"

    overview_table = summary_ws.tables["ModelOverview"]
    overview_bounds = range_boundaries(overview_table.ref)
    term_table = summary_ws.tables["TermInference"]
    term_min_col, term_min_row, term_max_col, term_max_row = range_boundaries(term_table.ref)
    assert term_min_row == overview_bounds[3] + 3
    term_headers = [
        summary_ws.cell(row=term_min_row, column=column).value
        for column in range(term_min_col, term_max_col + 1)
    ]
    assert term_headers == EXPECTED_SUMMARY_TERM_HEADERS

    overview = {row["Metric"]: row["Value"] for row in _table_records(summary_ws, "ModelOverview")}
    assert isinstance(overview["Observations"], int)
    assert not isinstance(overview["Observations"], bool)
    assert isinstance(overview["AIC"], float)
    assert isinstance(overview["Converged"], bool)

    terms = _table_records(summary_ws, "TermInference")
    assert [(row["Term"], row["Kind"]) for row in terms] == [
        (row.term, row.kind) for row in expected_summary.terms
    ]
    assert {row["Kind"] for row in terms} == {"coefficient", "level", "smooth"}
    assert summary_ws.cell(row=term_max_row + 3, column=1).value == "Notes"
    assert summary_ws.cell(row=term_max_row + 3, column=1).font.bold
    assert [
        summary_ws.cell(row=term_max_row + 4 + index, column=1).value
        for index in range(len(expected_summary.notes))
    ] == list(expected_summary.notes)
    assert all(
        "SuperGLM Results" not in str(cell.value) for row in summary_ws.iter_rows() for cell in row
    )
    assert summary_ws.column_dimensions["A"].width < 60


def test_summary_level_display_modes_do_not_change_collapsed_excel_export(tmp_path):
    model, X, y, w = _fit_export_model()
    session = EditorSession.from_model(
        model,
        terms=["region"],
        train_data=(X, y, w),
    )
    session.select_levels("region", ["B", "C"])
    collapsed = session.replace_with_collapsed_levels("region", method="fit")
    before_summary = build_summary_export_payload(collapsed)
    before_rating = build_rating_table_payload(collapsed, X, y, sample_weight=w)

    collapsed.summary(level_display="expanded")
    collapsed.summary(level_display="grouped")

    assert build_summary_export_payload(collapsed) == before_summary
    after_rating = build_rating_table_payload(collapsed, X, y, sample_weight=w)
    _assert_rating_payload_equal(after_rating, before_rating)

    region = next(block.table for block in before_rating.main_effects if block.name == "region")
    assert region["region"].tolist() == ["A", "B", "C"]
    expected_weights = X.assign(_weight=w).groupby("region", sort=False)["_weight"].sum()
    np.testing.assert_allclose(
        region["Weight"].to_numpy(),
        expected_weights.reindex(region["region"]).to_numpy(),
    )
    assert region.loc[region["region"] == "B", "Relativity"].iloc[0] == pytest.approx(
        region.loc[region["region"] == "C", "Relativity"].iloc[0]
    )

    output = tmp_path / "collapsed_summary_modes.xlsx"
    collapsed.export_rating_tables(output, X, y, sample_weight=w)
    workbook = load_workbook(output, data_only=True)
    summary_ws = workbook["Model Summary"]
    term_min_col, term_min_row, term_max_col, _ = range_boundaries(
        summary_ws.tables["TermInference"].ref
    )
    term_headers = [
        summary_ws.cell(row=term_min_row, column=column).value
        for column in range(term_min_col, term_max_col + 1)
    ]
    assert term_headers == EXPECTED_SUMMARY_TERM_HEADERS
    assert "Level group" not in term_headers
    assert "G1" not in {cell.value for row in summary_ws.iter_rows() for cell in row}

    rating_ws = workbook["Rating Tables"]
    assert [rating_ws.cell(row=7, column=column).value for column in range(4, 7)] == [
        "region",
        "Relativity",
        "Weight",
    ]
    workbook_region = pd.DataFrame(
        {
            "region": [rating_ws.cell(row=row, column=4).value for row in range(8, 11)],
            "Relativity": [rating_ws.cell(row=row, column=5).value for row in range(8, 11)],
            "Weight": [rating_ws.cell(row=row, column=6).value for row in range(8, 11)],
        }
    )
    pd.testing.assert_frame_equal(workbook_region, region.reset_index(drop=True))


def test_ordered_spline_workbook_keeps_only_global_inference(tmp_path):
    model, levels = _fit_ordered_export_model()
    summary = build_summary_export_payload(model)
    output = tmp_path / "ordered_summary.xlsx"

    _write_workbook(_workbook_payload(summary), output)

    summary_ws = load_workbook(output, data_only=True)["Model Summary"]
    rows = [row for row in _table_records(summary_ws, "TermInference") if row["Group"] == "band"]
    smooth_rows = [row for row in rows if row["Kind"] == "smooth"]
    level_rows = [row for row in rows if row["Kind"] == "level"]

    assert len(smooth_rows) == 1
    assert isinstance(smooth_rows[0]["P Value"], float)
    assert len(level_rows) == len(levels)
    for row in level_rows:
        assert row["Estimate"] is not None
        assert row["Std Error"] is not None
        assert row["CI Lower"] is not None
        assert row["CI Upper"] is not None
        assert row["Statistic"] is None
        assert row["P Value"] is None
        assert row["Significance"] in (None, "")


def _fit_ordered_specials_export_model():
    rng = np.random.default_rng(20260805)
    ordered = [f"L{i}" for i in range(7)]
    codes = np.repeat(np.arange(len(ordered)), 90)
    band_ordered = np.asarray(ordered, dtype=object)[codes]
    x = codes / (len(ordered) - 1.0)
    eta_ordered = -0.8 + 0.9 * x
    band_missing = np.full(240, "MISSING", dtype=object)
    eta_missing = np.full(240, -0.8 - 0.5)
    band = np.concatenate([band_ordered, band_missing])
    eta = np.concatenate([eta_ordered, eta_missing])
    region = np.resize(np.array(["N", "S", "E"], dtype=object), band.size)
    weights = rng.uniform(0.6, 1.8, band.size)
    y = rng.poisson(np.exp(eta) * weights).astype(float)
    X = pd.DataFrame({"band": band, "region": region})
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={
            "band": OrderedCategorical(
                order=ordered,
                specials=["MISSING"],
                base="first",
                basis=Spline(kind="ps", k=5),
            ),
            "region": Categorical(base="N"),
        },
    )
    model.fit(X, y, sample_weight=weights)
    return model, X, y, weights, ordered


def test_summary_sheet_marks_a_term_that_contains_free_levels():
    # False today: a specials term's group row is Kind="smooth", identical to
    # a term with no specials, so the workbook records nothing about free
    # levels. Also pins that the Wood note survives the new kind value.
    model, _, _, _, _ = _fit_ordered_specials_export_model()
    payload = build_summary_export_payload(model)
    band_rows = [row for row in payload.terms if row.group == "band"]

    marked = [row for row in band_rows if row.kind == "smooth+free"]
    assert len(marked) == 1
    assert marked[0].term == "band"
    assert isinstance(marked[0].p_value, float)
    assert not [row for row in band_rows if row.kind == "smooth"]
    assert "Smooth p-values use Wood (2013) Bayesian tests." in payload.notes


def test_summary_sheet_marks_the_free_level_row_itself():
    # False today: every level row is Kind="level", so the workbook cannot say
    # WHICH level was fitted free. export/summary.py:301 already emits one row
    # per level (test_rating_table_export.py:1325-1348 pins that), so this is
    # the kind string alone — no new column and no rating-sheet change.
    model, _, _, _, ordered = _fit_ordered_specials_export_model()
    payload = build_summary_export_payload(model)
    level_rows = [
        row for row in payload.terms if row.group == "band" and row.kind in {"level", "free level"}
    ]

    assert [row.term for row in level_rows] == [f"band[{level}]" for level in [*ordered, "MISSING"]]
    assert [row.kind for row in level_rows] == ["level"] * len(ordered) + ["free level"]
    free_row = level_rows[-1]
    assert free_row.estimate is not None
    assert free_row.std_error is not None
    assert free_row.p_value is None


def test_summary_sheet_level_kinds_are_unchanged_without_specials():
    # Guards the width/format contract the other direction: a term with no
    # specials must keep every level row at Kind="level".
    model, levels = _fit_ordered_export_model()
    payload = build_summary_export_payload(model)
    level_rows = [
        row for row in payload.terms if row.group == "band" and row.kind in {"level", "free level"}
    ]

    assert [row.kind for row in level_rows] == ["level"] * len(levels)


def test_specials_workbook_keeps_summary_columns_and_rating_block_layout(tmp_path):
    # False today: nothing exercises a specials model through the workbook, so
    # neither the fixed Summary header set nor the 3-column rating blocks are
    # pinned against a marker column creeping onto the rating sheet.
    model, X, y, weights, _ = _fit_ordered_specials_export_model()
    output = tmp_path / "specials.xlsx"

    model.export_rating_tables(output, X, y, sample_weight=weights, n_bins=20)

    wb = load_workbook(output, data_only=True)
    summary_ws = wb["Model Summary"]
    term_min_col, term_min_row, term_max_col, _ = range_boundaries(
        summary_ws.tables["TermInference"].ref
    )
    term_headers = [
        summary_ws.cell(row=term_min_row, column=column).value
        for column in range(term_min_col, term_max_col + 1)
    ]
    assert term_headers == EXPECTED_SUMMARY_TERM_HEADERS
    kinds = {row["Kind"] for row in _table_records(summary_ws, "TermInference")}
    assert "smooth+free" in kinds
    assert "free level" in kinds

    rating_ws = wb["Rating Tables"]
    assert rating_ws["A5"].value == "band"
    assert [rating_ws.cell(row=7, column=column).value for column in range(1, 4)] == [
        "band",
        "Relativity",
        "Weight",
    ]
    # Block 2 must still start at column 4: excel.py:176 keys start_col and
    # excel.py:186/188 key number formats on a 3-column stride.
    assert rating_ws["D5"].value == "region"
    assert [rating_ws.cell(row=7, column=column).value for column in range(4, 7)] == [
        "region",
        "Relativity",
        "Weight",
    ]


def test_excel_workbook_writes_to_binary_stream():
    summary = SummaryExportPayload(
        overview=(SummaryOverviewRow("Fit", "Observations", 12),),
        terms=(
            SummaryTermRow(
                term="Intercept",
                group="",
                kind="coefficient",
                estimate=0.25,
                std_error=0.1,
                statistic=2.5,
                statistic_type="z",
                p_value=0.012,
                ci_lower=0.05,
                ci_upper=0.45,
                edf=None,
                smoothing_lambda=None,
                active=True,
                significance="*",
                warning="",
            ),
        ),
        notes=("Typed workbook stream.",),
    )
    target = BytesIO()

    _write_workbook(_workbook_payload(summary), target)

    assert not target.closed
    target.seek(0)
    wb = load_workbook(target, data_only=True)
    assert wb.sheetnames == ["Rating Tables", "Discretization Impact", "Model Summary"]
    assert set(wb["Model Summary"].tables) == {"ModelOverview", "TermInference"}


def test_excel_workbook_empty_terms_has_valid_blank_table_row():
    summary = SummaryExportPayload(
        overview=(SummaryOverviewRow("Fit", "Observations", 0),),
        terms=(),
        notes=(),
    )
    target = BytesIO()

    _write_workbook(_workbook_payload(summary), target)

    target.seek(0)
    summary_ws = load_workbook(target, data_only=True)["Model Summary"]
    table = summary_ws.tables["TermInference"]
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    assert max_row == min_row + 1
    assert [
        summary_ws.cell(row=max_row, column=column).value for column in range(min_col, max_col + 1)
    ] == [None] * 15


def test_excel_workbook_includes_fit_offset_multiplier(tmp_path):
    model, X, y, w = _fit_offset_export_model(distinct_terms=2)
    output = tmp_path / "tables_with_offset.xlsx"

    model.export_rating_tables(output, X, y, sample_weight=w)

    ws = load_workbook(output, data_only=True)["Rating Tables"]
    assert ws["D5"].value == "Offset Multiplier"
    assert ws["D7"].value == "Offset Multiplier"
    assert ws["E7"].value == "Relativity"
    assert sorted([ws["D8"].value, ws["D9"].value]) == [1.0, 3.0]


def test_excel_workbook_includes_source_aware_fit_offset(tmp_path):
    model, X, y, w, term, offset = _fit_term_offset_export_model(distinct_terms=2)
    output = tmp_path / "tables_with_source_offset.xlsx"

    model.export_rating_tables(
        output,
        X,
        y,
        sample_weight=w,
        offset=offset,
        offset_source=term,
        offset_name="Term",
    )

    ws = load_workbook(output, data_only=True)["Rating Tables"]
    assert ws["D5"].value == "Term"
    assert ws["D7"].value == "Term"
    assert ws["E7"].value == "Relativity"
    assert sorted([ws["D8"].value, ws["D9"].value]) == [12.0, 36.0]
    rows = sorted(
        [(ws["D8"].value, ws["E8"].value), (ws["D9"].value, ws["E9"].value)],
        key=lambda row: row[0],
    )
    assert [row[0] for row in rows] == [12.0, 36.0]
    np.testing.assert_allclose([row[1] for row in rows], [1.0, 3.0])


def test_interactions_start_two_blank_rows_below_main_effects(tmp_path):
    rng = np.random.default_rng(321)
    n = 400
    X = pd.DataFrame(
        {
            "region": rng.choice(["A", "B", "C"], n),
            "type": rng.choice(["X", "Y"], n),
        }
    )
    y = rng.poisson(1.0 + 0.2 * (X["region"] == "B")).astype(float)
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={"region": Categorical(base="first"), "type": Categorical(base="first")},
        interactions=[("region", "type")],
    )
    model.fit(X, y)
    output = tmp_path / "interaction.xlsx"

    model.export_rating_tables(output, X, y)

    ws = load_workbook(output, data_only=True)["Rating Tables"]
    main_last_row = (
        8
        + max(len(block.table) for block in build_rating_table_payload(model, X, y).main_effects)
        - 1
    )
    interaction_title_row = main_last_row + 3
    assert ws.cell(row=interaction_title_row - 1, column=1).value is None
    assert ws.cell(row=interaction_title_row, column=1).value == "region:type"


def test_continuous_interaction_export_uses_selected_bins():
    rng = np.random.default_rng(456)
    n = 500
    X = pd.DataFrame(
        {
            "age": rng.uniform(18, 80, n),
            "vehicle_age": rng.uniform(0, 20, n),
        }
    )
    eta = (
        -1.0
        + 0.08 * np.sin(X["age"].to_numpy() / 9.0)
        + 0.05 * np.cos(X["vehicle_age"].to_numpy() / 4.0)
        + 0.03 * (X["age"].to_numpy() - 45.0) * (X["vehicle_age"].to_numpy() - 8.0) / 100.0
    )
    y = rng.poisson(np.exp(eta)).astype(float)
    model = SuperGLM(
        family="poisson",
        selection_penalty=0.0,
        features={
            "age": Spline(n_knots=7),
            "vehicle_age": Spline(n_knots=6),
        },
        interactions=[("age", "vehicle_age")],
    )
    model.fit(X, y)

    payload = build_rating_table_payload(model, X, y, n_bins=12)

    interaction = next(block for block in payload.interactions if block.name == "age:vehicle_age")
    assert interaction.table.shape == (12, 13)
    assert interaction.table.columns[0] == "age"
    assert np.isfinite(interaction.table.iloc[:, 1:].to_numpy(dtype=float)).all()


def test_export_rejects_unsupported_format(tmp_path):
    model, X, y, w = _fit_export_model()
    with pytest.raises(ValueError, match="Unsupported rating table export format"):
        model.export_rating_tables(tmp_path / "tables.csv", X, y, sample_weight=w)


def test_export_validates_lengths(tmp_path):
    model, X, y, w = _fit_export_model()
    with pytest.raises(ValueError, match="same length"):
        model.export_rating_tables(tmp_path / "tables.xlsx", X.iloc[:-1], y, sample_weight=w)


def test_summary_sheet_keeps_a_deselected_smooth_inactive():
    # False today. `active` asks whether ANY source group of the row survived
    # selection, and an OrderedCategorical specials term contributes two: the
    # penalized spline block and the special block, which is built
    # `penalized=False` and is therefore never deselected. So the workbook marks
    # the smooth row active even when the spline was dropped -- and it disagrees
    # with model.summary(), which reports the same row inactive and suppresses
    # its Wood statistics (coef_tables.py, fixed for the summary path only).
    #
    # The exported workbook is the artefact that leaves the building, so the two
    # must not say different things about whether a term is in the model.
    from tests.test_ordered_categorical_inference import (
        _fit_deselected_spline_with_live_special,
        _smooth_row,
    )

    model = _fit_deselected_spline_with_live_special()
    assert set(model.result.rank_info.selected_group_names) == {"band:special"}
    assert _smooth_row(model.summary()).active is False

    payload = build_summary_export_payload(model)
    smooth_row = next(
        row for row in payload.terms if row.group == "band" and row.kind == "smooth+free"
    )
    assert smooth_row.active is False

    # The free level is still fitted, so its own row must stay active: this is
    # about the smooth, not about suppressing the one estimate the term made.
    level_rows = [row for row in payload.terms if row.group == "band" and row.kind != "smooth+free"]
    assert any(row.active for row in level_rows)


def test_summary_sheet_keeps_dropped_smooth_level_rows_inactive():
    # The other half of the active-flag fix. Scoping the smooth GROUP row to the
    # spline block left the LEVEL rows still ORing over every source group, and
    # `_source_groups_for_row` returns both `band` and `band:special` for a row
    # like `band[L0]`. So when selection drops the spline, every smoothed level
    # is still exported active -- inheriting it from the free special, which is
    # the one part of the term that survived. Only a level actually fitted free
    # may take its activity from the special block.
    from tests.test_ordered_categorical_inference import (
        _fit_deselected_spline_with_live_special,
    )

    model = _fit_deselected_spline_with_live_special()
    assert set(model.result.rank_info.selected_group_names) == {"band:special"}

    payload = build_summary_export_payload(model)
    band = [row for row in payload.terms if row.group == "band"]
    free = [row for row in band if row.kind == "free level"]
    smooth_levels = [row for row in band if row.kind == "level"]

    assert free and smooth_levels, "fixture must produce both row kinds"
    assert all(row.active for row in free), "the special is still fitted"
    assert not any(row.active for row in smooth_levels), (
        "smoothed levels are no longer fitted once the spline block is dropped"
    )


def test_summary_sheet_level_activity_filter_fails_closed_on_a_renamed_field():
    # The test above pins the BEHAVIOUR; this pins that the behaviour cannot be
    # undone silently. The filter used to read
    # `getattr(group, "subgroup_type", None) != "special"`, which fails OPEN:
    # rename the field and `None != "special"` holds for every group, so the
    # filter becomes a no-op and every smoothed level row goes straight back to
    # inheriting activity from the special block -- the regression above,
    # restored by a rename, with the suite still green.
    #
    # Simulate the rename by hiding the attribute on the model's groups and
    # assert the export REFUSES rather than quietly reverting. The compact
    # summary is built first, off the untouched model, so only the level-row
    # activity filter sees the renamed groups.
    from superglm.export.summary import _adapt_compact_summary, _term_rows
    from tests.test_ordered_categorical_inference import (
        _fit_deselected_spline_with_live_special,
    )

    model = _fit_deselected_spline_with_live_special()
    source = _adapt_compact_summary(model.summary(detail="compact"))

    class _Renamed:
        """A GroupSlice whose `subgroup_type` has been renamed away."""

        def __init__(self, group):
            self._group = group
            self.block_type = group.subgroup_type

        def __getattr__(self, name):
            if name == "subgroup_type":
                raise AttributeError(name)
            return getattr(self._group, name)

    # Sanity: the untouched model does produce inactive smoothed level rows, so
    # a silent revert really would be visible here.
    baseline = _term_rows(model, source)
    assert not any(row.active for row in baseline if row.group == "band" and row.kind == "level")

    model._groups = [_Renamed(group) for group in model._groups]
    with pytest.raises(AttributeError, match="subgroup_type"):
        _term_rows(model, source)


# Ten significant digits, i.e. 5e-10 relative.  This bounds only what a HUMAN
# reads off the sheet: the stored value is exact and every reconstruction reads
# that, so a display format cannot move a premium.  It is the number that
# multiplies every row of the tariff, though, so a reader who cannot reconcile
# it to ten digits cannot check the workbook at all.  (The bin boundaries and
# axis values beside it are printed at full round-trip precision, because those
# are keys a consumer converts back to floats rather than numbers a reader
# reads -- see ``_format_number``.)
_REQUIRED_BASE_PRECISION = 5e-10

# Bases spanning the magnitudes a fitted intercept actually reaches.  A Poisson
# claim-frequency model sits around ``exp(-3)`` = 5%; ``exp(-14)`` is the
# near-1e-6 case ``excel.py`` cites as its reason for not using a fixed-decimal
# mask; ``exp(8)`` is a severity model's scale.  These are the magnitudes the
# format has to hold precision at, and testing one of them tests nothing.
_BASE_MAGNITUDES = [math.exp(k) for k in (8.0, 2.0, -1.0, -3.0, -5.0, -9.0, -14.0)]


@pytest.mark.parametrize("centering", ["native", "mean"])
def test_the_base_relativity_cell_is_rendered_at_full_precision(tmp_path, centering):
    """The one cell that multiplies every row must not render at two decimals.

    ``C2`` is column 3, so the global number-format loop's ``column % 3 == 0``
    arm claims it and formats it ``#,##0.00``.  The stored value stays exact,
    which is why no reader of the file object ever noticed: ``float(ws["C2"])``
    is right and every reconstruction test passes.  A human reading or
    copy-pasting the sheet sees the rounded number, and the base multiplies the
    entire tariff -- measured 0.3719954211385351 displaying as 0.37, a 5.4e-03
    relative error on every risk.  Under ``centering="mean"`` that cell also
    carries the whole transferred centering constant, so it is exactly the
    number a consumer would reconcile two exports with.

    The assertion is on what the format RESOLVES, not on how it is spelled, so
    the cell is rendered here the way Excel would and the result compared to
    the stored value.
    """
    model, X, y, w = _fit_export_model()
    output = tmp_path / f"base-{centering}.xlsx"
    model.export_rating_tables(output, X, y, sample_weight=w, n_bins=20, centering=centering)

    cell = load_workbook(output)["Rating Tables"]["C2"]
    stored = float(cell.value)
    assert stored > 0.0
    displayed = _as_excel_renders(stored, cell.number_format)
    assert displayed == pytest.approx(stored, rel=_REQUIRED_BASE_PRECISION), (
        f"base {stored!r} renders as {displayed!r} under number_format {cell.number_format!r}"
    )


@pytest.mark.parametrize("base", _BASE_MAGNITUDES)
def test_the_base_cell_holds_its_precision_at_every_magnitude_a_fit_reaches(tmp_path, base):
    """The precision has to survive the magnitude, not just the fixture's base.

    The fixture above happens to land near 0.37.  A format can hold ten
    significant digits there and lose them three decades down -- which is
    exactly what ``General`` does, because ECMA-376 Part 1 s18.8.30 budgets it
    in DISPLAY CHARACTERS ("max overall length for cell display is 11, not
    including negative sign, but includes leading zeros and decimal
    separator"), not in significant digits.  Leading zeros are charged to the
    same budget as digits, so the guarantee decays as the base shrinks: at
    ``exp(-3)`` = 0.049787068367863944 the eleven characters buy eight
    significant digits and a 7.4e-09 error, fifteen times the tolerance.

    So the base is swept over the magnitudes a fitted intercept reaches.  The
    payload is rebuilt with each one rather than fitted to it -- reaching
    ``exp(-14)`` by fitting would need a response nobody writes -- and it goes
    through the real renderer, so this pins the shipped format and not a copy
    of it.
    """
    model, X, y, w = _fit_export_model()
    payload = build_rating_table_payload(model, X, y, sample_weight=w, n_bins=20)
    output = tmp_path / f"base-{base:.3e}.xlsx"
    _write_workbook(replace(payload, base_relativity=base), output)

    cell = load_workbook(output)["Rating Tables"]["C2"]
    # The STORED value, which is what a reconstruction reads.  Not bit-exact:
    # the sheet serialises a float at sixteen significant digits where round-
    # tripping a float64 takes seventeen, so a value can come back one ulp
    # away.  Sixteen digits is 5e-16 relative at worst; 1e-15 is twice that,
    # and six orders of magnitude inside what this test is really about.
    assert float(cell.value) == pytest.approx(base, rel=1e-15)
    displayed = _as_excel_renders(base, cell.number_format)
    assert displayed is not None, (
        f"number_format {cell.number_format!r} leaves the rendering of {base!r} "
        "undetermined by ECMA-376, so no precision can be claimed for it"
    )
    assert displayed == pytest.approx(base, rel=_REQUIRED_BASE_PRECISION), (
        f"base {base!r} renders as {displayed!r} under number_format {cell.number_format!r}"
    )


# Bases whose ``str()`` is SHORT: exactly ``1.0``, a round ``0.5`` or ``100.0``,
# and a value near the smallest base the export admits, whose ``str`` is the
# six characters ``1e-300``.  ``_autosize`` measures ``str(cell.value)``, the
# base cell is in the column it sizes, and the floor is 12 -- so the column is
# too narrow precisely when the base's raw ``str`` is short and nothing longer
# shares the column.  A base like ``exp(-3)`` prints as twenty characters and
# pulls the column to 22 on its own, which is why the precision sweep above,
# which only ever used such bases, never saw this.
_SHORT_STR_BASES = [1.0, 0.5, 100.0, 1e-300]


@pytest.mark.parametrize("base", _SHORT_STR_BASES)
def test_the_base_relativity_cell_is_wide_enough_to_render(tmp_path, base):
    """A cell Excel renders as ``########`` is a cell nobody can check.

    The base carries eleven significant digits by number format, which is the
    whole point of that format; the rendering is sixteen characters wide, and
    seventeen at a three-digit exponent.  ``_autosize`` sizes a column from
    ``str(cell.value)`` -- the RAW float, not the format -- so the width it
    picks has nothing to do with what the cell will display (issue #290).

    Sized here with no main-effect block, which is both the shape the issue
    names -- an intercept-only export -- and the strongest form of the claim:
    the base cell must fit its own rendering on its own merits, not because a
    neighbouring ``Weight`` column happened to be long.  Measured against the
    unfixed renderer: 12 for every base here, against renderings of 16 and 17
    characters; and 20 for the same bases with the fixture's blocks present,
    which is why this is reachable rather than universal.

    Cosmetic and loud rather than silent: the stored value is exact and every
    reconstruction in this suite is unaffected.  It is worth pinning because
    this is the one cell the mean-centering fix exists to make trustworthy, and
    a reader who cannot see it cannot check it.

    Asserted against the width the FORMAT needs, derived from the format string
    rather than from the observed column width, so it stays true if either
    moves.
    """
    model, X, y, w = _fit_export_model()
    payload = build_rating_table_payload(model, X, y, sample_weight=w, n_bins=20)
    output = tmp_path / f"width-{base:.3e}.xlsx"
    _write_workbook(
        replace(payload, main_effects=[], interactions=[], base_relativity=base), output
    )

    ws = load_workbook(output)["Rating Tables"]
    cell = ws["C2"]
    assert cell.number_format == _BASE_RELATIVITY_NUMBER_FORMAT

    mantissa = _BASE_RELATIVITY_NUMBER_FORMAT.split("E")[0]
    decimals = len(mantissa.split(".")[-1]) if "." in mantissa else 0
    rendered = f"{float(cell.value):.{decimals}E}"
    assert len(rendered) >= 16, "the format really is wider than the autosize floor of 12"

    width = ws.column_dimensions[cell.column_letter].width
    assert width >= len(rendered), (
        f"base {base!r} renders as {rendered!r} ({len(rendered)} characters) in a column "
        f"{width} wide, so Excel shows ########"
    )


def test_every_block_key_column_is_wide_enough_for_the_keys_it_carries(tmp_path):
    """The same failure as the base cell, one column over, and quieter.

    ``_autosize`` caps at 36 characters. That cap is a readability limit for
    prose -- the summary sheet's wrapped ``Notes`` column depends on it -- and a
    block's KEY column is not prose. Once the bin boundaries are printed at
    round-trip precision an interval pair runs to 40 characters on this
    fixture, so the column is clipped.

    Excel clips rather than overflows, because the neighbouring ``Relativity``
    column is populated on every row. The reader therefore sees
    ``[20.463050119288255, 23.16627506`` -- which still reads as an interval,
    with a plausible and *wrong* right edge. That is worse than ``########``:
    hash marks announce themselves, a truncated number does not.

    Asserted for every block, not only the binned one, so a future block kind
    with a long key is covered; and against the content of the column rather
    than against 36 or 40, so it survives either number moving.
    """
    model, X, y, w = _fit_export_model()
    payload = build_rating_table_payload(model, X, y, sample_weight=w, n_bins=20)
    output = tmp_path / "keys.xlsx"
    _write_workbook(payload, output)
    ws = load_workbook(output)["Rating Tables"]

    binned = [block for block in payload.main_effects if block.kind == "continuous"]
    assert binned, "the fixture carries a block whose keys are interval strings"

    for idx, block in enumerate(payload.main_effects):
        letter = ws.cell(row=1, column=1 + idx * 3).column_letter
        longest = max(len(str(cell.value)) for cell in ws[letter] if cell.value is not None)
        width = ws.column_dimensions[letter].width
        assert width >= longest, (
            f"block {block.name!r} key column {letter} is {width} wide against a "
            f"{longest}-character cell, so Excel truncates the key"
        )

    # And the fixture really does exceed the prose cap, or this measures nothing.
    key_letter = ws.cell(row=1, column=1).column_letter
    assert max(len(str(c.value)) for c in ws[key_letter] if c.value is not None) > 36


def test_the_formats_this_cell_did_not_get_would_each_have_lost_the_precision():
    """The chosen format has to beat the alternatives, or the choice is arbitrary.

    Without this, ``_as_excel_renders`` could model anything at all and the
    test above would still pass on a fixture whose base happens to be benign.
    Each rejected candidate is required to miss the tolerance somewhere in the
    swept range, so the format actually shipped is the one that survives it.
    """

    def worst(number_format: str) -> float:
        errors = []
        for base in _BASE_MAGNITUDES:
            displayed = _as_excel_renders(base, number_format)
            if displayed is None:
                return math.inf  # not even determined -- strictly worse than wrong
            errors.append(abs(displayed - base) / base)
        return max(errors)

    # What the global ``column % 3 == 0`` arm would have left on C2.
    assert worst("#,##0.00") > _REQUIRED_BASE_PRECISION
    # The piecewise columns' fixed-decimal mask: fine for a log relativity,
    # which is bounded, and hopeless for an unbounded exp().
    assert worst(_PIECEWISE_NUMBER_FORMAT) > _REQUIRED_BASE_PRECISION
    # ``General``, whose eleven characters are spent on leading zeros.
    assert worst("General") > _REQUIRED_BASE_PRECISION

    # And the one that shipped clears it everywhere.
    assert worst(_BASE_RELATIVITY_NUMBER_FORMAT) <= _REQUIRED_BASE_PRECISION


# Excel's ``General`` is budgeted in display characters, not significant
# digits: ECMA-376 Part 1 s18.8.30, "Floating point rule: For general
# formatting in cells, max overall length for cell display is 11, not including
# negative sign, but includes leading zeros and decimal separator."
_GENERAL_DISPLAY_CHARACTERS = 11
# The same clause switches ``General`` to exponential once the decimal exponent
# drops below -3, and pins no digit count for that branch -- display is left
# "based on the available cell width".  Below this exponent ``General`` is
# therefore not modellable, which is the point: a precision guarantee cannot be
# built on it.
_GENERAL_SCIENTIFIC_BELOW_EXPONENT = -3


def _as_excel_renders(value: float, number_format: str) -> float | None:
    """The value a reader gets back off the sheet, for the formats used here.

    ``None`` means the format's rendering of this value is not determined by
    ECMA-376 -- which disqualifies it from carrying a precision claim just as
    surely as rendering it wrongly would.
    """
    if number_format == "General":
        return _as_general_renders(value)
    mantissa = number_format.split("E")[0]
    if "E+" in number_format or "E-" in number_format:
        # ``0.0000000000E+00``: the zeros left of the ``E`` are mantissa digit
        # placeholders, so the rendering is fixed-significand scientific.
        decimals = len(mantissa.split(".")[-1]) if "." in mantissa else 0
        return float(f"{value:.{decimals}E}")
    decimals = len(number_format.split(".")[-1]) if "." in number_format else 0
    return round(value, decimals)


def _as_general_renders(value: float) -> float | None:
    """``General``, per ECMA-376's eleven-CHARACTER budget.

    The budget is spent on the leading ``0``, the decimal separator and every
    leading zero before it reaches a significant digit, which is why the
    significant digits it delivers fall as the value does:

        0.36787944117144233 -> ``0.367879441``  9 sig digits, 4.7e-10
        0.049787068367863944 -> ``0.049787068``  8 sig digits, 7.4e-09
        0.006737946999085467 -> ``0.006737947``  7 sig digits, 1.4e-10
    """
    if value == 0.0:
        return 0.0
    exponent = math.floor(math.log10(abs(value)))
    if exponent < _GENERAL_SCIENTIFIC_BELOW_EXPONENT:
        return None
    integer_characters = max(exponent + 1, 1)
    decimals = max(_GENERAL_DISPLAY_CHARACTERS - integer_characters - 1, 0)
    return float(f"{value:.{decimals}f}")
