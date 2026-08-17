"""Excel renderer for rating-table export payloads."""

from __future__ import annotations

from os import PathLike
from pathlib import Path
from typing import BinaryIO

import pandas as pd

_TERM_COLUMNS = (
    ("Term", "term"),
    ("Group", "group"),
    ("Kind", "kind"),
    ("Estimate", "estimate"),
    ("Std Error", "std_error"),
    ("Statistic", "statistic"),
    ("Statistic Type", "statistic_type"),
    ("P Value", "p_value"),
    ("CI Lower", "ci_lower"),
    ("CI Upper", "ci_upper"),
    ("EDF", "edf"),
    ("Lambda", "smoothing_lambda"),
    ("Active", "active"),
    ("Significance", "significance"),
    ("Warning", "warning"),
)
_TERM_NUMERIC_COLUMNS = frozenset(
    {"Estimate", "Std Error", "Statistic", "P Value", "CI Lower", "CI Upper", "EDF", "Lambda"}
)

# The width of an ordinary main-effect block: a key column, a relativity and a
# weight.  It is no longer the STRIDE -- blocks are placed by their own widths
# (``_main_effect_start_columns``) because a ppform block is seven columns wide,
# and the number formats are chosen by a column's role in its block rather than
# by ``column % 3``.  What the constant still names is the mandatory
# three-column prefix every block carries: the downstream loader locates blocks
# by a header signature requiring ``relativity`` at +1 and ``weight`` at +2, so
# anything a block adds is added to the RIGHT of these three, and that is what
# makes "past the prefix" a usable test for the exact-form columns.
_MAIN_EFFECT_BLOCK_STRIDE = 3

# Blocks sit end to end with no blank column between them, which is what the
# fixed stride of 3 already did when every block was exactly three wide.  Named
# rather than left implicit because it is now a layout choice instead of an
# arithmetic accident: keeping it at 0 is what makes the all-three-wide default
# layout land on columns 1, 4, 7 exactly as it did before.
_MAIN_EFFECT_BLOCK_GAP = 0


def _main_effect_start_columns(blocks) -> list[int]:
    """The first column of each main-effect block, by cumulative width.

    A ppform block is seven columns and every other block is three, so "block
    ``i`` starts at ``1 + i * 3``" stopped being true the moment the first wide
    block could appear in the list.  Under the old rule a wide block's
    right-hand neighbour was written straight over its coefficients, and
    openpyxl raised nothing -- the sheet simply lost four columns and repeated a
    ``Relativity``/``Weight`` pair where they had been.

    Every caller that places or re-reaches a block goes through this, so the
    placement, the number formats, the piecewise notes and the key-column widths
    cannot come to disagree about where a block is.
    """
    starts: list[int] = []
    column = 1
    for block in blocks:
        starts.append(column)
        column += len(block.table.columns) + _MAIN_EFFECT_BLOCK_GAP
    return starts


_MAIN_EFFECT_TITLE_ROW = 5
_MAIN_EFFECT_NOTE_ROW = 6
_MAIN_EFFECT_HEADER_ROW = 7

# The piecewise block's two numeric columns are re-formatted after the per-block
# format pass.  ``Log relativity`` is not one of the roles that pass knows, and
# under the ``column % 3`` loop that preceded it the column rendered at two
# decimal places: the value stored in the cell stays exact either way, but a
# human reading or copy-pasting the sheet would see ``0.00``, which defeats the
# entire purpose of publishing that column.
_PIECEWISE_NUMBER_FORMAT = "0.000000000000"

# The ppform block's four exact-form columns.  Coefficients are not factors and
# must not be read at six decimal places: a reader copying ``b`` into a stored
# procedure needs every digit that was fitted, and the magnitude is not bounded
# the way a relativity's is -- ``a``, ``b``, ``c`` and ``d`` are signed
# log-space numbers whose scale follows the fitted curve rather than the [0, 1]
# local variable they multiply.  Significant digits, not decimal places, for the
# same reason the base relativity gets a scientific format.  (The interval
# bounds need no format at all: they live in the text key, not in columns of
# their own.)
_PPFORM_NUMBER_FORMAT = "0.00000000000000E+00"

# The two formats the main-effect blocks have always carried, keyed by the
# column's ROLE rather than by its position.  The loop these replace keyed on
# ``cell.column % _MAIN_EFFECT_BLOCK_STRIDE``, which means "the second and third
# column of a three-wide block" and means nothing at all once blocks differ in
# width: with one seven-column block on the sheet, ``% 3 == 2`` lands on the
# ``b`` coefficient and then on the NEXT block's interval key, and ``% 3 == 0``
# on ``c`` and then on that block's ``Relativity``.  Reading the header the
# block itself declared cannot drift that way.
_MAIN_EFFECT_ROLE_FORMATS = {
    "Relativity": "0.000000",
    "Weight": "#,##0.00",
}

# The base relativity cell, for the same reason and re-applied at the same
# point.  ``C2`` sits five rows above the first block, so the per-block format
# pass never reaches it -- and under the ``column % 3`` loop that preceded that
# pass it was worse than unformatted, because the loop walked the whole sheet
# and its ``== 0`` arm claimed column 3, rendering the base at two decimal
# places.  Either way the value stored keeps its sixteen significant digits, so
# no reader of the file object notices, but a human reading the sheet does.  It
# is the one cell that multiplies EVERY row of the tariff, and under
# ``centering="mean"`` it additionally carries the whole transferred centering
# constant, so a reader rating off the displayed number is uniformly wrong.
# Measured on a two-term Poisson fit: base
# 0.3719954211385351 displays as 0.37, a 5.4e-03 relative error on every risk;
# 0.4027922135365106 -> 0.40 (6.9e-03) in the mean-centered export of the same
# model.
#
# Scientific rather than a fixed-decimal format, which is where this differs
# from the piecewise columns.  Those hold log relativities, bounded by the
# tariff's own design, so twelve decimals is always enough.  The base is
# ``exp(intercept)`` and its magnitude is bounded by nothing: an ordinary
# claim-frequency Poisson fit sits near ``exp(-3)`` = 0.05 and a low-frequency
# one near 1e-6, where twelve decimals leaves seven significant digits.
#
# And scientific rather than ``General``, which was the first choice here and
# does not survive its own justification.  ECMA-376 Part 1 s18.8.30 budgets
# ``General`` in DISPLAY CHARACTERS, not significant digits -- "max overall
# length for cell display is 11, not including negative sign, but includes
# leading zeros and decimal separator" -- so leading zeros are charged against
# the same eleven, and the precision decays exactly as the base shrinks: at
# ``exp(-3)`` those eleven characters buy eight significant digits and a
# 7.4e-09 error.  Below a decimal exponent of -3 the same clause switches
# ``General`` to exponential and pins no digit count at all, leaving display
# "based on the available cell width", so there is no precision to claim.
#
# ``0.0000000000E+00`` is ten mantissa decimals plus the leading digit: eleven
# significant digits at EVERY magnitude, and well inside the fifteen Excel
# retains.  The base should not be the least precisely stated number on a sheet
# whose other public keys -- bin boundaries, interaction axis values -- are
# printed at full round-trip precision.
_BASE_RELATIVITY_CELL = "C2"
_BASE_RELATIVITY_NUMBER_FORMAT = "0.0000000000E+00"

# ...and a format Excel cannot fit renders as ``########``, which is the one
# cell on the sheet a reader most needs to be able to read.  ``_autosize``
# cannot see it: it measures ``str(cell.value)``, the raw float, and the base
# sits in a column whose other entries are a block's ``Weight`` values -- short
# whole numbers on an ordinary tariff, and absent entirely on an intercept-only
# export -- so the column lands on the floor of 12 while the rendering needs
# more.  Derived from the format rather than observed: one leading digit, the
# decimal separator, ten mantissa decimals, ``E``, the exponent sign, and up to
# three exponent digits (float64 reaches E+308) is 17 characters; the base is
# ``exp`` of a real number so it is never negative and never needs an
# eighteenth.  Plus the two columns of padding ``_autosize`` already applies.
_BASE_RELATIVITY_MIN_WIDTH = 17 + 2

# The same failure, one column over.  ``_autosize`` caps at 36 characters, which
# is a readability limit for PROSE -- the summary sheet's wrapped ``Notes``
# column depends on it -- and a block's KEY column is not prose.  Its cells are
# interval strings printed at round-trip precision, and a pair of them runs to
# 40 characters on an ordinary fixture.  Excel clips rather than overflows,
# because the neighbouring ``Relativity`` column is populated on every row, so
# the reader sees ``[20.463050119288255, 23.16627506`` -- which still reads as
# an interval, with a plausible and wrong right edge.  That is strictly worse
# than the ``########`` this module already refuses for the base: hash marks
# announce themselves, a truncated number does not.
#
# Derived, not observed: ``repr`` of a float64 is at most 24 characters
# (``-1.2345678901234567e-308``), so the widest interval this module can print
# is ``[`` + 24 + ``, `` + 24 + ``)`` = 52, plus the two columns of padding
# ``_autosize`` applies.
_KEY_COLUMN_MAX_WIDTH = 1 + 24 + 2 + 24 + 1 + 2


def _resolve_workbook_target(
    target: str | PathLike[str] | BinaryIO,
) -> Path | BinaryIO:
    if isinstance(target, str | PathLike):
        out = Path(target)
        out.parent.mkdir(parents=True, exist_ok=True)
        return out
    return target


def _add_excel_table(
    ws,
    *,
    display_name: str,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int,
) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    reference = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=display_name, ref=reference)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _write_summary_sheet(ws, summary) -> None:
    from openpyxl.styles import Alignment, Font

    ws["A1"] = "Model Summary"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Fit and model overview"
    ws["A3"].font = Font(bold=True)

    overview_header_row = 4
    overview_headers = ("Section", "Metric", "Value")
    for column, header in enumerate(overview_headers, start=1):
        ws.cell(row=overview_header_row, column=column, value=header).font = Font(bold=True)
    for row_number, overview_row in enumerate(summary.overview, start=overview_header_row + 1):
        ws.cell(row=row_number, column=1, value=overview_row.section)
        ws.cell(row=row_number, column=2, value=overview_row.metric)
        value_cell = ws.cell(row=row_number, column=3, value=overview_row.value)
        if isinstance(overview_row.value, int) and not isinstance(overview_row.value, bool):
            value_cell.number_format = "0"
        elif isinstance(overview_row.value, float):
            value_cell.number_format = "0.000000"

    overview_end_row = overview_header_row + len(summary.overview)
    _add_excel_table(
        ws,
        display_name="ModelOverview",
        min_row=overview_header_row,
        max_row=overview_end_row,
        min_col=1,
        max_col=len(overview_headers),
    )

    term_header_row = overview_end_row + 3
    for column, (header, _) in enumerate(_TERM_COLUMNS, start=1):
        ws.cell(row=term_header_row, column=column, value=header).font = Font(bold=True)

    term_rows = summary.terms or (None,)
    for row_number, term_row in enumerate(term_rows, start=term_header_row + 1):
        for column, (header, attribute) in enumerate(_TERM_COLUMNS, start=1):
            value = None if term_row is None else getattr(term_row, attribute)
            cell = ws.cell(row=row_number, column=column, value=value)
            if value is not None and header in _TERM_NUMERIC_COLUMNS:
                cell.number_format = "0.000000E+00" if header == "P Value" else "0.000000"

    term_end_row = term_header_row + len(term_rows)
    _add_excel_table(
        ws,
        display_name="TermInference",
        min_row=term_header_row,
        max_row=term_end_row,
        min_col=1,
        max_col=len(_TERM_COLUMNS),
    )

    notes_header_row = term_end_row + 3
    ws.cell(row=notes_header_row, column=1, value="Notes").font = Font(bold=True)
    for row_number, note in enumerate(summary.notes, start=notes_header_row + 1):
        cell = ws.cell(row=row_number, column=1, value=note)
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A5"


def _write_dataframe(ws, df: pd.DataFrame, start_row: int, start_col: int) -> tuple[int, int]:
    from openpyxl.styles import Font

    for c, column in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=c, value=column)
        cell.font = Font(bold=True)
    for r, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for c, value in enumerate(row, start=start_col):
            ws.cell(row=r, column=c, value=value)
    return start_row + len(df), start_col + len(df.columns) - 1


def _piecewise_interpolation_note(
    table: pd.DataFrame, extrapolation: str, centering_shift: float = 0.0
) -> str:
    """The interpolation and extrapolation rule for one piecewise block.

    Derived from the block's own printed columns, so a reader can check the
    stated slopes against the two end rows rather than take them on trust.

    The slopes are written at full round-trip precision on purpose.  The whole
    claim of a piecewise rating table is that the workbook reproduces the model
    exactly; a boundary slope rounded for readability would put a small,
    silent discrepancy back into the tariff at exactly the rows -- the ones
    outside the rated range -- where nobody would look for it.

    That exactness claim holds in both centerings.  ``centering="mean"``
    shifts this block by a constant so its relativities have geometric mean 1,
    and the exported base relativity carries the same constant back, so the
    base-times-blocks product is the same number either way.  The note states
    the constant it was shifted by, because a reader holding the two workbooks
    side by side can otherwise only see that the columns disagree.

    The out-of-range rule is the term's ``extrapolation`` parameter, and it is
    stated here rather than assumed because both directions exist in practice:
    holding flat beyond the outermost knots is the default (the library's
    splines hold flat too), extending the boundary segments is the stated
    alternative.  What ``lower``/``upper`` buy is that the knots sit where the
    tariff's rated range says they should.
    """
    knots = [float(value) for value in table.iloc[:, 0]]
    interpolate = (
        "Interpolate linearly on Log relativity (equivalently, geometrically on Relativity). "
    )
    # The exactness claim, stated in the sheet itself, with the centering
    # constant named so a reader can check the two exports against each other.
    # Printed at full round-trip precision for the same reason the boundary
    # slopes are: it is the number that reconciles this block with the base.
    scope = (
        " Exact reproduction of the fitted model holds in either centering."
        if centering_shift == 0.0
        else (
            " Exact reproduction of the fitted model holds in either centering: "
            f"these Log relativity values are the fitted ones less {centering_shift}, "
            "and the base relativity carries that constant back."
        )
    )
    if extrapolation == "extend":
        log_relativity = [float(value) for value in table["Log relativity"]]
        slope_low = (log_relativity[1] - log_relativity[0]) / (knots[1] - knots[0])
        slope_high = (log_relativity[-1] - log_relativity[-2]) / (knots[-1] - knots[-2])
        return (
            interpolate
            + (
                "Beyond the tabulated range the boundary segments continue: "
                f"below {knots[0]} use slope {slope_low}; "
                f"above {knots[-1]} use slope {slope_high}."
            )
            + scope
        )
    if extrapolation == "error":
        return (
            interpolate
            + (
                f"Values outside [{knots[0]}, {knots[-1]}] are not rated: the model "
                "refuses them (extrapolation='error')."
            )
            + scope
        )
    return (
        interpolate
        + (
            "Beyond the tabulated range hold the end rows flat: "
            f"below {knots[0]} use the {knots[0]} row; "
            f"above {knots[-1]} use the {knots[-1]} row."
        )
        + scope
    )


def _main_effect_number_format(block, column_name: str, offset: int) -> str | None:
    """The format one column of one block should carry, or ``None`` for plain.

    ``Relativity`` and ``Weight`` are recognised by name wherever they appear,
    so the lookup half of a ppform block renders exactly like the binned block
    it stands in for.  Everything past that block's mandatory three-column
    prefix is its exact form, and gets significant digits rather than decimal
    places -- a coefficient rendered as ``0.00`` is stored exactly and read
    wrongly, which is the failure this whole export exists to remove.

    Keyed on the block's ``kind`` rather than on the four column names, so a
    coefficient column renamed in ``rating_tables`` cannot quietly fall back to
    the default rendering.
    """
    role = _MAIN_EFFECT_ROLE_FORMATS.get(column_name)
    if role is not None:
        return role
    if block.kind == "continuous_ppform" and offset >= _MAIN_EFFECT_BLOCK_STRIDE:
        return _PPFORM_NUMBER_FORMAT
    return None


def _format_main_effect_blocks(ws, main_effects) -> None:
    """Apply each block's number formats within its own columns.

    Per block rather than over the whole sheet, because a sheet-wide rule can
    only be expressed in column arithmetic and column arithmetic stopped being
    able to say which column is which the moment blocks could differ in width.
    """
    for block, start_col in zip(
        main_effects, _main_effect_start_columns(main_effects), strict=True
    ):
        first_row = _MAIN_EFFECT_HEADER_ROW + 1
        for offset, column_name in enumerate(block.table.columns):
            number_format = _main_effect_number_format(block, str(column_name), offset)
            if number_format is None:
                continue
            for row in range(first_row, first_row + len(block.table)):
                ws.cell(row=row, column=start_col + offset).number_format = number_format


def _annotate_piecewise_blocks(ws, main_effects) -> None:
    """Re-format the piecewise cells and write each block's interpolation note.

    Runs *after* the per-block format pass and touches nothing else: the block
    placement stays exactly as it was, so every other block keeps its
    coordinates and its formats.
    """
    starts = _main_effect_start_columns(main_effects)
    for idx, block in enumerate(main_effects):
        if block.kind != "piecewise":
            continue
        start_col = starts[idx]
        first_row = _MAIN_EFFECT_HEADER_ROW + 1
        for row in range(first_row, first_row + len(block.table)):
            for offset in (1, 2):
                cell = ws.cell(row=row, column=start_col + offset)
                cell.number_format = _PIECEWISE_NUMBER_FORMAT
        # Row 6 sits between the block title and the dataframe header and is
        # otherwise unused, so the note needs no layout change to land in.
        ws.cell(
            row=_MAIN_EFFECT_NOTE_ROW,
            column=start_col,
            value=_piecewise_interpolation_note(
                block.table,
                block.extrapolation or "clip",
                block.centering_shift,
            ),
        )


def _autosize(ws) -> None:
    from openpyxl.utils import get_column_letter

    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells
        )
        ws.column_dimensions[letter].width = min(max(max_length + 2, 12), 36)


def _widen_block_key_columns(ws, main_effects) -> None:
    """Let each block's key column fit its own keys, past the prose cap.

    Only the FIRST column of each main-effect block, because that is the one
    holding the keys a reader looks a risk up by; the ``Relativity`` and
    ``Weight`` columns beside it are formatted numbers and are already short.
    The cap stays, at the width the printer can actually produce, so the rule
    is still a bound rather than "as wide as it takes".
    """
    from openpyxl.utils import get_column_letter

    for start_col in _main_effect_start_columns(main_effects):
        letter = get_column_letter(start_col)
        longest = max(
            (len(str(cell.value)) for cell in ws[letter] if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[letter].width = min(
            max(ws.column_dimensions[letter].width or 0, longest + 2),
            _KEY_COLUMN_MAX_WIDTH,
        )


def write_rating_table_workbook(
    payload,
    target: str | PathLike[str] | BinaryIO,
    *,
    sheet_name: str,
    summary_sheet_name: str,
    impact_sheet_name: str,
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    out = _resolve_workbook_target(target)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.freeze_panes = "A8"

    ws["A2"] = "Base"
    ws["A2"].font = Font(bold=True)
    ws[_BASE_RELATIVITY_CELL] = float(payload.base_relativity)

    max_main_row = 8
    for block, start_col in zip(
        payload.main_effects,
        _main_effect_start_columns(payload.main_effects),
        strict=True,
    ):
        title_cell = ws.cell(row=_MAIN_EFFECT_TITLE_ROW, column=start_col, value=block.name)
        title_cell.font = Font(bold=True)
        end_row, _ = _write_dataframe(ws, block.table, _MAIN_EFFECT_HEADER_ROW, start_col)
        max_main_row = max(max_main_row, end_row)

    _format_main_effect_blocks(ws, payload.main_effects)

    ws[_BASE_RELATIVITY_CELL].number_format = _BASE_RELATIVITY_NUMBER_FORMAT
    _annotate_piecewise_blocks(ws, payload.main_effects)

    interaction_row = max_main_row + 3
    for block in payload.interactions:
        title_cell = ws.cell(row=interaction_row, column=1, value=block.name)
        title_cell.font = Font(bold=True)
        end_row, _ = _write_dataframe(ws, block.table, interaction_row + 2, 1)
        for row in ws.iter_rows(
            min_row=interaction_row + 3,
            max_row=end_row,
            min_col=2,
            max_col=block.table.shape[1],
        ):
            for cell in row:
                cell.number_format = "0.000000"
        interaction_row = end_row + 3

    impact_ws = wb.create_sheet(impact_sheet_name)
    _write_dataframe(impact_ws, payload.discretization_impact, 1, 1)

    summary_ws = wb.create_sheet(summary_sheet_name)
    _write_summary_sheet(summary_ws, payload.summary)

    for sheet in wb.worksheets:
        _autosize(sheet)

    # After ``_autosize``, and floors rather than assignments, so a column that
    # is already wider keeps its width.
    base_column = ws[_BASE_RELATIVITY_CELL].column_letter
    ws.column_dimensions[base_column].width = max(
        ws.column_dimensions[base_column].width,
        _BASE_RELATIVITY_MIN_WIDTH,
    )
    _widen_block_key_columns(ws, payload.main_effects)

    summary_ws.column_dimensions["A"].width = 32
    summary_ws.column_dimensions["B"].width = 24
    for column in "CDEFGHIJKLMNO":
        summary_ws.column_dimensions[column].width = min(
            max(summary_ws.column_dimensions[column].width, 14),
            18,
        )
    wb.save(out)
