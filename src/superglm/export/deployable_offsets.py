"""Rating-workbook integration for deployable fixed offsets."""

from __future__ import annotations

from pathlib import Path

from superglm.offsets import fixed_offset_frame

from .rating_tables import export_rating_tables as _export_rating_tables

FIXED_OFFSET_SHEET_NAME = "Fixed Offsets"


def _append_fixed_offset_sheet(
    model,
    workbook_path: str | Path,
    *,
    sheet_name: str = FIXED_OFFSET_SHEET_NAME,
) -> None:
    frame = fixed_offset_frame(model)
    if frame.empty:
        return

    from openpyxl import load_workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    workbook = load_workbook(workbook_path)
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    worksheet = workbook.create_sheet(sheet_name, 1)
    worksheet.freeze_panes = "A2"

    for column_index, column_name in enumerate(frame.columns, start=1):
        cell = worksheet.cell(row=1, column=column_index, value=column_name)
        cell.font = Font(bold=True)

    for row_index, row in enumerate(frame.itertuples(index=False), start=2):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row=row_index, column=column_index, value=value)

    numeric_formats = {
        "Reference Value": "0.###############",
        "Coefficient": "0.###############",
    }
    for column_name, number_format in numeric_formats.items():
        column_index = frame.columns.get_loc(column_name) + 1
        for row_index in range(2, len(frame) + 2):
            worksheet.cell(row=row_index, column=column_index).number_format = number_format

    for column_index, column_name in enumerate(frame.columns, start=1):
        values = [column_name, *frame[column_name].astype(str).tolist()]
        width = min(max(max(len(value) for value in values) + 2, 12), 60)
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    workbook.save(workbook_path)


def export_rating_tables(
    model,
    file_path,
    X,
    y,
    sample_weight=None,
    *,
    fixed_offset_sheet_name: str = FIXED_OFFSET_SHEET_NAME,
    **kwargs,
):
    """Export rating tables and any structured fixed offsets to one workbook."""
    output = _export_rating_tables(
        model,
        file_path,
        X,
        y,
        sample_weight=sample_weight,
        **kwargs,
    )
    _append_fixed_offset_sheet(
        model,
        output,
        sheet_name=fixed_offset_sheet_name,
    )
    return output


__all__ = ["FIXED_OFFSET_SHEET_NAME", "export_rating_tables"]
